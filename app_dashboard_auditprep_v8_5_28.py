# ============================================================
# AuditPrep IA - V8.5.28 — correctifs Cloud ML + contraste Light/Dark
# Base : V8.4 sécurisée + interface portail métier Convergence
# Fichier à placer dans : C:\PFE Omar\app\app_dashboard_auditprep_v8_5_28.py
#
# Lancement PowerShell :
#   cd "C:\PFE Omar"
#   conda run -n auditprep python -m streamlit run "app\app_dashboard_auditprep_v8_5_28.py"
# ============================================================

import base64
import hashlib
import hmac
import html
import io
import math
import os
import re
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import psycopg2
from psycopg2 import sql
import streamlit as st

# Version V8.3 : base V8.2 + correction des anciens modes de génération
# manquants et synthèse IA plus lisible.


# ============================================================
# 1. CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="AuditPrep IA | Convergence",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ============================================================
# 2. CSS V6 - correction lisibilité dark/light + largeur
# ============================================================

st.markdown(
    """
<style>
/* ============================================================
   AuditPrep IA - CSS V6 sécurisé Light/Dark
   Objectif : rendu propre en mode clair ET sombre sans cartes grisées
   ============================================================ */

:root {
    --audit-red: #ff4b4b;
    --audit-red-dark: #d93636;
    --audit-border-soft: rgba(148, 163, 184, 0.34);
    --audit-border-strong: rgba(148, 163, 184, 0.52);
    --audit-shadow: 0 14px 34px rgba(15, 23, 42, 0.08);
}

/* Zone principale */
.main .block-container {
    max-width: 1340px;
    padding-top: 2rem;
    padding-bottom: 3rem;
}

/* Ne pas forcer une couleur globale : on respecte Light/Dark Streamlit */
[data-testid="stAppViewContainer"] {
    color: inherit;
}

/* Sidebar */
[data-testid="stSidebar"] {
    border-right: 1px solid var(--audit-border-soft);
}

[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span {
    color: inherit !important;
}

/* Cartes : compatibles avec le thème actif */
.audit-hero,
.audit-card,
.audit-soft-card,
.audit-kpi-card,
.audit-info-box,
.audit-warning-box,
.audit-success-box,
.audit-alert-card {
    color: inherit !important;
    border-radius: 18px;
    border: 1px solid var(--audit-border-soft);
    padding: 1.15rem 1.35rem;
    margin-bottom: 1rem;
    box-shadow: var(--audit-shadow);
}

.audit-hero,
.audit-card,
.audit-soft-card,
.audit-kpi-card,
.audit-alert-card {
    background: rgba(148, 163, 184, 0.08);
}

.audit-hero {
    background: linear-gradient(135deg, rgba(148, 163, 184, 0.18), rgba(148, 163, 184, 0.06));
}

.audit-card h3,
.audit-soft-card h4,
.audit-kpi-card,
.audit-alert-card,
.audit-info-box,
.audit-warning-box,
.audit-success-box,
.audit-card p,
.audit-hero div {
    color: inherit !important;
}

.audit-title {
    font-size: clamp(1.45rem, 2.4vw, 2.2rem);
    font-weight: 850;
    line-height: 1.15;
    margin-bottom: 0.45rem;
}

.audit-subtitle {
    font-size: 1rem;
    opacity: 0.82;
}

.audit-section-title {
    font-size: 1.55rem;
    font-weight: 800;
    margin-top: 1.2rem;
    margin-bottom: 0.4rem;
}

.audit-section-subtitle {
    opacity: 0.78;
    margin-bottom: 1rem;
}

.audit-mini-label {
    font-size: 0.82rem;
    font-weight: 800;
    opacity: 0.72;
    text-transform: uppercase;
    letter-spacing: 0.04em;
}

.audit-big-number {
    font-size: clamp(1.7rem, 2.5vw, 2.35rem);
    font-weight: 780;
    margin-top: 0.2rem;
}

.audit-pill {
    display: inline-block;
    padding: 0.22rem 0.58rem;
    border-radius: 999px;
    font-size: 0.78rem;
    font-weight: 850;
    margin-left: 0.45rem;
    white-space: nowrap;
}

.audit-pill-high {
    background: rgba(248, 113, 113, 0.18);
    color: #ef4444 !important;
}

.audit-pill-mid {
    background: rgba(245, 158, 11, 0.18);
    color: #b45309 !important;
}

.audit-pill-low {
    background: rgba(34, 197, 94, 0.14);
    color: #15803d !important;
}

.audit-alert-card {
    border-left: 7px solid rgba(100, 116, 139, 0.9);
}

.audit-alert-title {
    font-weight: 850;
    margin-bottom: 0.3rem;
}

.audit-success-box {
    background: rgba(34, 197, 94, 0.12);
    border-color: rgba(34, 197, 94, 0.30);
}

.audit-warning-box {
    background: rgba(245, 158, 11, 0.14);
    border-color: rgba(245, 158, 11, 0.34);
}

.audit-info-box {
    background: rgba(59, 130, 246, 0.11);
    border-color: rgba(59, 130, 246, 0.27);
}

/* Boutons */
.stButton > button {
    border-radius: 12px !important;
    font-weight: 850 !important;
    min-height: 2.75rem;
}

.stButton > button[kind="primary"] {
    background-color: var(--audit-red) !important;
    border-color: var(--audit-red) !important;
    color: white !important;
}

/* Écran de connexion entreprise */
.audit-login-shell {
    max-width: 520px;
    margin: 7vh auto 1rem auto;
    padding: 2rem 2.2rem;
    border: 1px solid var(--audit-border-soft);
    border-radius: 24px;
    background: linear-gradient(145deg, rgba(15, 23, 42, 0.04), rgba(148, 163, 184, 0.10));
    box-shadow: 0 24px 60px rgba(15, 23, 42, 0.13);
    text-align: center;
}

.audit-login-mark {
    width: 58px;
    height: 58px;
    margin: 0 auto 1rem auto;
    display: grid;
    place-items: center;
    border-radius: 18px;
    background: linear-gradient(135deg, #ef4444, #b91c1c);
    color: white;
    font-size: 1.65rem;
    box-shadow: 0 12px 30px rgba(239, 68, 68, 0.26);
}

.audit-login-title {
    font-size: 1.75rem;
    font-weight: 850;
    margin-bottom: 0.35rem;
}

.audit-login-company {
    font-size: 0.86rem;
    font-weight: 800;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    opacity: 0.64;
}

.audit-user-chip {
    padding: 0.75rem 0.9rem;
    border: 1px solid var(--audit-border-soft);
    border-radius: 14px;
    background: rgba(148, 163, 184, 0.08);
    margin-bottom: 0.9rem;
}

/* L'authentification est gérée dans l'application : les contrôles Streamlit
   de déploiement ne font pas partie du parcours métier local. */
[data-testid="stStatusWidget"] {
    visibility: hidden;
}

.stButton > button[kind="primary"]:hover {
    background-color: var(--audit-red-dark) !important;
    border-color: var(--audit-red-dark) !important;
}

/* Inputs */
.stTextInput input,
.stNumberInput input,
.stPasswordInput input,
.stSelectbox div[data-baseweb="select"] > div {
    border-radius: 12px !important;
}

/* Dataframes */
[data-testid="stDataFrame"] {
    border-radius: 14px;
    overflow: hidden;
    border: 1px solid var(--audit-border-soft);
}

/* Tabs */
button[data-baseweb="tab"] {
    font-weight: 780;
}

button[data-baseweb="tab"][aria-selected="true"] {
    color: var(--audit-red) !important;
    border-bottom-color: var(--audit-red) !important;
}

/* Expander */
.streamlit-expanderHeader {
    font-weight: 780 !important;
}

/* Pas de @media prefers-color-scheme : on laisse Streamlit gérer Light/Dark correctement. */

/* Grilles et cartes utilisées par le dashboard */
.audit-grid-3 {
    display: grid;
    grid-template-columns: repeat(3, minmax(0, 1fr));
    gap: 1rem;
    margin: 1rem 0 1.4rem 0;
}
.audit-grid-4 {
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 1rem;
    margin: 1rem 0 1.4rem 0;
}
.audit-mini-card {
    color: inherit !important;
    background: rgba(148, 163, 184, 0.08);
    border: 1px solid var(--audit-border-soft);
    border-radius: 18px;
    padding: 1.15rem 1.35rem;
    box-shadow: var(--audit-shadow);
    min-height: 130px;
}
.audit-mini-card h3 { margin: 0.35rem 0 0.55rem 0; }
.audit-mini-card p { margin: 0; opacity: 0.86; }
.audit-kpi-label {
    font-size: 0.78rem;
    font-weight: 850;
    opacity: 0.74;
    text-transform: uppercase;
    letter-spacing: 0.04em;
}
.audit-kpi-value {
    font-size: 2.1rem;
    font-weight: 850;
    line-height: 1.15;
    margin-top: 0.35rem;
}
.audit-step-bubble {
    width: 2.25rem;
    height: 2.25rem;
    border-radius: 999px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    font-weight: 900;
    background: rgba(248, 113, 113, 0.18);
    color: var(--audit-red);
    margin-bottom: 0.6rem;
}
.audit-badge {
    display: inline-block;
    padding: 0.22rem 0.58rem;
    border-radius: 999px;
    font-size: 0.78rem;
    font-weight: 850;
    margin-left: 0.35rem;
}
.badge-high { background: rgba(248, 113, 113, 0.18); color: #ef4444 !important; }
.badge-medium { background: rgba(245, 158, 11, 0.18); color: #b45309 !important; }
.badge-low { background: rgba(34, 197, 94, 0.14); color: #15803d !important; }
.audit-warning,
.audit-success,
.audit-note {
    border-radius: 14px;
    padding: 0.95rem 1.1rem;
    margin: 0.8rem 0 1rem 0;
    border: 1px solid var(--audit-border-soft);
}
.audit-warning { background: rgba(245, 158, 11, 0.14); border-color: rgba(245, 158, 11, 0.34); }
.audit-success { background: rgba(34, 197, 94, 0.12); border-color: rgba(34, 197, 94, 0.30); }
.audit-note { background: rgba(59, 130, 246, 0.10); border-color: rgba(59, 130, 246, 0.24); }
.audit-separator {
    border: none;
    border-top: 1px solid var(--audit-border-soft);
    margin: 1.8rem 0;
}

/* ============================================================
   V8.5 - identité Convergence inspirée d'un portail métier
   ============================================================ */
:root {
    --audit-navy: #102a68;
    --audit-navy-dark: #081b49;
    --audit-orange: #ff5a1f;
    --audit-orange-dark: #e94a12;
    --audit-page: #f5f6f8;
}

[data-testid="stAppViewContainer"] {
    background: var(--audit-page);
}

.main .block-container {
    max-width: 1180px;
    padding-top: 1.25rem;
}

.audit-topbar {
    min-height: 64px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 1rem;
    padding: 0.72rem 1.25rem;
    margin: -1.25rem 0 1.6rem 0;
    border-radius: 0 0 16px 16px;
    background: linear-gradient(135deg, var(--audit-navy-dark), var(--audit-navy));
    color: white !important;
    box-shadow: 0 12px 30px rgba(8, 27, 73, 0.18);
}

.audit-brand {
    display: flex;
    align-items: center;
    gap: 0.72rem;
    font-weight: 850;
}

.audit-brand-mark {
    width: 38px;
    height: 38px;
    display: inline-grid;
    place-items: center;
    border-radius: 10px;
    background: var(--audit-orange);
    color: white !important;
    font-size: 0.84rem;
    letter-spacing: 0.03em;
}

.audit-brand-name {
    padding: 0.35rem 0.65rem;
    border-radius: 7px;
    background: var(--audit-orange);
    color: white !important;
    font-size: 0.88rem;
    letter-spacing: 0.02em;
}

.audit-nav-items {
    display: flex;
    align-items: center;
    gap: 1.15rem;
    font-size: 0.86rem;
    color: rgba(255, 255, 255, 0.76) !important;
}

.audit-nav-active {
    color: var(--audit-orange) !important;
    font-weight: 850;
}

.audit-hero {
    background: linear-gradient(135deg, var(--audit-navy-dark), var(--audit-navy)) !important;
    border: none !important;
    color: white !important;
    text-align: center;
    padding: 3rem 2.2rem !important;
    box-shadow: 0 18px 38px rgba(8, 27, 73, 0.20) !important;
}

.audit-hero *,
.audit-hero div,
.audit-hero p {
    color: white !important;
}

.audit-hero-title {
    font-size: clamp(1.8rem, 3vw, 2.65rem) !important;
    font-weight: 900 !important;
    letter-spacing: -0.02em;
}

.audit-hero-subtitle {
    max-width: 760px;
    margin: 0.75rem auto 0 auto;
    color: rgba(255,255,255,0.78) !important;
}

.audit-mini-card,
.audit-card,
.audit-soft-card,
.audit-kpi-card,
.audit-alert-card {
    background: #ffffff !important;
    border-color: rgba(15, 42, 104, 0.10) !important;
    box-shadow: 0 10px 26px rgba(15, 42, 104, 0.08) !important;
}

.audit-grid-3 .audit-mini-card {
    text-align: center;
}

.audit-kpi-value,
.audit-big-number {
    color: var(--audit-orange) !important;
}

.audit-step-bubble {
    background: rgba(255, 90, 31, 0.12) !important;
    color: var(--audit-orange) !important;
}

.stButton > button[kind="primary"] {
    background: var(--audit-orange) !important;
    border-color: var(--audit-orange) !important;
    color: white !important;
}

.stButton > button[kind="primary"]:hover {
    background: var(--audit-orange-dark) !important;
    border-color: var(--audit-orange-dark) !important;
}

.audit-login-shell {
    background: linear-gradient(135deg, var(--audit-navy-dark), var(--audit-navy)) !important;
    border: none !important;
    color: white !important;
    box-shadow: 0 22px 52px rgba(8, 27, 73, 0.24) !important;
}

.audit-login-shell *,
.audit-login-shell div {
    color: white !important;
}

.audit-login-mark {
    background: var(--audit-orange) !important;
    box-shadow: 0 12px 30px rgba(255, 90, 31, 0.28) !important;
}

.audit-info-box,
.audit-note {
    border-left: 4px solid var(--audit-navy) !important;
}

[data-testid="stSidebar"] {
    background: #ffffff;
    border-right: 1px solid rgba(15, 42, 104, 0.10);
}

[data-testid="stDataFrame"] {
    background: white;
}

/* Responsive */
@media (max-width: 900px) {
    .main .block-container {
        padding-left: 1rem;
        padding-right: 1rem;
    }
    .audit-card,
    .audit-soft-card,
    .audit-kpi-card,
    .audit-info-box,
    .audit-warning-box,
    .audit-success-box,
    .audit-alert-card {
        padding: 0.95rem 1rem;
    }
    .audit-nav-items {
        display: none;
    }
    .audit-topbar {
        margin-top: -1rem;
    }
}
</style>
""",
    unsafe_allow_html=True,
)

# ============================================================
# 2B. THÈME VISUEL FINAL INTÉGRÉ — V8.5.27 / INTER
#     Modification purement visuelle : logique métier inchangée
# ============================================================

st.markdown(
    r"""
<style>
/* =========================================================
   AUDITPREP — UI THEME
   Palette: Violet #7D4FFE + Lavender #C49FFF + Soft Pink #FFD0E6
   Typography: Inter
   ========================================================= */

@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

:root {
    --ap-lime: #7D4FFE;          /* Primary violet */
    --ap-lime-dark: #6036D6;
    --ap-lime-deep: #4422A8;
    --ap-lime-soft: #DCCFFF;
    --ap-lime-pale: #F0EAFF;

    --ap-lime-grey: #C49FFF;
    --ap-lime-grey-dark: #956FD0;
    --ap-lime-grey-soft: #E9DCFF;

    --ap-blue-grey: #FFD0E6;
    --ap-blue-grey-dark: #B75F89;
    --ap-blue-grey-deep: #5A3147;
    --ap-blue-grey-soft: #FFE0EE;
    --ap-blue-grey-pale: #FFF0F7;

    --ap-bg: #EEE9F8;
    --ap-surface: #F8F4FF;
    --ap-surface-alt: #F2ECFA;
    --ap-text: #30243F;
    --ap-muted: #716580;
    --ap-border: #D9CCE8;
    --ap-shadow: 0 10px 30px rgba(76, 48, 112, 0.10);
    --ap-shadow-hover: 0 14px 34px rgba(76, 48, 112, 0.17);
    --ap-radius: 16px;
}

/* ---------- Global ---------- */
html, body, [class*="css"], [data-testid="stAppViewContainer"] {
    font-family: 'Inter', system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
}

[data-testid="stAppViewContainer"] {
    background:
        radial-gradient(circle at 8% 2%, rgba(125, 79, 254, 0.12), transparent 25%),
        radial-gradient(circle at 96% 8%, rgba(255, 208, 230, 0.32), transparent 28%),
        var(--ap-bg);
    color: var(--ap-text);
}

[data-testid="stAppViewContainer"] > .main {
    background: transparent;
}

.main .block-container {
    max-width: 1420px;
    margin: 0 auto;
    padding-top: 2.2rem;
    padding-bottom: 3.5rem;
    padding-left: clamp(1rem, 3vw, 3rem);
    padding-right: clamp(1rem, 3vw, 3rem);
}

/* Smooth arrival without making the interface distracting */
.main .block-container > div {
    animation: apFadeUp 0.38s ease both;
}

@keyframes apFadeUp {
    from { opacity: 0; transform: translateY(5px); }
    to   { opacity: 1; transform: translateY(0); }
}

/* ---------- Typography hierarchy ---------- */
h1, h2, h3, h4, h5, h6 {
    font-family: 'Inter', system-ui, sans-serif !important;
    letter-spacing: -0.025em;
    color: var(--ap-blue-grey-deep);
    text-wrap: balance;
}

h1 {
    font-size: clamp(2rem, 3vw, 2.9rem) !important;
    line-height: 1.08 !important;
    font-weight: 800 !important;
    text-align: center;
    margin: 0.25rem auto 1.65rem !important;
    max-width: 1050px;
    position: relative;
}

h1::after {
    content: "";
    display: block;
    width: 78px;
    height: 5px;
    margin: 0.8rem auto 0;
    border-radius: 999px;
    background: linear-gradient(90deg, var(--ap-lime), var(--ap-blue-grey));
}

h2 {
    font-size: clamp(1.45rem, 2vw, 1.95rem) !important;
    line-height: 1.2 !important;
    font-weight: 750 !important;
    margin-top: 2.2rem !important;
    margin-bottom: 1rem !important;
    padding-left: 0.85rem;
    border-left: 5px solid var(--ap-lime);
}

h3 {
    font-size: clamp(1.08rem, 1.45vw, 1.3rem) !important;
    line-height: 1.35 !important;
    font-weight: 700 !important;
    color: var(--ap-blue-grey-dark) !important;
    margin-top: 1.5rem !important;
    margin-bottom: 0.75rem !important;
}

p, li, label, .stMarkdown, [data-testid="stText"] {
    color: var(--ap-text);
    line-height: 1.65;
}

small, .caption, [data-testid="stCaptionContainer"] {
    color: var(--ap-muted) !important;
}

/* ---------- Sidebar ---------- */
[data-testid="stSidebar"] {
    background:
        linear-gradient(180deg, rgba(196, 214, 0, 0.08), rgba(127, 149, 163, 0.08)),
        #F8F4FF;
    border-right: 1px solid var(--ap-border);
}

[data-testid="stSidebar"] .block-container {
    padding-top: 1.7rem;
}

[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
    text-align: left !important;
}

[data-testid="stSidebar"] h1::after {
    margin-left: 0;
}

/* ---------- Generic surfaces / cards ---------- */
[data-testid="stVerticalBlockBorderWrapper"],
[data-testid="stForm"],
[data-testid="stExpander"] {
    border-radius: var(--ap-radius) !important;
}

[data-testid="stForm"] {
    background: rgba(255,255,255,0.92);
    border: 1px solid var(--ap-border) !important;
    box-shadow: var(--ap-shadow);
    padding: 1.15rem;
}

[data-testid="stExpander"] details {
    border: 1px solid var(--ap-border) !important;
    border-radius: 14px !important;
    overflow: hidden;
    background: rgba(255,255,255,0.85);
}

[data-testid="stExpander"] summary {
    font-weight: 650;
    color: var(--ap-blue-grey-deep);
}

/* ---------- Metrics ---------- */
[data-testid="stMetric"] {
    background: linear-gradient(145deg, #FFFFFF, var(--ap-blue-grey-pale));
    border: 1px solid var(--ap-border);
    border-top: 4px solid var(--ap-lime);
    border-radius: var(--ap-radius);
    padding: 1.05rem 1.15rem;
    box-shadow: var(--ap-shadow);
    min-height: 118px;
    transition: transform .22s ease, box-shadow .22s ease, border-color .22s ease;
}

[data-testid="stMetric"]:hover {
    transform: translateY(-3px);
    box-shadow: var(--ap-shadow-hover);
    border-color: rgba(196,214,0,.55);
}

[data-testid="stMetricLabel"] {
    color: var(--ap-muted) !important;
    font-weight: 600;
}

[data-testid="stMetricValue"] {
    color: var(--ap-blue-grey-deep) !important;
    font-weight: 800;
    letter-spacing: -0.035em;
}

/* ---------- Buttons ---------- */
.stButton > button,
.stDownloadButton > button,
[data-testid="stFormSubmitButton"] > button {
    position: relative;
    overflow: hidden;
    min-height: 44px;
    border-radius: 12px !important;
    border: 1px solid var(--ap-lime-dark) !important;
    background: linear-gradient(135deg, var(--ap-lime) 0%, #C49FFF 100%) !important;
    color: #FFFFFF !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 750 !important;
    letter-spacing: -0.01em;
    box-shadow: 0 6px 16px rgba(138, 150, 0, .16);
    transition:
        transform .18s cubic-bezier(.2,.8,.2,1),
        box-shadow .18s ease,
        filter .18s ease,
        border-color .18s ease;
}

.stButton > button::before,
.stDownloadButton > button::before,
[data-testid="stFormSubmitButton"] > button::before {
    content: "";
    position: absolute;
    top: -100%;
    left: -35%;
    width: 32%;
    height: 300%;
    transform: rotate(24deg);
    background: rgba(255,255,255,.36);
    transition: left .5s ease;
    pointer-events: none;
}

.stButton > button:hover,
.stDownloadButton > button:hover,
[data-testid="stFormSubmitButton"] > button:hover {
    transform: translateY(-2px) scale(1.012);
    box-shadow: 0 10px 24px rgba(67, 88, 102, .20);
    filter: saturate(1.05) brightness(1.02);
    border-color: var(--ap-blue-grey-dark) !important;
}

.stButton > button:hover::before,
.stDownloadButton > button:hover::before,
[data-testid="stFormSubmitButton"] > button:hover::before {
    left: 120%;
}

.stButton > button:active,
.stDownloadButton > button:active,
[data-testid="stFormSubmitButton"] > button:active {
    transform: translateY(0) scale(.985);
    box-shadow: 0 3px 9px rgba(38,57,67,.14);
}

.stButton > button:focus-visible,
.stDownloadButton > button:focus-visible,
[data-testid="stFormSubmitButton"] > button:focus-visible {
    outline: 3px solid rgba(127,149,163,.28) !important;
    outline-offset: 2px;
}

/* Secondary buttons stay cool gray-blue when Streamlit marks them secondary */
.stButton > button[kind="secondary"],
.stDownloadButton > button[kind="secondary"] {
    background: linear-gradient(135deg, #FFFFFF, var(--ap-blue-grey-pale)) !important;
    color: var(--ap-blue-grey-deep) !important;
    border-color: #E3D4F2 !important;
    box-shadow: 0 5px 14px rgba(67,88,102,.10);
}

/* ---------- Inputs ---------- */
[data-baseweb="input"] > div,
[data-baseweb="select"] > div,
[data-baseweb="textarea"] > div,
.stNumberInput > div > div,
.stDateInput > div > div {
    border-radius: 11px !important;
}

[data-baseweb="input"] > div,
[data-baseweb="select"] > div,
[data-baseweb="textarea"] > div {
    border-color: #E0D3EC !important;
    background: rgba(255,255,255,.96) !important;
    transition: border-color .18s ease, box-shadow .18s ease;
}

[data-baseweb="input"] > div:focus-within,
[data-baseweb="select"] > div:focus-within,
[data-baseweb="textarea"] > div:focus-within {
    border-color: var(--ap-blue-grey) !important;
    box-shadow: 0 0 0 3px rgba(196,214,0,.14) !important;
}

/* ---------- Tabs ---------- */
[data-baseweb="tab-list"] {
    gap: .35rem;
    background: rgba(255,255,255,.75);
    border: 1px solid var(--ap-border);
    border-radius: 13px;
    padding: .3rem;
}

button[data-baseweb="tab"] {
    border-radius: 9px !important;
    color: var(--ap-muted) !important;
    font-weight: 650 !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
}

button[data-baseweb="tab"][aria-selected="true"] {
    background: var(--ap-lime-pale) !important;
    color: var(--ap-blue-grey-deep) !important;
}

[data-baseweb="tab-highlight"] {
    background-color: var(--ap-lime) !important;
}

/* ---------- Radio / checkbox ---------- */
[data-testid="stCheckbox"] label,
[data-testid="stRadio"] label {
    font-weight: 520;
}

/* ---------- Alerts ---------- */
[data-testid="stAlert"] {
    border-radius: 13px !important;
    border: 1px solid var(--ap-border) !important;
    box-shadow: 0 4px 14px rgba(38,57,67,.05);
}

/* ---------- Dataframes / tables ---------- */
[data-testid="stDataFrame"],
[data-testid="stTable"] {
    border: 1px solid var(--ap-border);
    border-radius: 14px;
    overflow: hidden;
    background: white;
    box-shadow: 0 7px 22px rgba(76,48,112,.08);
}

/* ---------- Progress ---------- */
[data-testid="stProgress"] > div > div > div > div {
    background: linear-gradient(90deg, var(--ap-lime), var(--ap-blue-grey)) !important;
}

/* ---------- Dividers ---------- */
hr {
    border: 0 !important;
    height: 1px !important;
    background: linear-gradient(90deg, transparent, #E2D4EE, transparent) !important;
    margin: 1.8rem 0 !important;
}

/* ---------- Links ---------- */
a {
    color: var(--ap-blue-grey-dark) !important;
    text-decoration-color: rgba(196,214,0,.8) !important;
    text-underline-offset: 3px;
}

a:hover {
    color: var(--ap-lime-dark) !important;
}

/* ---------- Custom utility classes for st.markdown HTML ---------- */
.ap-center {
    text-align: center;
    margin-left: auto;
    margin-right: auto;
}

.ap-hero {
    max-width: 1040px;
    margin: 0 auto 1.4rem;
    padding: 1.35rem 1.5rem;
    text-align: center;
    background: linear-gradient(135deg, rgba(255,255,255,.95), rgba(242,246,248,.95));
    border: 1px solid var(--ap-border);
    border-radius: 20px;
    box-shadow: var(--ap-shadow);
}

.ap-card {
    background: rgba(255,255,255,.94);
    border: 1px solid var(--ap-border);
    border-radius: var(--ap-radius);
    padding: 1.15rem 1.2rem;
    box-shadow: var(--ap-shadow);
    height: 100%;
    transition: transform .22s ease, box-shadow .22s ease;
}

.ap-card:hover {
    transform: translateY(-3px);
    box-shadow: var(--ap-shadow-hover);
}

.ap-badge {
    display: inline-flex;
    align-items: center;
    gap: .35rem;
    padding: .38rem .72rem;
    border-radius: 999px;
    background: var(--ap-lime-pale);
    color: var(--ap-blue-grey-deep);
    border: 1px solid rgba(196,214,0,.42);
    font-size: .82rem;
    font-weight: 700;
}

.ap-subtitle {
    max-width: 850px;
    margin: -.65rem auto 1.5rem;
    text-align: center;
    color: var(--ap-muted) !important;
    font-size: 1rem;
}

/* ---------- Responsive ---------- */
@media (max-width: 768px) {
    .main .block-container {
        padding-top: 1.35rem;
        padding-left: .9rem;
        padding-right: .9rem;
    }

    h1 {
        font-size: 2rem !important;
    }

    h2 {
        font-size: 1.35rem !important;
    }

    [data-testid="stMetric"] {
        min-height: 104px;
    }
}

/* Accessibility: respect reduced motion */
@media (prefers-reduced-motion: reduce) {
    *, *::before, *::after {
        animation-duration: .01ms !important;
        animation-iteration-count: 1 !important;
        transition-duration: .01ms !important;
        scroll-behavior: auto !important;
    }
}


/* =========================================================
   AUDITPREP V8.5 — FINAL VISUAL OVERRIDES
   Existing business/layout classes are preserved.
   Palette: #7D4FFE / #C49FFF / #FFD0E6
   ========================================================= */
:root {
    --audit-red: #7D4FFE;
    --audit-red-dark: #6036D6;
    --audit-navy: #B75F89;
    --audit-navy-dark: #5A3147;
    --audit-orange: #7D4FFE;
    --audit-orange-dark: #6036D6;
    --audit-page: #EEE9F8;
}

/* Main page alignment */
.main .block-container {
    width: calc(100% - 28px) !important;
    max-width: 1600px !important;
    margin-left: auto !important;
    margin-right: auto !important;
    padding-left: clamp(1rem, 1.8vw, 2rem) !important;
    padding-right: clamp(1rem, 1.8vw, 2rem) !important;
    padding-top: 1.35rem !important;
}

/* Streamlit headings: deliberate hierarchy */
.main h1 {
    text-align: center !important;
    max-width: 1000px;
    margin-left: auto !important;
    margin-right: auto !important;
}
.main h2 {
    text-align: center !important;
    padding-left: 0 !important;
    border-left: 0 !important;
    margin-left: auto !important;
    margin-right: auto !important;
    color: var(--ap-blue-grey-deep) !important;
}
.main h2::after {
    content: "";
    display: block;
    width: 62px;
    height: 4px;
    margin: .55rem auto 0;
    border-radius: 999px;
    background: linear-gradient(90deg, var(--ap-lime), var(--ap-blue-grey));
}
.main h3 {
    color: var(--ap-blue-grey-dark) !important;
}
.audit-mini-card h3,
.audit-card h3,
.audit-hero-title,
.audit-login-title {
    text-align: center !important;
}

/* Keep sidebar headings practical rather than centered */
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
    text-align: left !important;
    border-left: 0 !important;
    padding-left: 0 !important;
}
[data-testid="stSidebar"] h2::after { margin-left: 0 !important; }

/* Top navigation */
.audit-topbar {
    width: 100% !important;
    max-width: none !important;
    min-height: 76px !important;
    margin: -1.35rem auto 1.9rem auto !important;
    padding: 0.88rem clamp(1.25rem, 2.4vw, 2.25rem) !important;
    background: linear-gradient(125deg, var(--ap-blue-grey-deep), var(--ap-blue-grey-dark)) !important;
    border-bottom: 3px solid var(--ap-lime) !important;
    box-shadow: 0 14px 32px rgba(76,48,112,.18) !important;
}
.audit-brand-mark {
    background: var(--ap-lime) !important;
    color: #FFFFFF !important;
    box-shadow: 0 5px 16px rgba(125,79,254,.24);
}
.audit-brand-name {
    background: rgba(255,255,255,.10) !important;
    color: #fff !important;
    border: 1px solid rgba(255,255,255,.15);
}
.audit-nav-items { color: rgba(255,255,255,.82) !important; }
.audit-nav-active {
    color: var(--ap-lime) !important;
    position: relative;
}

/* Hero / central focus */
.audit-hero {
    max-width: 1120px;
    margin: 0 auto 1.5rem !important;
    background:
        radial-gradient(circle at 15% 15%, rgba(125,79,254,.18), transparent 28%),
        linear-gradient(135deg, var(--ap-blue-grey-deep), #7D4FFE) !important;
    border: 1px solid rgba(196,159,255,.34) !important;
    border-radius: 22px !important;
    padding: clamp(2rem, 5vw, 3.2rem) clamp(1.2rem, 4vw, 2.6rem) !important;
    box-shadow: 0 18px 44px rgba(76,48,112,.18) !important;
}
.audit-hero-title {
    max-width: 900px;
    margin: 0 auto !important;
    color: #fff !important;
    line-height: 1.12 !important;
}
.audit-hero-title::after {
    content: "";
    display: block;
    width: 74px;
    height: 5px;
    margin: .9rem auto .2rem;
    border-radius: 999px;
    background: var(--ap-lime);
}
.audit-hero-subtitle {
    max-width: 790px !important;
    margin: .8rem auto 0 !important;
    text-align: center !important;
    color: rgba(255,255,255,.82) !important;
    line-height: 1.65;
}

/* Existing dashboard cards re-skinned */
.audit-card,
.audit-soft-card,
.audit-kpi-card,
.audit-mini-card,
.audit-alert-card {
    background: rgba(248,244,255,.96) !important;
    border: 1px solid var(--ap-border) !important;
    box-shadow: var(--ap-shadow) !important;
    color: var(--ap-text) !important;
}
.audit-card,
.audit-mini-card {
    transition: transform .2s ease, box-shadow .2s ease, border-color .2s ease;
}
.audit-card:hover,
.audit-mini-card:hover {
    transform: translateY(-3px);
    border-color: rgba(125,79,254,.48) !important;
    box-shadow: var(--ap-shadow-hover) !important;
}
.audit-grid-3,
.audit-grid-4 {
    width: 100%;
    max-width: 1160px;
    margin-left: auto !important;
    margin-right: auto !important;
    align-items: stretch;
}
.audit-grid-3 .audit-mini-card,
.audit-grid-4 .audit-mini-card { text-align: center !important; }
.audit-kpi-value,
.audit-big-number {
    color: var(--ap-blue-grey-deep) !important;
}
.audit-kpi-label {
    color: var(--ap-muted) !important;
}
.audit-step-bubble {
    background: var(--ap-lime-soft) !important;
    color: var(--ap-lime-deep) !important;
    border: 1px solid rgba(125,79,254,.42);
    box-shadow: inset 0 0 0 3px rgba(255,255,255,.50);
}

/* Login */
.audit-login-shell {
    background:
        radial-gradient(circle at 50% 0%, rgba(125,79,254,.18), transparent 34%),
        linear-gradient(145deg, var(--ap-blue-grey-deep), var(--ap-blue-grey-dark)) !important;
    border: 1px solid rgba(196,159,255,.34) !important;
    box-shadow: 0 24px 60px rgba(76,48,112,.24) !important;
}
.audit-login-mark {
    background: var(--ap-lime) !important;
    color: #FFFFFF !important;
    box-shadow: 0 10px 28px rgba(125,79,254,.25) !important;
}

/* Sidebar and user chip */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #F8F4FF 0%, #EEE7F7 100%) !important;
    border-right: 1px solid var(--ap-border) !important;
}
.audit-user-chip {
    background: #F8F4FF !important;
    border: 1px solid var(--ap-border) !important;
    border-left: 4px solid var(--ap-lime) !important;
    box-shadow: 0 5px 16px rgba(76,48,112,.08);
}

/* Primary actions use the V8.5.27 violet palette, overriding legacy orange/red selectors */
.stButton > button[kind="primary"],
.stDownloadButton > button[kind="primary"],
[data-testid="stFormSubmitButton"] > button[kind="primary"] {
    background: linear-gradient(135deg, var(--ap-lime), #C49FFF) !important;
    border-color: var(--ap-lime-dark) !important;
    color: #FFFFFF !important;
}
.stButton > button[kind="primary"]:hover,
.stDownloadButton > button[kind="primary"]:hover,
[data-testid="stFormSubmitButton"] > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #956FD0, var(--ap-lime)) !important;
    border-color: var(--ap-blue-grey-dark) !important;
    color: #FFFFFF !important;
}

/* Tints/shades on supporting information blocks */
.audit-info-box,
.audit-note {
    background: var(--ap-blue-grey-pale) !important;
    border-color: #E5D4EE !important;
    border-left: 4px solid var(--ap-blue-grey) !important;
}
.audit-warning {
    background: #FFF0F7 !important;
    border-color: #E7C6D8 !important;
    border-left: 4px solid var(--ap-lime-dark) !important;
}
.audit-success {
    background: #F0EAFF !important;
    border-color: #DCCFFF !important;
    border-left: 4px solid var(--ap-lime-grey-dark) !important;
}

/* Tabs: remove previous orange active state */
button[data-baseweb="tab"][aria-selected="true"] {
    color: var(--ap-blue-grey-deep) !important;
    border-bottom-color: var(--ap-lime) !important;
}

/* Center CTA rows and common form containers */
[data-testid="stForm"] {
    margin-left: auto;
    margin-right: auto;
}

/* Responsive central layout */
@media (max-width: 980px) {
    .audit-grid-4 { grid-template-columns: repeat(2, minmax(0,1fr)) !important; }
}
@media (max-width: 760px) {
    .audit-grid-3,
    .audit-grid-4 { grid-template-columns: 1fr !important; }
    .audit-topbar { border-radius: 0 0 14px 14px !important; }
    .audit-hero { border-radius: 18px !important; }
    .main h2 { font-size: 1.35rem !important; }
}


/* =========================================================
   V8.5.27 — HEADER LARGE ÉCRAN
   ========================================================= */
.audit-brand {
    gap: .82rem !important;
}

.audit-brand-mark {
    width: 42px !important;
    height: 42px !important;
    font-size: .88rem !important;
    border-radius: 11px !important;
}

.audit-brand-name {
    font-size: .92rem !important;
    padding: .42rem .72rem !important;
}

.audit-nav-items {
    gap: clamp(.95rem, 1.45vw, 1.45rem) !important;
    font-size: clamp(.82rem, .82vw, .94rem) !important;
    white-space: nowrap !important;
}

/* On large displays the header almost reaches both sides of the content area. */
@media (min-width: 1200px) {
    .audit-topbar {
        width: 100% !important;
        max-width: none !important;
    }
}

/* Keep the header clean on tablets/mobile. */
@media (max-width: 900px) {
    .main .block-container {
        width: 100% !important;
        max-width: 100% !important;
        padding-left: .9rem !important;
        padding-right: .9rem !important;
    }

    .audit-topbar {
        min-height: 66px !important;
        padding: .72rem 1rem !important;
    }
}


/* =========================================================
   V8.5.27 — GLASS DARK FIGMA-INSPIRED THEME
   Visual-only override
   Source inspiration: dark purple figma glass effect reference
   ========================================================= */

:root {
    --gf-bg: #181319;
    --gf-bg-2: #21162A;
    --gf-bg-3: #2B1840;
    --gf-panel: rgba(255,255,255,0.06);
    --gf-panel-2: rgba(255,255,255,0.08);
    --gf-panel-3: rgba(255,255,255,0.10);
    --gf-border: rgba(255,255,255,0.14);
    --gf-border-soft: rgba(255,255,255,0.08);

    --gf-text: #FFFFFF;
    --gf-text-soft: #E9E4EE;
    --gf-muted: #C9C0D3;
    --gf-subtle: #AFA5BC;

    --gf-purple-1: #5F2EFF;
    --gf-purple-2: #7E46FF;
    --gf-purple-3: #9A61FF;
    --gf-pink: #C28BFF;
    --gf-blue-glow: #45C7FF;

    --gf-shadow-xl: 0 24px 70px rgba(0,0,0,.35);
    --gf-shadow-lg: 0 18px 48px rgba(0,0,0,.28);
    --gf-shadow-md: 0 12px 30px rgba(0,0,0,.24);
    --gf-glass-blur: blur(18px) saturate(140%);
}

/* -------- GLOBAL TYPO -------- */
html,
body,
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stSidebar"],
button,
input,
textarea,
select,
.main,
.main * {
    font-family: "Inter", system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
}

html, body, .stApp,
[data-testid="stAppViewContainer"] {
    background-color: var(--gf-bg) !important;
}

/* Main page dark purple animated background */
[data-testid="stAppViewContainer"] {
    background:
        radial-gradient(circle at 18% 18%, rgba(126,70,255,.23), transparent 28%),
        radial-gradient(circle at 85% 12%, rgba(194,139,255,.12), transparent 24%),
        radial-gradient(circle at 78% 78%, rgba(95,46,255,.20), transparent 30%),
        radial-gradient(circle at 38% 72%, rgba(69,199,255,.07), transparent 22%),
        linear-gradient(135deg, #181319 0%, #1D1421 24%, #23152C 44%, #261637 62%, #22152F 78%, #181319 100%) !important;
    background-size: 140% 140%, 145% 145%, 150% 150%, 145% 145%, 100% 100% !important;
    animation: gfAmbientDrift 22s ease-in-out infinite alternate !important;
}

/* Top white host header strip if present */
header[data-testid="stHeader"] {
    background: rgba(10, 8, 12, 0.22) !important;
    backdrop-filter: blur(10px) !important;
}

@keyframes gfAmbientDrift {
    0% {
        background-position: 0% 0%, 100% 0%, 86% 100%, 30% 80%, 0 0;
    }
    50% {
        background-position: 8% 10%, 92% 8%, 76% 88%, 38% 73%, 0 0;
    }
    100% {
        background-position: 16% 6%, 86% 16%, 68% 82%, 45% 68%, 0 0;
    }
}

/* -------- MAIN CONTAINER -------- */
.main .block-container {
    background: transparent !important;
}

/* -------- TEXT -------- */
body,
.main,
.main p,
.main li,
.main label,
.main div,
.main span,
.main small,
.main strong,
.main em,
.main code,
[data-testid="stMarkdownContainer"],
[data-testid="stMarkdownContainer"] *,
[data-testid="stCaptionContainer"],
[data-testid="stText"],
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] * {
    color: var(--gf-text-soft) !important;
}

.main h1,
.main h2,
.main h3,
.main h4,
.main h5,
.main h6,
[data-testid="stMarkdownContainer"] h1,
[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3 {
    color: var(--gf-text) !important;
    letter-spacing: -0.02em !important;
}

/* Links */
a,
.main a,
[data-testid="stMarkdownContainer"] a {
    color: #E8D8FF !important;
}
a:hover,
.main a:hover {
    color: #FFFFFF !important;
}

/* -------- SIDEBAR -------- */
[data-testid="stSidebar"] {
    background:
        linear-gradient(180deg, rgba(28,21,32,.90), rgba(20,16,24,.96)) !important;
    border-right: 1px solid rgba(255,255,255,.08) !important;
    box-shadow: 8px 0 24px rgba(0,0,0,.22) !important;
}

[data-testid="stSidebar"] .block-container {
    background: transparent !important;
}

[data-testid="stSidebar"] *,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] div {
    color: var(--gf-text-soft) !important;
}

[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
    color: #FFFFFF !important;
}

[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,.08) !important;
}

.audit-version-block {
    background: linear-gradient(135deg, rgba(255,255,255,.07), rgba(255,255,255,.03)) !important;
    border: 1px solid rgba(255,255,255,.12) !important;
    color: var(--gf-muted) !important;
    box-shadow: var(--gf-shadow-md) !important;
    backdrop-filter: var(--gf-glass-blur) !important;
    -webkit-backdrop-filter: var(--gf-glass-blur) !important;
}
.audit-version-block * {
    color: var(--gf-muted) !important;
}
.audit-version-block b,
.audit-version-title {
    color: #FFFFFF !important;
}

/* -------- GLASS PANELS / CARDS -------- */
.audit-topbar,
.audit-hero,
.ap-hero,
.ap-card,
.audit-note,
.audit-mini-card,
[data-testid="stMetric"],
[data-testid="stForm"],
[data-testid="stExpander"] details,
[data-testid="stDataFrame"],
[data-testid="stTable"],
[data-baseweb="tab-list"],
[data-baseweb="input"] > div,
[data-baseweb="select"] > div,
[data-baseweb="textarea"] > div,
div[data-baseweb="notification"],
[data-testid="stAlert"] {
    background:
        linear-gradient(135deg, rgba(255,255,255,.09), rgba(255,255,255,.035)) !important;
    border: 1px solid rgba(255,255,255,.12) !important;
    box-shadow: var(--gf-shadow-lg) !important;
    backdrop-filter: var(--gf-glass-blur) !important;
    -webkit-backdrop-filter: var(--gf-glass-blur) !important;
}

/* -------- TOPBAR -------- */
.audit-topbar {
    position: relative !important;
    overflow: hidden !important;
    background:
        linear-gradient(135deg, rgba(65,31,96,.78), rgba(47,27,83,.72)) !important;
    border-bottom: 1px solid rgba(126,70,255,.38) !important;
    box-shadow:
        0 18px 42px rgba(0,0,0,.24),
        0 2px 0 rgba(126,70,255,.80),
        inset 0 1px 0 rgba(255,255,255,.12) !important;
}

.audit-topbar::before {
    content: "" !important;
    position: absolute !important;
    inset: 0 !important;
    background:
        radial-gradient(circle at 16% 20%, rgba(255,255,255,.10), transparent 18%),
        radial-gradient(circle at 82% 18%, rgba(126,70,255,.25), transparent 24%),
        linear-gradient(90deg, rgba(255,255,255,.04), transparent 25%, transparent 75%, rgba(255,255,255,.03));
    pointer-events: none !important;
}

.audit-brand-mark {
    background: linear-gradient(135deg, var(--gf-purple-2), var(--gf-purple-3)) !important;
    color: #FFFFFF !important;
    box-shadow: 0 10px 22px rgba(126,70,255,.28) !important;
}

.audit-brand-name {
    background: rgba(255,255,255,.08) !important;
    color: #FFFFFF !important;
    border: 1px solid rgba(255,255,255,.13) !important;
}

.audit-nav-items,
.audit-nav-items a,
.audit-nav-items span {
    color: rgba(255,255,255,.90) !important;
}
.audit-nav-items a:hover,
.audit-nav-items .active {
    color: #F3E9FF !important;
    text-shadow: 0 0 12px rgba(194,139,255,.20) !important;
}

/* -------- HERO -------- */
.audit-hero,
.ap-hero {
    position: relative !important;
    overflow: hidden !important;
    background:
        linear-gradient(135deg, rgba(22,14,28,.88) 0%, rgba(40,20,67,.88) 56%, rgba(52,24,88,.85) 100%) !important;
    border: 1px solid rgba(255,255,255,.10) !important;
    box-shadow:
        0 28px 64px rgba(0,0,0,.34),
        inset 0 1px 0 rgba(255,255,255,.11) !important;
}

.audit-hero::before,
.ap-hero::before {
    content: "" !important;
    position: absolute !important;
    inset: 0 !important;
    background:
        radial-gradient(circle at 14% 18%, rgba(126,70,255,.18), transparent 22%),
        radial-gradient(circle at 84% 22%, rgba(255,255,255,.06), transparent 18%),
        radial-gradient(circle at 74% 76%, rgba(95,46,255,.18), transparent 24%);
    pointer-events: none !important;
}

.audit-hero h1,
.audit-hero h2,
.ap-hero h1,
.ap-hero h2 {
    color: #FFFFFF !important;
    font-weight: 860 !important;
    letter-spacing: -0.03em !important;
}

.audit-hero p,
.audit-hero .stMarkdown,
.ap-hero p,
.ap-hero .stMarkdown {
    color: rgba(255,255,255,.92) !important;
}

.audit-hero h1::after,
.audit-hero h2::after,
.ap-hero h1::after,
.ap-hero h2::after {
    content: "" !important;
    display: block !important;
    width: 106px !important;
    height: 4px !important;
    margin: .9rem auto 0 auto !important;
    border-radius: 999px !important;
    background: linear-gradient(90deg, rgba(255,255,255,.12), var(--gf-purple-3), rgba(255,255,255,.12)) !important;
}

/* -------- MINI CARDS / NOTES / STEPS -------- */
.audit-mini-card,
.audit-note,
.ap-card {
    color: var(--gf-text-soft) !important;
}
.audit-mini-card h3,
.audit-note b,
.audit-step-bubble {
    color: #FFFFFF !important;
}

.audit-step-bubble {
    background: linear-gradient(135deg, var(--gf-purple-2), var(--gf-purple-3)) !important;
    box-shadow: 0 8px 22px rgba(126,70,255,.24) !important;
}

/* -------- FORMS / INPUTS -------- */
input,
textarea,
[data-baseweb="input"] input,
[data-baseweb="textarea"] textarea,
.stTextInput input,
.stNumberInput input,
.stTextArea textarea {
    color: #FFFFFF !important;
    caret-color: #FFFFFF !important;
}

[data-baseweb="input"] > div,
[data-baseweb="select"] > div,
[data-baseweb="textarea"] > div,
.stTextInput input,
.stNumberInput input,
.stTextArea textarea {
    background: rgba(255,255,255,.045) !important;
    border-color: rgba(255,255,255,.12) !important;
}

input::placeholder,
textarea::placeholder {
    color: rgba(233,228,238,.56) !important;
    opacity: 1 !important;
}

[data-baseweb="select"] *,
[data-baseweb="menu"] *,
[data-baseweb="popover"] *,
[data-baseweb="tag"] *,
[role="option"] *,
[role="listbox"] * {
    color: #FFFFFF !important;
}
[data-baseweb="menu"],
[data-baseweb="popover"],
[role="listbox"] {
    background: #24182E !important;
    border: 1px solid rgba(255,255,255,.10) !important;
}
[role="option"]:hover,
[aria-selected="true"][role="option"] {
    background: rgba(126,70,255,.22) !important;
}

/* -------- BUTTONS -------- */
.stButton > button,
.stDownloadButton > button,
[data-testid="stFormSubmitButton"] > button {
    background:
        linear-gradient(180deg, rgba(255,255,255,.10), rgba(255,255,255,.04)) !important;
    color: #FFFFFF !important;
    border: 1px solid rgba(255,255,255,.14) !important;
    border-radius: 999px !important;
    box-shadow:
        0 12px 28px rgba(0,0,0,.22),
        inset 0 1px 0 rgba(255,255,255,.12) !important;
}

.stButton > button:hover,
.stDownloadButton > button:hover,
[data-testid="stFormSubmitButton"] > button:hover {
    background:
        linear-gradient(180deg, rgba(255,255,255,.15), rgba(255,255,255,.06)) !important;
    border-color: rgba(194,139,255,.26) !important;
    box-shadow:
        0 16px 34px rgba(0,0,0,.28),
        inset 0 1px 0 rgba(255,255,255,.14) !important;
}

/* Let main CTA keep a purple premium emphasis */
button[kind="primary"],
.stButton > button[kind="primary"],
[data-testid="stFormSubmitButton"] > button {
    background:
        linear-gradient(135deg, rgba(126,70,255,.82), rgba(95,46,255,.76)) !important;
    border-color: rgba(194,139,255,.26) !important;
}

/* -------- METRICS / TABLES / TABS -------- */
[data-testid="stMetric"] * {
    color: var(--gf-text-soft) !important;
}
[data-testid="stMetricValue"] {
    color: #FFFFFF !important;
}
[data-testid="stMetricLabel"] {
    color: var(--gf-muted) !important;
}

[data-baseweb="tab"] {
    color: var(--gf-muted) !important;
}
[data-baseweb="tab"][aria-selected="true"] {
    color: #FFFFFF !important;
}

[data-testid="stDataFrame"],
[data-testid="stTable"],
[data-testid="stDataFrame"] *,
[data-testid="stTable"] * {
    color: var(--gf-text-soft) !important;
}

/* -------- EXPANDERS / ALERTS / CODE -------- */
[data-testid="stExpander"] summary,
[data-testid="stExpander"] summary *,
[data-testid="stExpander"] details,
[data-testid="stExpander"] details * {
    color: var(--gf-text-soft) !important;
}

pre,
code,
[data-testid="stCodeBlock"],
[data-testid="stCodeBlock"] * {
    color: #F2EDFA !important;
}
pre,
[data-testid="stCodeBlock"] {
    background: rgba(18,13,22,.92) !important;
    border: 1px solid rgba(255,255,255,.08) !important;
}

[data-testid="stAlert"],
[data-testid="stAlert"] * {
    color: #FFFFFF !important;
}

/* -------- SMALL POLISH -------- */
hr {
    border-color: rgba(255,255,255,.08) !important;
}

.audit-topbar,
.audit-hero,
.ap-hero,
.ap-card,
.audit-mini-card,
.audit-note,
[data-testid="stMetric"],
[data-testid="stForm"],
[data-testid="stExpander"] details {
    border-radius: 22px !important;
}

.audit-topbar {
    border-radius: 0 0 20px 20px !important;
}

@media (prefers-reduced-motion: reduce) {
    [data-testid="stAppViewContainer"] {
        animation: none !important;
    }
}


/* =========================================================
   V8.5.27 — READABILITY / CONTRAST PATCH
   Fix white/gray text that is too hard to read.
   Visual-only patch.
   ========================================================= */

:root {
    --gf-text: #FFFFFF;
    --gf-text-soft: #F6F2FA;
    --gf-muted: #E2DAEA;
    --gf-subtle: #CBBFD7;
    --gf-label: #EFE8F6;
    --gf-dim: #CFC4DA;
}

/* Global content */
body,
.main,
.main p,
.main li,
.main div,
.main span,
.main small,
.main strong,
.main em,
.main code,
.main label,
[data-testid="stMarkdownContainer"],
[data-testid="stMarkdownContainer"] *,
[data-testid="stCaptionContainer"],
[data-testid="stText"],
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] *,
[data-testid="stAlert"],
[data-testid="stAlert"] *,
div[data-baseweb="notification"],
div[data-baseweb="notification"] * {
    color: var(--gf-text-soft) !important;
}

/* Headings */
.main h1,
.main h2,
.main h3,
.main h4,
.main h5,
.main h6,
[data-testid="stMarkdownContainer"] h1,
[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3 {
    color: #FFFFFF !important;
}

/* Sidebar readability */
[data-testid="stSidebar"] *,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div,
[data-testid="stSidebar"] small {
    color: var(--gf-text-soft) !important;
}

[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] strong,
[data-testid="stSidebar"] b {
    color: #FFFFFF !important;
}

[data-testid="stSidebar"] [data-testid="stCaptionContainer"],
[data-testid="stSidebar"] .stCaption {
    color: var(--gf-dim) !important;
}

/* Topbar / nav */
.audit-topbar,
.audit-topbar *,
.audit-brand-name,
.audit-nav-items,
.audit-nav-items *,
.audit-brand-mark {
    color: #FFFFFF !important;
}

.audit-nav-items a,
.audit-nav-items span {
    color: rgba(255,255,255,.92) !important;
}

.audit-nav-items a:hover,
.audit-nav-items .active {
    color: #FFFFFF !important;
}

/* Hero and key cards */
.audit-hero *,
.ap-hero *,
.audit-mini-card *,
.audit-note *,
.ap-card * {
    color: var(--gf-text-soft) !important;
}

.audit-hero h1,
.audit-hero h2,
.ap-hero h1,
.ap-hero h2,
.audit-mini-card h3,
.audit-note b,
.ap-card h3 {
    color: #FFFFFF !important;
}

/* Inputs / text fields */
input,
textarea,
select,
[data-baseweb="input"] input,
[data-baseweb="textarea"] textarea,
[data-baseweb="select"] input,
.stTextInput input,
.stNumberInput input,
.stTextArea textarea {
    color: #FFFFFF !important;
    caret-color: #FFFFFF !important;
}

[data-baseweb="input"] > div,
[data-baseweb="select"] > div,
[data-baseweb="textarea"] > div,
.stTextInput input,
.stNumberInput input,
.stTextArea textarea {
    color: #FFFFFF !important;
}

input::placeholder,
textarea::placeholder {
    color: #D7CDE1 !important;
    opacity: 1 !important;
}

/* Dropdown menus / options */
[data-baseweb="menu"] *,
[data-baseweb="popover"] *,
[data-baseweb="select"] *,
[data-baseweb="tag"] *,
[role="listbox"] *,
[role="option"] * {
    color: #FFFFFF !important;
}

[data-baseweb="menu"],
[data-baseweb="popover"],
[role="listbox"] {
    background: #24182E !important;
}

[role="option"] {
    color: #FFFFFF !important;
}

[role="option"]:hover,
[aria-selected="true"][role="option"] {
    color: #FFFFFF !important;
    background: rgba(126,70,255,.24) !important;
}

/* Tabs */
[data-baseweb="tab"] *,
[data-baseweb="tab"] {
    color: var(--gf-dim) !important;
}

[data-baseweb="tab"][aria-selected="true"],
[data-baseweb="tab"][aria-selected="true"] *,
button[role="tab"][aria-selected="true"],
button[role="tab"][aria-selected="true"] * {
    color: #FFFFFF !important;
}

/* Metrics */
[data-testid="stMetricLabel"],
[data-testid="stMetricLabel"] * {
    color: var(--gf-muted) !important;
}

[data-testid="stMetricValue"],
[data-testid="stMetricValue"] * {
    color: #FFFFFF !important;
}

/* Tables / dataframes */
[data-testid="stTable"],
[data-testid="stTable"] *,
[data-testid="stDataFrame"],
[data-testid="stDataFrame"] * {
    color: #F6F2FA !important;
}

[data-testid="stDataFrame"] [role="columnheader"],
[data-testid="stDataFrame"] [role="gridcell"],
[data-testid="stDataFrame"] [data-testid="StyledDataFrameCell"],
[data-testid="stDataFrame"] [data-testid="StyledDataFrameCell"] * {
    color: #F6F2FA !important;
}

/* Expander */
[data-testid="stExpander"] summary,
[data-testid="stExpander"] summary *,
[data-testid="stExpander"] details,
[data-testid="stExpander"] details * {
    color: #F4EFF8 !important;
}

/* Code / pre */
pre,
code,
[data-testid="stCodeBlock"],
[data-testid="stCodeBlock"] * {
    color: #F8F4FC !important;
}

/* Version block */
.audit-version-block,
.audit-version-block * {
    color: var(--gf-muted) !important;
}
.audit-version-block b,
.audit-version-title {
    color: #FFFFFF !important;
}

/* Labels / forms / helper copy */
label,
.stSelectbox label,
.stMultiSelect label,
.stTextInput label,
.stNumberInput label,
.stTextArea label,
.stDateInput label,
.stTimeInput label,
.stRadio label,
.stCheckbox label,
.stToggle label,
[data-testid="stWidgetLabel"] label,
[data-testid="stWidgetLabel"] span {
    color: var(--gf-label) !important;
}

/* Buttons text */
.stButton > button,
.stButton > button *,
.stDownloadButton > button,
.stDownloadButton > button *,
[data-testid="stFormSubmitButton"] > button,
[data-testid="stFormSubmitButton"] > button * {
    color: #FFFFFF !important;
}

/* Keep weaker metadata visible enough */
small,
.stCaption,
[data-testid="stCaptionContainer"] {
    color: var(--gf-dim) !important;
}


/* =========================================================
   V8.5.27 — TARGETED CONTRAST FIX
   Fixes the hard-to-read components visible in the UI screenshot:
   - sidebar user card
   - secondary and disabled buttons
   - KPI cards / metrics
   - info banners
   - tabs
   - expander headers
   - labels on violet surfaces
   ========================================================= */

/* -------- SIDEBAR user / identity card -------- */
/* Make the light card readable: dark text on light surface */
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] {
    border-color: rgba(255,255,255,.10) !important;
}

[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMarkdownContainer"],
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMarkdownContainer"] *,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] p,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] span,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] div,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] small,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] strong {
    color: #2B2035 !important;
}

/* If the sidebar contains a light internal card, force light background + dark text */
[data-testid="stSidebar"] .stContainer,
[data-testid="stSidebar"] [data-testid="stElementContainer"] .stAlert,
[data-testid="stSidebar"] [data-testid="stElementContainer"] .st-emotion-cache-1r6slb0,
[data-testid="stSidebar"] [data-testid="stElementContainer"] .st-emotion-cache-ocqkz7 {
    color: inherit !important;
}

/* -------- Buttons -------- */
/* Primary button stays bright and readable */
.stButton > button,
.stDownloadButton > button,
[data-testid="stFormSubmitButton"] > button {
    color: #FFFFFF !important;
    font-weight: 700 !important;
}

/* Secondary buttons: give them a stronger readable surface */
.stButton > button:not([kind="primary"]),
.stDownloadButton > button:not([kind="primary"]) {
    background: linear-gradient(180deg, rgba(255,255,255,.16), rgba(255,255,255,.08)) !important;
    color: #FFFFFF !important;
    border-color: rgba(255,255,255,.18) !important;
}

/* Disabled buttons should still be readable */
.stButton > button:disabled,
.stDownloadButton > button:disabled,
[data-testid="stFormSubmitButton"] > button:disabled,
button[disabled] {
    opacity: 1 !important;
    color: #F2ECF8 !important;
    background: linear-gradient(180deg, rgba(255,255,255,.14), rgba(255,255,255,.08)) !important;
    border: 1px solid rgba(255,255,255,.16) !important;
    box-shadow: none !important;
    cursor: not-allowed !important;
}
.stButton > button:disabled *,
.stDownloadButton > button:disabled *,
[data-testid="stFormSubmitButton"] > button:disabled *,
button[disabled] * {
    color: #F2ECF8 !important;
}

/* -------- Expanders / section bars -------- */
[data-testid="stExpander"] summary,
[data-testid="stExpander"] summary *,
details summary,
details summary * {
    color: #F8F4FC !important;
    font-weight: 700 !important;
}

/* -------- KPI cards / metric cards -------- */
[data-testid="stMetric"] {
    background:
        linear-gradient(135deg, rgba(84,45,122,.60), rgba(58,30,88,.64)) !important;
    border: 1px solid rgba(255,255,255,.14) !important;
}

[data-testid="stMetricLabel"],
[data-testid="stMetricLabel"] *,
[data-testid="stMetric"] label,
[data-testid="stMetric"] small {
    color: #E8E0F0 !important;
    font-weight: 650 !important;
}

[data-testid="stMetricValue"],
[data-testid="stMetricValue"] *,
[data-testid="stMetric"] [data-testid="stMarkdownContainer"] strong {
    color: #FFFFFF !important;
    font-weight: 800 !important;
}

/* -------- Info banners / notes / explanation blocks -------- */
.audit-note,
.audit-note *,
.ap-card,
.ap-card *,
.audit-mini-card,
.audit-mini-card * {
    color: #F5F1F9 !important;
}

.audit-note b,
.ap-card h3,
.audit-mini-card h3 {
    color: #FFFFFF !important;
}

/* Generic markdown/info strips that look too dim */
div[data-baseweb="notification"],
div[data-baseweb="notification"] *,
[data-testid="stAlert"],
[data-testid="stAlert"] * {
    color: #F8F4FC !important;
}

/* -------- Tabs -------- */
[data-baseweb="tab-list"] {
    background: linear-gradient(180deg, rgba(255,255,255,.08), rgba(255,255,255,.04)) !important;
    border: 1px solid rgba(255,255,255,.10) !important;
}

[data-baseweb="tab"],
button[role="tab"] {
    color: #E2D8EC !important;
    font-weight: 600 !important;
    opacity: 1 !important;
}

[data-baseweb="tab"] *,
button[role="tab"] * {
    color: #E2D8EC !important;
    opacity: 1 !important;
}

[data-baseweb="tab"][aria-selected="true"],
button[role="tab"][aria-selected="true"] {
    background: rgba(255,255,255,.12) !important;
    color: #FFFFFF !important;
    border-radius: 10px !important;
}

[data-baseweb="tab"][aria-selected="true"] *,
button[role="tab"][aria-selected="true"] * {
    color: #FFFFFF !important;
}

/* -------- Form labels / helper text -------- */
label,
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] *,
.stSelectbox label,
.stMultiSelect label,
.stTextInput label,
.stNumberInput label,
.stTextArea label,
.stDateInput label,
.stTimeInput label,
.stRadio label,
.stCheckbox label {
    color: #F1EBF7 !important;
    font-weight: 650 !important;
}

/* Helper captions need to be readable too */
small,
.stCaption,
[data-testid="stCaptionContainer"] {
    color: #D9CF E3 !important;
}

/* fix accidental invalid color token from previous theme if browser ignores it */
small,
.stCaption,
[data-testid="stCaptionContainer"] {
    color: #D9CFE3 !important;
}

/* -------- Dataframe / table header text -------- */
[data-testid="stDataFrame"] [role="columnheader"],
[data-testid="stDataFrame"] [role="columnheader"] *,
[data-testid="stTable"] th,
[data-testid="stTable"] th * {
    color: #FFFFFF !important;
    font-weight: 700 !important;
}

[data-testid="stDataFrame"] [role="gridcell"],
[data-testid="stDataFrame"] [role="gridcell"] *,
[data-testid="stTable"] td,
[data-testid="stTable"] td * {
    color: #F5F1F9 !important;
}

/* -------- Markdown paragraphs directly on dark background -------- */
.main [data-testid="stMarkdownContainer"] p,
.main [data-testid="stMarkdownContainer"] li,
.main [data-testid="stMarkdownContainer"] span {
    color: #F2ECF8 !important;
}

/* -------- Explicit stronger contrast for the left 'Version' list -------- */
.audit-version-block,
.audit-version-block * {
    color: #E6DDED !important;
}
.audit-version-block b,
.audit-version-title {
    color: #FFFFFF !important;
}


/* =========================================================
   V8.5.27 — WHITE / LIGHT CARD TEXT FIX
   Any text inside light / white frames becomes dark for readability.
   ========================================================= */

/* Sidebar user/info light card */
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"],
[data-testid="stSidebar"] .stAlert,
[data-testid="stSidebar"] .stContainer {
    color: inherit !important;
}

/* Strong dark text for content inside light cards in sidebar */
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] p,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] span,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] div,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] small,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] strong,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] b,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] label,
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMarkdownContainer"],
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMarkdownContainer"] * {
    color: #221A2B !important;
}

/* If a button becomes very light, make its text dark */
.stButton > button[disabled],
.stDownloadButton > button[disabled],
[data-testid="stFormSubmitButton"] > button[disabled],
button[disabled] {
    color: #241C2D !important;
}

.stButton > button[disabled] *,
.stDownloadButton > button[disabled] *,
[data-testid="stFormSubmitButton"] > button[disabled] *,
button[disabled] * {
    color: #241C2D !important;
}

/* Selected tab with very light background -> dark text */
[data-baseweb="tab"][aria-selected="true"],
button[role="tab"][aria-selected="true"] {
    color: #241C2D !important;
}

[data-baseweb="tab"][aria-selected="true"] *,
button[role="tab"][aria-selected="true"] * {
    color: #241C2D !important;
}

/* Any intentionally light block should render dark text */
.audit-light-surface,
.audit-light-surface *,
.audit-white-card,
.audit-white-card * {
    color: #221A2B !important;
}

/* Safety net for light backgrounds commonly seen in Streamlit blocks */
[style*="background: rgb(255, 255, 255)"],
[style*="background-color: rgb(255, 255, 255)"],
[style*="background:#fff"],
[style*="background: #fff"],
[style*="background-color:#fff"],
[style*="background-color: #fff"],
[style*="background:white"],
[style*="background: white"],
[style*="background-color:white"],
[style*="background-color: white"] {
    color: #221A2B !important;
}

[style*="background: rgb(255, 255, 255)"] *,
[style*="background-color: rgb(255, 255, 255)"] *,
[style*="background:#fff"] *,
[style*="background: #fff"] *,
[style*="background-color:#fff"] *,
[style*="background-color: #fff"] *,
[style*="background:white"] *,
[style*="background: white"] *,
[style*="background-color:white"] *,
[style*="background-color: white"] * {
    color: #221A2B !important;
}


/* =========================================================
   V8.5.27 — NAVBAR FUNCTIONAL FIX
   Real internal links, same visual theme.
   ========================================================= */

.audit-topbar {
    position: relative !important;
    z-index: 20 !important;
}

.audit-topbar::before,
.audit-topbar::after {
    pointer-events: none !important;
}

.audit-brand,
.audit-nav-items {
    position: relative !important;
    z-index: 22 !important;
}

.audit-nav-items {
    pointer-events: auto !important;
}

.audit-nav-link {
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    position: relative !important;
    z-index: 23 !important;
    pointer-events: auto !important;
    cursor: pointer !important;
    text-decoration: none !important;
    color: rgba(255,255,255,.90) !important;
    font-weight: 650 !important;
    white-space: nowrap !important;
    padding: .42rem .18rem !important;
    transition:
        color .18s ease,
        opacity .18s ease,
        transform .18s ease !important;
}

.audit-nav-link:visited {
    color: rgba(255,255,255,.90) !important;
}

.audit-nav-link:hover {
    color: #FFFFFF !important;
    text-decoration: none !important;
    transform: translateY(-1px) !important;
}

.audit-nav-link::after {
    content: "" !important;
    position: absolute !important;
    left: 50% !important;
    bottom: .08rem !important;
    width: 0 !important;
    height: 2px !important;
    border-radius: 999px !important;
    background: #C49FFF !important;
    transform: translateX(-50%) !important;
    transition: width .18s ease !important;
}

.audit-nav-link:hover::after,
.audit-nav-link.audit-nav-active::after {
    width: 72% !important;
}

.audit-nav-link.audit-nav-active {
    color: #FFFFFF !important;
    font-weight: 800 !important;
}

.audit-nav-user {
    color: rgba(255,255,255,.76) !important;
    font-weight: 550 !important;
    white-space: nowrap !important;
    pointer-events: none !important;
}

/* Invisible scroll targets. Negative margin gives comfortable visual offset. */
.audit-anchor {
    position: relative !important;
    display: block !important;
    width: 1px !important;
    height: 1px !important;
    margin-top: -18px !important;
    padding-top: 18px !important;
    visibility: hidden !important;
}

/* Smooth in-page navigation. */
html {
    scroll-behavior: smooth !important;
}

@media (prefers-reduced-motion: reduce) {
    html {
        scroll-behavior: auto !important;
    }
}


/* =========================================================
   V8.5.27 — CARTES DE VIGILANCE : TEXTE NOIR
   ========================================================= */

/* Fond clair => texte noir pour une lecture nette */
.audit-alert-card,
.audit-alert-card *,
.audit-alert-card b,
.audit-alert-card strong {
    color: #111111 !important;
}

/* Les badges Élevée / Modérée / Faible restent colorés en fond,
   mais leur texte devient noir aussi */
.audit-alert-card .audit-badge,
.audit-alert-card .audit-badge *,
.audit-alert-card .badge-high,
.audit-alert-card .badge-medium,
.audit-alert-card .badge-low {
    color: #111111 !important;
}

/* Renforce légèrement les libellés */
.audit-alert-card > b:first-child {
    font-weight: 800 !important;
}


/* =========================================================
   V8.5.27 — TEXTES DEMANDÉS EN NOIR
   Ciblage uniquement des blocs indiqués par l'utilisateur.
   ========================================================= */

.audit-requested-black,
.audit-requested-black *,
.audit-requested-black b,
.audit-requested-black strong,
.audit-requested-black span,
.audit-requested-black p,
.audit-requested-black h1,
.audit-requested-black h2,
.audit-requested-black h3,
.audit-requested-black small {
    color: #111111 !important;
}

/* Badge du lot sélectionné */
.audit-requested-black .audit-badge,
.audit-requested-black .audit-badge * {
    color: #111111 !important;
}

/* Titre Variables textuelles les plus influentes */
.audit-requested-black-heading,
h3.audit-requested-black-heading {
    color: #111111 !important;
    font-weight: 800 !important;
}

/* Carte identité sidebar : Administrateur AuditPrep / Administrateur */
.audit-user-chip.audit-requested-black,
.audit-user-chip.audit-requested-black *,
.audit-user-chip.audit-requested-black b,
.audit-user-chip.audit-requested-black span {
    color: #111111 !important;
}

/* If these targeted cards use a light surface, keep it clearly readable. */
.audit-card.audit-requested-black,
.audit-info-box.audit-requested-black,
.audit-user-chip.audit-requested-black {
    text-shadow: none !important;
}


/* =========================================================
   V8.5.27 — MISSIONS SELECT + PRINCIPE DE SÉCURITÉ EN NOIR
   ========================================================= */

/* Selectbox des missions : fond clair + texte noir */
[data-baseweb="select"] > div {
    background: #FFFFFF !important;
    color: #111111 !important;
    border-color: rgba(125,79,254,.22) !important;
}

[data-baseweb="select"] span,
[data-baseweb="select"] div,
[data-baseweb="select"] input {
    color: #111111 !important;
}

/* Liste déroulante des missions */
[data-baseweb="popover"],
[data-baseweb="menu"],
[role="listbox"] {
    background: #FFFFFF !important;
}

[role="option"],
[role="option"] *,
[data-baseweb="menu"] *,
[data-baseweb="popover"] * {
    color: #111111 !important;
}

[role="option"]:hover,
[role="option"][aria-selected="true"] {
    background: #EEE6FF !important;
    color: #111111 !important;
}

[role="option"]:hover *,
[role="option"][aria-selected="true"] * {
    color: #111111 !important;
}

/* Principe de sécurité et autres warning sur fond clair */
.audit-warning,
.audit-warning *,
.audit-warning b,
.audit-warning strong,
.audit-warning span,
.audit-warning p {
    color: #111111 !important;
    text-shadow: none !important;
}


/* =========================================================
   V8.5.27 — INFOBULLES D'AIDE LISIBLES
   Exemple : "Les métriques sont calculées sur 30 %..."
   ========================================================= */

[role="tooltip"],
[role="tooltip"] *,
[data-baseweb="tooltip"],
[data-baseweb="tooltip"] *,
[data-baseweb="popover"] [role="tooltip"],
[data-baseweb="popover"] [role="tooltip"] * {
    color: #111111 !important;
}

/* Fond clair pour garantir noir sur clair */
[role="tooltip"],
[data-baseweb="tooltip"] {
    background: #FFFFFF !important;
    border: 1px solid rgba(125,79,254,.20) !important;
    box-shadow: 0 10px 28px rgba(0,0,0,.18) !important;
}


/* =========================================================
   V8.5.27 — PREMIUM SECTION HEADERS
   Six requested headings only.
   ========================================================= */

.audit-section-head {
    --section-accent-1: #7D4FFE;
    --section-accent-2: #C49FFF;

    position: relative !important;
    isolation: isolate !important;
    overflow: hidden !important;

    display: flex !important;
    align-items: center !important;

    width: 100% !important;
    min-height: 76px !important;
    margin: 1.65rem 0 1rem 0 !important;
    padding: 1.05rem 1.35rem 1.05rem 1.55rem !important;

    border: 1px solid rgba(255,255,255,.14) !important;
    border-radius: 18px !important;

    background:
        linear-gradient(
            105deg,
            rgba(255,255,255,.095) 0%,
            rgba(255,255,255,.055) 38%,
            rgba(125,79,254,.10) 100%
        ) !important;

    box-shadow:
        0 14px 34px rgba(0,0,0,.20),
        inset 0 1px 0 rgba(255,255,255,.10) !important;

    backdrop-filter: blur(18px) saturate(135%) !important;
    -webkit-backdrop-filter: blur(18px) saturate(135%) !important;
}

/* Accent rail */
.audit-section-head::before {
    content: "" !important;
    position: absolute !important;
    left: 0 !important;
    top: 12px !important;
    bottom: 12px !important;
    width: 5px !important;
    border-radius: 0 999px 999px 0 !important;
    background: linear-gradient(
        180deg,
        var(--section-accent-1),
        var(--section-accent-2)
    ) !important;
    box-shadow: 0 0 20px color-mix(in srgb, var(--section-accent-1) 48%, transparent) !important;
}

/* Soft glow on the right */
.audit-section-head::after {
    content: "" !important;
    position: absolute !important;
    z-index: -1 !important;
    width: 210px !important;
    height: 210px !important;
    right: -78px !important;
    top: -88px !important;
    border-radius: 50% !important;
    background: radial-gradient(
        circle,
        color-mix(in srgb, var(--section-accent-2) 24%, transparent) 0%,
        transparent 68%
    ) !important;
    pointer-events: none !important;
}

.audit-section-head h2 {
    position: relative !important;
    z-index: 2 !important;

    margin: 0 !important;
    padding: 0 !important;

    color: #FFFFFF !important;
    font-family: "Inter", system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
    font-size: clamp(1.28rem, 1.65vw, 1.72rem) !important;
    line-height: 1.2 !important;
    font-weight: 780 !important;
    letter-spacing: -0.025em !important;
    text-shadow: 0 2px 18px rgba(0,0,0,.18) !important;
}

/* Semantic accent variations — same component, cleaner hierarchy */
.audit-section-head--why {
    --section-accent-1: #7D4FFE;
    --section-accent-2: #C49FFF;
}

.audit-section-head--chain {
    --section-accent-1: #8D63FF;
    --section-accent-2: #7FB7FF;
}

.audit-section-head--tools {
    --section-accent-1: #A873FF;
    --section-accent-2: #FFD0E6;
}

.audit-section-head--summary {
    --section-accent-1: #6F55FF;
    --section-accent-2: #A7B8FF;
}

.audit-section-head--alerts {
    --section-accent-1: #B96CFF;
    --section-accent-2: #FF9DCB;
}

.audit-section-head--decisions {
    --section-accent-1: #7D4FFE;
    --section-accent-2: #73D5FF;
}

/* Micro interaction, without changing layout */
.audit-section-head {
    transition:
        transform .22s ease,
        border-color .22s ease,
        box-shadow .22s ease !important;
}

.audit-section-head:hover {
    transform: translateY(-1px) !important;
    border-color: rgba(196,159,255,.28) !important;
    box-shadow:
        0 18px 40px rgba(0,0,0,.24),
        inset 0 1px 0 rgba(255,255,255,.13) !important;
}

@media (max-width: 760px) {
    .audit-section-head {
        min-height: 66px !important;
        margin: 1.25rem 0 .85rem 0 !important;
        padding: .9rem 1rem .9rem 1.25rem !important;
        border-radius: 15px !important;
    }

    .audit-section-head h2 {
        font-size: 1.22rem !important;
    }
}

@media (prefers-reduced-motion: reduce) {
    .audit-section-head {
        transition: none !important;
    }

    .audit-section-head:hover {
        transform: none !important;
    }
}


/* =========================================================
   V8.5.27 — WORKFLOW STEP HEADERS
   Étapes 1→5 : même taille, thème distinct des grands titres.
   ========================================================= */

.audit-workflow-head {
    position: relative !important;
    overflow: hidden !important;

    display: flex !important;
    align-items: center !important;
    gap: 1rem !important;

    width: 100% !important;
    min-height: 68px !important;
    margin: 1.35rem 0 .8rem 0 !important;
    padding: .78rem 1.15rem .78rem .9rem !important;

    border: 1px solid rgba(196,159,255,.22) !important;
    border-radius: 15px !important;

    background:
        linear-gradient(
            90deg,
            rgba(125,79,254,.13) 0%,
            rgba(255,255,255,.065) 24%,
            rgba(255,255,255,.035) 100%
        ) !important;

    box-shadow:
        0 10px 26px rgba(0,0,0,.16),
        inset 0 1px 0 rgba(255,255,255,.08) !important;

    backdrop-filter: blur(14px) saturate(125%) !important;
    -webkit-backdrop-filter: blur(14px) saturate(125%) !important;
}

/* Thin progression line at bottom — deliberately unlike the previous section cards */
.audit-workflow-head::after {
    content: "" !important;
    position: absolute !important;
    left: 0 !important;
    right: 0 !important;
    bottom: 0 !important;
    height: 2px !important;
    background: linear-gradient(
        90deg,
        #7D4FFE 0%,
        #C49FFF 46%,
        rgba(196,159,255,0) 100%
    ) !important;
}

/* Step number */
.audit-workflow-num {
    flex: 0 0 auto !important;

    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;

    width: 42px !important;
    height: 42px !important;
    border-radius: 12px !important;

    background:
        linear-gradient(135deg, #7D4FFE 0%, #A86DFF 100%) !important;

    color: #FFFFFF !important;
    font-family: "Inter", system-ui, sans-serif !important;
    font-size: .84rem !important;
    line-height: 1 !important;
    font-weight: 850 !important;
    letter-spacing: .04em !important;

    border: 1px solid rgba(255,255,255,.18) !important;
    box-shadow:
        0 8px 20px rgba(125,79,254,.24),
        inset 0 1px 0 rgba(255,255,255,.16) !important;
}

/* Exactly the same title dimensions for steps 1, 2, 3, 4 and 5 */
.audit-workflow-head h2 {
    margin: 0 !important;
    padding: 0 !important;

    color: #FFFFFF !important;
    font-family: "Inter", system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
    font-size: clamp(1.22rem, 1.5vw, 1.52rem) !important;
    line-height: 1.22 !important;
    font-weight: 760 !important;
    letter-spacing: -0.022em !important;
}

/* Subtle hover only */
.audit-workflow-head {
    transition:
        border-color .18s ease,
        background .18s ease,
        box-shadow .18s ease !important;
}

.audit-workflow-head:hover {
    border-color: rgba(196,159,255,.34) !important;
    background:
        linear-gradient(
            90deg,
            rgba(125,79,254,.18) 0%,
            rgba(255,255,255,.08) 24%,
            rgba(255,255,255,.045) 100%
        ) !important;
    box-shadow:
        0 12px 30px rgba(0,0,0,.20),
        inset 0 1px 0 rgba(255,255,255,.10) !important;
}

@media (max-width: 760px) {
    .audit-workflow-head {
        min-height: 62px !important;
        gap: .75rem !important;
        padding: .72rem .85rem !important;
        border-radius: 13px !important;
    }

    .audit-workflow-num {
        width: 38px !important;
        height: 38px !important;
        border-radius: 10px !important;
        font-size: .78rem !important;
    }

    .audit-workflow-head h2 {
        font-size: 1.12rem !important;
    }
}

@media (prefers-reduced-motion: reduce) {
    .audit-workflow-head {
        transition: none !important;
    }
}


/* =========================================================
   V8.5.27 — REFINED WORKFLOW TIMELINE
   Plus premium, plus léger, même taille pour les étapes 1→5.
   ========================================================= */

.audit-workflow-head {
    position: relative !important;
    overflow: visible !important;

    display: flex !important;
    align-items: center !important;
    gap: 1rem !important;

    width: 100% !important;
    min-height: 64px !important;
    margin: 1.45rem 0 .9rem 0 !important;
    padding: .72rem 1.2rem .72rem .78rem !important;

    border: 1px solid rgba(255,255,255,.10) !important;
    border-radius: 17px !important;

    background:
        linear-gradient(
            105deg,
            rgba(255,255,255,.075) 0%,
            rgba(255,255,255,.045) 55%,
            rgba(125,79,254,.10) 100%
        ) !important;

    box-shadow:
        0 10px 26px rgba(0,0,0,.15),
        inset 0 1px 0 rgba(255,255,255,.08) !important;

    backdrop-filter: blur(16px) saturate(130%) !important;
    -webkit-backdrop-filter: blur(16px) saturate(130%) !important;

    transition:
        transform .2s ease,
        border-color .2s ease,
        box-shadow .2s ease !important;
}

/* Slim accent line */
.audit-workflow-head::before {
    content: "" !important;
    position: absolute !important;
    left: 0 !important;
    top: 13px !important;
    bottom: 13px !important;
    width: 3px !important;
    border-radius: 0 999px 999px 0 !important;
    background: linear-gradient(180deg, #C49FFF, #7D4FFE) !important;
    box-shadow: 0 0 14px rgba(125,79,254,.30) !important;
}

/* Remove the old bottom progress line */
.audit-workflow-head::after {
    display: none !important;
}

/* Step number = elegant glass circle */
.audit-workflow-num {
    flex: 0 0 auto !important;

    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;

    width: 44px !important;
    height: 44px !important;
    border-radius: 50% !important;

    background:
        radial-gradient(circle at 30% 25%, rgba(255,255,255,.22), transparent 30%),
        linear-gradient(145deg, #8A57FF 0%, #6D3FE0 100%) !important;

    color: #FFFFFF !important;
    font-family: "Inter", system-ui, sans-serif !important;
    font-size: .82rem !important;
    line-height: 1 !important;
    font-weight: 850 !important;
    letter-spacing: .045em !important;

    border: 1px solid rgba(255,255,255,.18) !important;

    box-shadow:
        0 8px 22px rgba(125,79,254,.23),
        inset 0 1px 0 rgba(255,255,255,.18) !important;
}

/* Same title size for all five steps */
.audit-workflow-head h2 {
    margin: 0 !important;
    padding: 0 !important;

    color: #FFFFFF !important;
    font-family: "Inter", system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
    font-size: clamp(1.22rem, 1.45vw, 1.48rem) !important;
    line-height: 1.2 !important;
    font-weight: 740 !important;
    letter-spacing: -0.025em !important;
}

/* Small glow on the far right */
.audit-workflow-head > h2::after {
    content: "" !important;
    position: absolute !important;
    right: 20px !important;
    top: 50% !important;
    width: 44px !important;
    height: 4px !important;
    border-radius: 999px !important;
    transform: translateY(-50%) !important;
    background: linear-gradient(90deg, rgba(196,159,255,.18), rgba(125,79,254,.75)) !important;
    opacity: .72 !important;
}

.audit-workflow-head:hover {
    transform: translateY(-1px) !important;
    border-color: rgba(196,159,255,.22) !important;
    box-shadow:
        0 14px 32px rgba(0,0,0,.19),
        inset 0 1px 0 rgba(255,255,255,.10) !important;
}

/* Make each number visually distinct but still coherent */
.audit-workflow-head:nth-of-type(1) .audit-workflow-num {
    background: linear-gradient(145deg, #7D4FFE, #A06BFF) !important;
}
.audit-workflow-head:nth-of-type(2) .audit-workflow-num {
    background: linear-gradient(145deg, #755BFF, #8E73FF) !important;
}
.audit-workflow-head:nth-of-type(3) .audit-workflow-num {
    background: linear-gradient(145deg, #6F64F8, #8A80FF) !important;
}
.audit-workflow-head:nth-of-type(4) .audit-workflow-num {
    background: linear-gradient(145deg, #7850E8, #A173FF) !important;
}
.audit-workflow-head:nth-of-type(5) .audit-workflow-num {
    background: linear-gradient(145deg, #7246D8, #9B63F0) !important;
}

@media (max-width: 760px) {
    .audit-workflow-head {
        min-height: 60px !important;
        gap: .78rem !important;
        padding: .68rem .85rem .68rem .72rem !important;
        border-radius: 14px !important;
    }

    .audit-workflow-num {
        width: 38px !important;
        height: 38px !important;
        font-size: .75rem !important;
    }

    .audit-workflow-head h2 {
        font-size: 1.08rem !important;
    }

    .audit-workflow-head > h2::after {
        display: none !important;
    }
}

@media (prefers-reduced-motion: reduce) {
    .audit-workflow-head {
        transition: none !important;
    }
    .audit-workflow-head:hover {
        transform: none !important;
    }
}


/* =========================================================
   V8.5.27 — TRAIN BUTTON TEXT IN BLACK
   Keep the current visual style, change only the writing color.
   ========================================================= */

.st-key-v82_train_models_advanced button,
.st-key-v82_train_models_advanced button *,
.st-key-v82_train_models_advanced [data-testid="stButton"] button,
.st-key-v82_train_models_advanced [data-testid="stButton"] button * {
    color: #111111 !important;
    fill: #111111 !important;
}

/* Keep readability on hover/focus/disabled states too */
.st-key-v82_train_models_advanced button:hover,
.st-key-v82_train_models_advanced button:hover *,
.st-key-v82_train_models_advanced button:focus,
.st-key-v82_train_models_advanced button:focus *,
.st-key-v82_train_models_advanced button:disabled,
.st-key-v82_train_models_advanced button:disabled * {
    color: #111111 !important;
    fill: #111111 !important;
}


/* =========================================================
   V8.5.27 — MESH GRADIENT ANIMÉ
   Plusieurs zones violet foncé se déplacent et se mélangent.
   Background only — aucun objet visible.
   ========================================================= */

[data-testid="stAppViewContainer"] {
    position: relative !important;

    background:
        radial-gradient(
            78% 72% at 8% 18%,
            rgba(92, 54, 160, .34) 0%,
            rgba(72, 41, 120, .20) 34%,
            transparent 72%
        ),
        radial-gradient(
            72% 68% at 88% 12%,
            rgba(111, 63, 171, .28) 0%,
            rgba(77, 42, 128, .15) 38%,
            transparent 74%
        ),
        radial-gradient(
            82% 78% at 78% 88%,
            rgba(78, 47, 145, .32) 0%,
            rgba(57, 34, 103, .18) 38%,
            transparent 74%
        ),
        radial-gradient(
            76% 70% at 16% 84%,
            rgba(128, 75, 157, .19) 0%,
            rgba(82, 45, 106, .11) 40%,
            transparent 72%
        ),
        radial-gradient(
            64% 58% at 52% 48%,
            rgba(105, 63, 165, .16) 0%,
            transparent 70%
        ),
        linear-gradient(
            135deg,
            #171219 0%,
            #1C1422 26%,
            #22152D 50%,
            #1F1629 72%,
            #161218 100%
        ) !important;

    background-size:
        170% 165%,
        175% 170%,
        180% 175%,
        170% 180%,
        160% 160%,
        100% 100% !important;

    background-position:
        0% 0%,
        100% 0%,
        100% 100%,
        0% 100%,
        50% 50%,
        0 0 !important;

    animation: auditMeshShift 38s cubic-bezier(.45,0,.25,1) infinite alternate !important;
}

/* Le mesh reste diffus : aucun cercle ni objet distinct. */
[data-testid="stAppViewContainer"]::before {
    content: "" !important;
    position: fixed !important;
    inset: -18% !important;
    z-index: 0 !important;
    pointer-events: none !important;

    background:
        linear-gradient(
            115deg,
            rgba(125,79,254,.08) 0%,
            rgba(196,159,255,.045) 28%,
            rgba(70,43,110,.06) 52%,
            rgba(112,62,155,.055) 72%,
            rgba(125,79,254,.06) 100%
        ) !important;

    filter: blur(72px) !important;
    opacity: .78 !important;
    transform: translate3d(0,0,0) scale(1.10) !important;
    animation: auditMeshVeil 46s ease-in-out infinite alternate !important;
}

/* Keep the dashboard above the animated background. */
[data-testid="stAppViewContainer"] > .main,
[data-testid="stAppViewContainer"] > section,
[data-testid="stAppViewContainer"] > div:not(:first-child) {
    position: relative !important;
    z-index: 1 !important;
}

@keyframes auditMeshShift {
    0% {
        background-position:
            0% 4%,
            98% 0%,
            100% 96%,
            2% 100%,
            48% 46%,
            0 0;
        background-size:
            170% 165%,
            175% 170%,
            180% 175%,
            170% 180%,
            160% 160%,
            100% 100%;
    }

    25% {
        background-position:
            14% 12%,
            86% 8%,
            90% 84%,
            12% 88%,
            56% 42%,
            0 0;
    }

    50% {
        background-position:
            22% 6%,
            78% 18%,
            78% 76%,
            22% 82%,
            46% 58%,
            0 0;
        background-size:
            185% 175%,
            165% 185%,
            172% 165%,
            184% 168%,
            170% 170%,
            100% 100%;
    }

    75% {
        background-position:
            10% 22%,
            90% 12%,
            86% 68%,
            30% 90%,
            58% 54%,
            0 0;
    }

    100% {
        background-position:
            18% 8%,
            82% 22%,
            72% 84%,
            8% 72%,
            42% 44%,
            0 0;
        background-size:
            176% 182%,
            184% 172%,
            168% 184%,
            176% 172%,
            164% 176%,
            100% 100%;
    }
}

@keyframes auditMeshVeil {
    0% {
        transform: translate3d(-3%, -2%, 0) scale(1.08) rotate(-1deg) !important;
        opacity: .65;
    }

    50% {
        transform: translate3d(3%, 3%, 0) scale(1.14) rotate(1.5deg) !important;
        opacity: .88;
    }

    100% {
        transform: translate3d(-1%, 5%, 0) scale(1.11) rotate(-.5deg) !important;
        opacity: .72;
    }
}

@media (prefers-reduced-motion: reduce) {
    [data-testid="stAppViewContainer"],
    [data-testid="stAppViewContainer"]::before {
        animation: none !important;
    }
}



/* =========================================================
   V8.5.27 — TEXTES EXPLICATIFS CIBLÉS EN BLANC
   ========================================================= */

.audit-caption-white,
p.audit-caption-white {
    color: #FFFFFF !important;
    opacity: 1 !important;
    font-family: "Inter", system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
    line-height: 1.55 !important;
    margin: .35rem 0 .65rem 0 !important;
}


/* =========================================================
   V8.5.27 — PREMIER AUDIT / SANS HISTORIQUE EN NOIR
   ========================================================= */

.audit-first-audit-black,
.audit-first-audit-black *,
.audit-first-audit-black b,
.audit-first-audit-black strong,
.audit-first-audit-black span {
    color: #111111 !important;
    opacity: 1 !important;
    text-shadow: none !important;
}


/* =========================================================
   V8.5.27 — PREMIER AUDIT : TEXTE NOIR FORCÉ
   ========================================================= */
.audit-info-box.audit-first-audit-black,
.audit-info-box.audit-first-audit-black *,
.audit-info-box.audit-first-audit-black b,
.audit-info-box.audit-first-audit-black span {
    color: #111111 !important;
    opacity: 1 !important;
    text-shadow: none !important;
}


/* =========================================================
   V8.5.28 — CONTRASTE STABLE STREAMLIT LIGHT / DARK
   L'application conserve son univers sombre quel que soit le choix du
   visiteur dans le menu Streamlit. Les composants natifs reçoivent donc
   des surfaces et couleurs explicites au lieu de dépendre du thème actif.
   ========================================================= */

html,
body {
    color-scheme: dark !important;
}

/* Texte éditorial posé directement sur le fond sombre. */
[data-testid="stAppViewContainer"] .main p,
[data-testid="stAppViewContainer"] .main li,
[data-testid="stAppViewContainer"] .main label,
[data-testid="stAppViewContainer"] .main [data-testid="stMarkdownContainer"],
[data-testid="stAppViewContainer"] .main [data-testid="stMarkdownContainer"] p,
[data-testid="stAppViewContainer"] .main [data-testid="stMarkdownContainer"] li {
    color: #F7F3FB !important;
    opacity: 1 !important;
}

/* Légendes, aides et texte sous la barre de progression. */
[data-testid="stCaptionContainer"],
[data-testid="stCaptionContainer"] *,
.stCaption,
.stCaption *,
[data-testid="stProgress"] p,
[data-testid="stProgress"] span,
[data-testid="stProgress"] label {
    color: #EDE6F5 !important;
    opacity: 1 !important;
}

[data-testid="stProgress"] [role="progressbar"],
[data-testid="stProgress"] > div > div {
    background-color: rgba(255,255,255,.16) !important;
}

[data-testid="stProgress"] [role="progressbar"] > div,
[data-testid="stProgress"] > div > div > div {
    background: linear-gradient(90deg, #8A57FF, #C49FFF) !important;
}

/* Tous les boutons restent lisibles, y compris l'entraînement désactivé. */
.stButton > button,
.stButton > button *,
.stDownloadButton > button,
.stDownloadButton > button *,
[data-testid="stFormSubmitButton"] > button,
[data-testid="stFormSubmitButton"] > button *,
.st-key-v82_train_models_advanced button,
.st-key-v82_train_models_advanced button * {
    color: #FFFFFF !important;
    fill: #FFFFFF !important;
    opacity: 1 !important;
}

.stButton > button:disabled,
.stDownloadButton > button:disabled,
[data-testid="stFormSubmitButton"] > button:disabled,
.st-key-v82_train_models_advanced button:disabled {
    color: #F7F2FB !important;
    background: linear-gradient(180deg, #655572, #4C3F58) !important;
    border-color: rgba(255,255,255,.24) !important;
    opacity: 1 !important;
}

.stButton > button:disabled *,
.stDownloadButton > button:disabled *,
[data-testid="stFormSubmitButton"] > button:disabled *,
.st-key-v82_train_models_advanced button:disabled * {
    color: #F7F2FB !important;
    fill: #F7F2FB !important;
    opacity: 1 !important;
}

/* Champs, sélecteurs et listes : toujours sombre + texte clair. */
[data-baseweb="input"] > div,
[data-baseweb="textarea"] > div,
[data-baseweb="select"] > div,
.stTextInput input,
.stNumberInput input,
.stTextArea textarea,
.stDateInput input,
.stTimeInput input {
    background: #251C2D !important;
    border-color: rgba(196,159,255,.30) !important;
    color: #FFFFFF !important;
}

[data-baseweb="input"] input,
[data-baseweb="textarea"] textarea,
[data-baseweb="select"] input,
[data-baseweb="select"] span,
[data-baseweb="select"] div,
.stTextInput input,
.stNumberInput input,
.stTextArea textarea,
.stDateInput input,
.stTimeInput input {
    color: #FFFFFF !important;
    -webkit-text-fill-color: #FFFFFF !important;
    opacity: 1 !important;
}

[data-baseweb="input"] input::placeholder,
[data-baseweb="textarea"] textarea::placeholder,
.stTextInput input::placeholder,
.stTextArea textarea::placeholder {
    color: #D8CFE1 !important;
    -webkit-text-fill-color: #D8CFE1 !important;
    opacity: 1 !important;
}

[data-baseweb="popover"],
[data-baseweb="menu"],
[role="listbox"] {
    background: #211827 !important;
    border-color: rgba(196,159,255,.26) !important;
}

[role="option"],
[role="option"] *,
[data-baseweb="menu"] *,
[role="listbox"] * {
    color: #FFFFFF !important;
    opacity: 1 !important;
}

[role="option"]:hover,
[role="option"][aria-selected="true"] {
    background: #563A78 !important;
}

/* Radios, cases et libellés de formulaires. */
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] *,
[data-testid="stRadio"] label,
[data-testid="stRadio"] label *,
[data-testid="stCheckbox"] label,
[data-testid="stCheckbox"] label *,
[data-testid="stToggle"] label,
[data-testid="stToggle"] label * {
    color: #F5EFFA !important;
    opacity: 1 !important;
}

/* Alertes Streamlit : surface foncée explicite, quel que soit le thème. */
[data-testid="stAlert"],
div[data-baseweb="notification"] {
    background: linear-gradient(135deg, #624B69, #4E3D5A) !important;
    border: 1px solid rgba(255,255,255,.18) !important;
}

[data-testid="stAlert"],
[data-testid="stAlert"] *,
div[data-baseweb="notification"],
div[data-baseweb="notification"] * {
    color: #FFFFFF !important;
    fill: #FFFFFF !important;
    opacity: 1 !important;
}

/* Info-bulles : choix assumé noir sur blanc, stable dans les deux modes. */
[role="tooltip"],
[role="tooltip"] > div,
[role="tooltip"] [data-testid="stMarkdownContainer"],
[data-baseweb="tooltip"] {
    background: #FFFFFF !important;
    color: #17111D !important;
    border-color: rgba(125,79,254,.24) !important;
}

[role="tooltip"] *,
[data-baseweb="tooltip"] * {
    color: #17111D !important;
    -webkit-text-fill-color: #17111D !important;
    opacity: 1 !important;
}

/* Onglets et expanders. */
[data-baseweb="tab"],
[data-baseweb="tab"] *,
button[role="tab"],
button[role="tab"] *,
[data-testid="stExpander"] summary,
[data-testid="stExpander"] summary * {
    color: #EEE7F5 !important;
    opacity: 1 !important;
}

[data-baseweb="tab"][aria-selected="true"],
[data-baseweb="tab"][aria-selected="true"] *,
button[role="tab"][aria-selected="true"],
button[role="tab"][aria-selected="true"] * {
    color: #FFFFFF !important;
}

/* Les cartes volontairement claires gardent leur texte sombre. */
.audit-requested-black,
.audit-requested-black *,
.audit-light-surface,
.audit-light-surface *,
.audit-white-card,
.audit-white-card *,
.audit-alert-card,
.audit-alert-card *,
.audit-warning,
.audit-warning *,
.audit-info-box.audit-first-audit-black,
.audit-info-box.audit-first-audit-black * {
    color: #111111 !important;
    -webkit-text-fill-color: #111111 !important;
    opacity: 1 !important;
    text-shadow: none !important;
}

</style>
""",
    unsafe_allow_html=True,
)


# ============================================================
# 3. FONCTIONS TECHNIQUES
# ============================================================

def get_conn(host, port, dbname, user, password):
    return psycopg2.connect(
        host=host,
        port=int(port),
        dbname=dbname,
        user=user,
        password=password,
        connect_timeout=5,
    )


@st.cache_data(show_spinner=False, ttl=30)
def read_sql_cached(query, host, port, dbname, user, password, params=None):
    with get_conn(host, port, dbname, user, password) as conn:
        return pd.read_sql_query(query, conn, params=params)


def execute_generation(host, port, dbname, user, password, target_code, source_code):
    with get_conn(host, port, dbname, user, password) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM auditprep.fn_generate_smart_checklist(%s, %s);",
                (target_code, source_code),
            )
            rows = cur.fetchall()
            cols = [d[0] for d in cur.description]
        conn.commit()
    return pd.DataFrame(rows, columns=cols)


# ============================================================
# 3B. FONCTIONS V7 - ALIMENTATION DES DONNÉES DEPUIS L'APP
# ============================================================

@st.cache_data(show_spinner=False, ttl=60)
def get_table_columns_cached(host, port, dbname, user, password, table_name, schema_name="auditprep"):
    query = """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = %s AND table_name = %s
        ORDER BY ordinal_position;
    """
    with get_conn(host, port, dbname, user, password) as conn:
        return pd.read_sql_query(query, conn, params=(schema_name, table_name))["column_name"].tolist()


@st.cache_data(show_spinner=False, ttl=60)
def get_reference_labels_cached(host, port, dbname, user, password, table_name):
    """Lit les libellés d'une table de référence PostgreSQL.

    V7.5 : les listes déroulantes de l'interface ne sont plus inventées dans le code.
    Elles reprennent les vraies valeurs présentes dans PostgreSQL, colonne label.
    """
    query = f"""
        SELECT label
        FROM auditprep.{table_name}
        ORDER BY 1;
    """
    with get_conn(host, port, dbname, user, password) as conn:
        df = pd.read_sql_query(query, conn)
    return [str(x) for x in df["label"].dropna().tolist()]


def table_exists(conn, table_name, schema_name="auditprep"):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT EXISTS (
                SELECT 1
                FROM information_schema.tables
                WHERE table_schema = %s AND table_name = %s
            );
            """,
            (schema_name, table_name),
        )
        return bool(cur.fetchone()[0])


def table_columns(conn, table_name, schema_name="auditprep"):
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = %s AND table_name = %s
            ORDER BY ordinal_position;
            """,
            (schema_name, table_name),
        )
        return [r[0] for r in cur.fetchall()]


def pick_col(cols, candidates):
    for c in candidates:
        if c in cols:
            return c
    return None


def normalize_blank(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    value = str(value).strip()
    return value if value else None


def to_pg_value(value):
    """Convertit les types pandas/numpy en types Python acceptés par psycopg2.

    Correction V7.8.1 : Streamlit récupère parfois mission_id depuis un DataFrame
    sous forme numpy.int64. psycopg2 ne sait pas adapter directement ce type,
    d'où l'erreur : can't adapt type 'numpy.int64'.
    """
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime()
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return value


def execute_dynamic_insert(conn, table_name, values, schema_name="auditprep", returning_col=None):
    values = {k: v for k, v in values.items() if v is not None}
    if not values:
        raise ValueError(f"Aucune valeur à insérer dans {table_name}.")

    cols = list(values.keys())
    query = sql.SQL("INSERT INTO {}.{} ({}) VALUES ({})").format(
        sql.Identifier(schema_name),
        sql.Identifier(table_name),
        sql.SQL(", ").join(sql.Identifier(c) for c in cols),
        sql.SQL(", ").join(sql.Placeholder() for _ in cols),
    )
    if returning_col:
        query += sql.SQL(" RETURNING {}").format(sql.Identifier(returning_col))

    with conn.cursor() as cur:
        cur.execute(query, [to_pg_value(values[c]) for c in cols])
        if returning_col:
            return cur.fetchone()[0]
    return None


def get_or_create_simple(conn, table_name, id_candidates, value_candidates, value, extra_values=None):
    value = normalize_blank(value)
    if not value or not table_exists(conn, table_name):
        return None

    cols = table_columns(conn, table_name)
    id_col = pick_col(cols, id_candidates)
    value_col = pick_col(cols, value_candidates)
    if not id_col or not value_col:
        return None

    with conn.cursor() as cur:
        cur.execute(
            sql.SQL("SELECT {} FROM auditprep.{} WHERE LOWER({}::text) = LOWER(%s) LIMIT 1").format(
                sql.Identifier(id_col), sql.Identifier(table_name), sql.Identifier(value_col)
            ),
            (value,),
        )
        row = cur.fetchone()
        if row:
            return row[0]

    insert_values = {value_col: value}
    if extra_values:
        for k, v in extra_values.items():
            if k in cols:
                insert_values[k] = v
    return execute_dynamic_insert(conn, table_name, insert_values, returning_col=id_col)


def get_standard_id(conn, standard_label):
    if not table_exists(conn, "standards"):
        return None
    cols = table_columns(conn, "standards")
    id_col = pick_col(cols, ["standard_id", "id"])
    if not id_col:
        return None

    label = normalize_blank(standard_label) or "ISO 9001:2015"
    code = "ISO 9001"
    version = "2015"
    if ":" in label:
        code, version = [x.strip() for x in label.split(":", 1)]
    elif label:
        code = label.strip()

    code_col = pick_col(cols, ["standard_code", "code"])
    version_col = pick_col(cols, ["standard_version", "version"])
    label_col = pick_col(cols, ["standard_label", "label", "standard_name", "name"])

    with conn.cursor() as cur:
        if code_col and version_col:
            cur.execute(
                sql.SQL("SELECT {} FROM auditprep.standards WHERE LOWER({}::text)=LOWER(%s) AND LOWER({}::text)=LOWER(%s) LIMIT 1").format(
                    sql.Identifier(id_col), sql.Identifier(code_col), sql.Identifier(version_col)
                ),
                (code, version),
            )
            row = cur.fetchone()
            if row:
                return row[0]
        if label_col:
            cur.execute(
                sql.SQL("SELECT {} FROM auditprep.standards WHERE LOWER({}::text)=LOWER(%s) LIMIT 1").format(
                    sql.Identifier(id_col), sql.Identifier(label_col)
                ),
                (label,),
            )
            row = cur.fetchone()
            if row:
                return row[0]
        cur.execute(sql.SQL("SELECT {} FROM auditprep.standards LIMIT 1").format(sql.Identifier(id_col)))
        row = cur.fetchone()
        return row[0] if row else None


def get_clause_id(conn, clause_code, clause_title=None, standard_id=None):
    clause_code = normalize_blank(clause_code)
    if not clause_code or not table_exists(conn, "standard_clauses"):
        return None
    cols = table_columns(conn, "standard_clauses")
    id_col = pick_col(cols, ["clause_id", "id"])
    code_col = pick_col(cols, ["clause_code", "code"])
    if not id_col or not code_col:
        return None

    with conn.cursor() as cur:
        cur.execute(
            sql.SQL("SELECT {} FROM auditprep.standard_clauses WHERE {}::text = %s LIMIT 1").format(
                sql.Identifier(id_col), sql.Identifier(code_col)
            ),
            (clause_code,),
        )
        row = cur.fetchone()
        if row:
            return row[0]

    values = {code_col: clause_code}
    title_col = pick_col(cols, ["clause_title", "title", "label"])
    if title_col:
        values[title_col] = normalize_blank(clause_title) or f"Clause {clause_code}"
    standard_col = pick_col(cols, ["standard_id"])
    if standard_col and standard_id:
        values[standard_col] = standard_id
    return execute_dynamic_insert(conn, "standard_clauses", values, returning_col=id_col)


def get_finding_type_id(conn, finding_type):
    if not table_exists(conn, "finding_types"):
        return None
    ft = str(finding_type or "").lower()
    if "non" in ft or "nc" in ft:
        code = "NC"
    elif "rem" in ft or "rq" in ft:
        code = "RQ"
    elif "am" in ft or "amélior" in ft or "amelior" in ft:
        code = "AM"
    else:
        code = "RQ"

    cols = table_columns(conn, "finding_types")
    id_col = pick_col(cols, ["finding_type_id", "id"])
    code_col = pick_col(cols, ["code", "finding_type_code"])
    label_col = pick_col(cols, ["label", "finding_type_label", "name"])
    if not id_col:
        return None

    with conn.cursor() as cur:
        if code_col:
            cur.execute(
                sql.SQL("SELECT {} FROM auditprep.finding_types WHERE UPPER({}::text)=UPPER(%s) LIMIT 1").format(
                    sql.Identifier(id_col), sql.Identifier(code_col)
                ),
                (code,),
            )
            row = cur.fetchone()
            if row:
                return row[0]
        if label_col:
            cur.execute(
                sql.SQL("SELECT {} FROM auditprep.finding_types WHERE LOWER({}::text) LIKE LOWER(%s) LIMIT 1").format(
                    sql.Identifier(id_col), sql.Identifier(label_col)
                ),
                (f"%{finding_type}%",),
            )
            row = cur.fetchone()
            if row:
                return row[0]
    return None


def get_id_from_label_table(conn, table_name, id_column, label_value):
    """Retourne l'ID correspondant à un label dans une table de référence.

    Correction V7.5 : dans ta base réelle, audit_types et mission_statuses
    utilisent la colonne label. Cette fonction évite les NULL sur les clés
    étrangères audit_type_id et mission_status_id.
    """
    label_value = normalize_blank(label_value)
    if not label_value or not table_exists(conn, table_name):
        return None

    cols = table_columns(conn, table_name)
    if id_column not in cols or "label" not in cols:
        return None

    with conn.cursor() as cur:
        # Recherche exacte sur le vrai libellé PostgreSQL.
        cur.execute(
            sql.SQL(
                "SELECT {} FROM auditprep.{} WHERE LOWER(label) = LOWER(%s) LIMIT 1"
            ).format(sql.Identifier(id_column), sql.Identifier(table_name)),
            (label_value,),
        )
        row = cur.fetchone()
        if row:
            return row[0]

        # Recherche souple, utile si l'utilisateur saisit une variante proche.
        cur.execute(
            sql.SQL(
                "SELECT {} FROM auditprep.{} WHERE LOWER(label) LIKE LOWER(%s) LIMIT 1"
            ).format(sql.Identifier(id_column), sql.Identifier(table_name)),
            (f"%{label_value}%",),
        )
        row = cur.fetchone()
        if row:
            return row[0]

        # Dernier filet de sécurité : prendre le premier ID disponible.
        # Cela évite que PostgreSQL reçoive NULL sur une colonne NOT NULL.
        cur.execute(
            sql.SQL("SELECT {} FROM auditprep.{} ORDER BY {} LIMIT 1").format(
                sql.Identifier(id_column),
                sql.Identifier(table_name),
                sql.Identifier(id_column),
            )
        )
        row = cur.fetchone()
        return row[0] if row else None


def get_audit_type_id(conn, audit_type_label):
    return get_id_from_label_table(
        conn,
        table_name="audit_types",
        id_column="audit_type_id",
        label_value=audit_type_label or "Audit interne",
    )


def get_mission_status_id(conn, status_label):
    return get_id_from_label_table(
        conn,
        table_name="mission_statuses",
        id_column="mission_status_id",
        label_value=status_label or "Brouillon",
    )


# ============================================================
# 3C. V7.8 - CONTEXTE MÉTIER PERSISTANT ET LOTS CONTEXTUALISÉS
# ============================================================

def ensure_mission_contexts_table(conn):
    """Crée la table métier si elle n'existe pas.

    Cette table évite de modifier brutalement audit_missions. Elle stocke le
    contexte nécessaire pour comparer les audits entre eux : secteur, processus,
    objectifs, périmètre, exigences, risques et mots-clés.
    """
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS auditprep.mission_contexts (
                mission_id INTEGER PRIMARY KEY
                    REFERENCES auditprep.audit_missions(mission_id)
                    ON DELETE CASCADE,
                sector TEXT,
                audited_process TEXT,
                audit_objective TEXT,
                audit_scope TEXT,
                specific_requirements TEXT,
                known_risks TEXT,
                keywords TEXT,
                created_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW()
            );
            """
        )


def upsert_mission_context(conn, mission_id, context_data):
    """Enregistre ou met à jour le contexte métier d'une mission."""
    mission_id = to_pg_value(mission_id)
    if mission_id is None:
        return
    mission_id = int(mission_id)

    ensure_mission_contexts_table(conn)

    fields = {
        "sector": normalize_blank(context_data.get("sector")),
        "audited_process": normalize_blank(context_data.get("audited_process")),
        "audit_objective": normalize_blank(context_data.get("audit_objective")),
        "audit_scope": normalize_blank(context_data.get("audit_scope")),
        "specific_requirements": normalize_blank(context_data.get("specific_requirements")),
        "known_risks": normalize_blank(context_data.get("known_risks")),
        "keywords": normalize_blank(context_data.get("keywords")),
    }

    # On évite d'écraser avec une ligne totalement vide.
    if not any(fields.values()):
        return

    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO auditprep.mission_contexts (
                mission_id,
                sector,
                audited_process,
                audit_objective,
                audit_scope,
                specific_requirements,
                known_risks,
                keywords,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
            ON CONFLICT (mission_id) DO UPDATE SET
                sector = EXCLUDED.sector,
                audited_process = EXCLUDED.audited_process,
                audit_objective = EXCLUDED.audit_objective,
                audit_scope = EXCLUDED.audit_scope,
                specific_requirements = EXCLUDED.specific_requirements,
                known_risks = EXCLUDED.known_risks,
                keywords = EXCLUDED.keywords,
                updated_at = NOW();
            """,
            (
                mission_id,
                fields["sector"],
                fields["audited_process"],
                fields["audit_objective"],
                fields["audit_scope"],
                fields["specific_requirements"],
                fields["known_risks"],
                fields["keywords"],
            ),
        )


@st.cache_data(show_spinner=False, ttl=30)
def read_mission_contexts_cached(host, port, dbname, user, password):
    """Charge les contextes métier existants.

    Si la table n'existe pas encore, elle est créée automatiquement.
    """
    with get_conn(host, port, dbname, user, password) as conn:
        ensure_mission_contexts_table(conn)
        query = """
            SELECT
                mission_id,
                sector,
                audited_process,
                audit_objective,
                audit_scope,
                specific_requirements,
                known_risks,
                keywords,
                created_at,
                updated_at
            FROM auditprep.mission_contexts;
        """
        return pd.read_sql_query(query, conn)


def context_for_mission(contexts_df, mission_id):
    """Retourne un dictionnaire de contexte pour une mission donnée."""
    default = {
        "sector": "",
        "audited_process": "",
        "audit_objective": "",
        "audit_scope": "",
        "specific_requirements": "",
        "known_risks": "",
        "keywords": "",
    }
    if contexts_df is None or contexts_df.empty or mission_id is None:
        return default
    try:
        filtered = contexts_df[contexts_df["mission_id"].astype(str) == str(mission_id)]
        if filtered.empty:
            return default
        row = filtered.iloc[0]
        return {k: normalize_blank(row.get(k)) or "" for k in default.keys()}
    except Exception:
        return default


def enrich_missions_with_context(missions_df, contexts_df):
    """Ajoute les colonnes de contexte métier aux DataFrames missions."""
    if missions_df is None or missions_df.empty:
        return missions_df
    out = missions_df.copy()
    if contexts_df is None or contexts_df.empty or "mission_id" not in out.columns:
        for col in ["sector", "audited_process", "audit_objective", "audit_scope", "specific_requirements", "known_risks", "keywords"]:
            if col not in out.columns:
                out[col] = ""
        return out
    ctx_cols = ["mission_id", "sector", "audited_process", "audit_objective", "audit_scope", "specific_requirements", "known_risks", "keywords"]
    ctx = contexts_df[[c for c in ctx_cols if c in contexts_df.columns]].copy()
    out = out.merge(ctx, on="mission_id", how="left")
    for col in ["sector", "audited_process", "audit_objective", "audit_scope", "specific_requirements", "known_risks", "keywords"]:
        if col not in out.columns:
            out[col] = ""
        else:
            out[col] = out[col].fillna("")
    return out


def save_context_for_existing_mission(host, port, dbname, user, password, mission_id, context_data):
    """Bouton manuel : sauvegarde le contexte métier de la mission sélectionnée."""
    with get_conn(host, port, dbname, user, password) as conn:
        upsert_mission_context(conn, mission_id, context_data)
        conn.commit()


# ============================================================
# 3D. V8.3 - ÉTIQUETTES EXPERTES ET DONNÉES SUPERVISÉES
# ============================================================

def normalize_priority_label(value, finding_type=""):
    """Normalise une étiquette historique vers Faible / Moyenne / Haute."""
    raw = str(value or "").strip().lower()
    for old, new in {"é": "e", "è": "e", "ê": "e", "à": "a", "ç": "c"}.items():
        raw = raw.replace(old, new)
    if any(word in raw for word in ["haut", "elev", "critique", "majeur", "grave"]):
        return "Haute"
    if any(word in raw for word in ["moy", "moder", "signific"]):
        return "Moyenne"
    if any(word in raw for word in ["faibl", "mineur", "leger"]):
        return "Faible"

    finding_raw = str(finding_type or "").strip().lower()
    for old, new in {"é": "e", "è": "e", "ê": "e", "à": "a", "ç": "c"}.items():
        finding_raw = finding_raw.replace(old, new)
    if "non-conform" in finding_raw or "non conform" in finding_raw:
        return "Haute"
    if "remarque" in finding_raw or "observation" in finding_raw:
        return "Moyenne"
    if "amelior" in finding_raw or "opportun" in finding_raw:
        return "Faible"
    return None


def priority_to_score(priority):
    """Score proxy uniquement si aucun score expert historique n'existe."""
    return {"Haute": 85.0, "Moyenne": 60.0, "Faible": 30.0}.get(str(priority), None)


FRENCH_STOP_WORDS = {
    "a", "afin", "ai", "ainsi", "alors", "apres", "au", "aucun", "aucune",
    "aussi", "autre", "aux", "avec", "avoir", "car", "ce", "ces", "cet",
    "cette", "comme", "dans", "de", "des", "doit", "donc", "du", "elle",
    "en", "entre", "est", "et", "etre", "fait", "faire", "il", "ils", "je",
    "la", "le", "les", "leur", "leurs", "lui", "mais", "mes", "meme", "ne", "non",
    "nos", "notre", "nous", "on", "ou", "par", "pas", "plus", "pour", "que",
    "quel", "quelle", "quelles", "quels", "qui", "sa", "sans", "se", "ses",
    "son", "sont", "sur", "tous", "tout", "toute", "toutes", "un", "une",
    "vos", "votre", "vous", "y",
    # Termes retirés tant que des cibles proxy gravité/type subsistent : ils
    # pourraient révéler directement l'étiquette au lieu d'apprendre le risque.
    "amelioration", "conformite", "critique", "elevee", "faible", "gravite",
    "majeur", "mineur", "moyenne", "nonconformite", "observation", "opportunite",
    "remarque", "severite",
}

LEAKAGE_TOKEN_PREFIXES = (
    "amelior", "conformit", "nonconform", "critic", "critiq", "elev", "faibl", "gravit",
    "majeur", "mineur", "moyenn", "observ", "opportun", "remarqu", "sever",
)


def clean_french_model_text(value):
    """Nettoyage reproductible sans dépendance NLTK/spaCy."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char)).lower()
    tokens = re.findall(r"[a-z0-9]+", text)
    useful = [
        token for token in tokens
        if len(token) > 1
        and token not in FRENCH_STOP_WORDS
        and not token.startswith(LEAKAGE_TOKEN_PREFIXES)
    ]
    return " ".join(useful)


def is_expert_label_source(value):
    source = str(value or "").strip().lower()
    return source.startswith("expert") or source == "historique_explicite"


def select_training_dataset(dataset_df):
    """Privilégie les étiquettes expertes dès qu'elles suffisent au split 70/30."""
    if dataset_df is None or dataset_df.empty:
        return pd.DataFrame(), "Aucune donnée", 0
    data = dataset_df.copy()
    expert_mask = data["target_source"].map(is_expert_label_source)
    expert_df = data[expert_mask].copy()
    counts = expert_df["target_priority"].value_counts() if not expert_df.empty else pd.Series(dtype="int64")
    expert_ready = len(expert_df) >= 10 and len(counts) >= 2 and int(counts.min()) >= 2
    if expert_ready:
        return expert_df.reset_index(drop=True), "Données expertes uniquement", int(len(expert_df))
    return data.reset_index(drop=True), "Mode transitoire : experts + proxy", int(len(expert_df))


def ensure_finding_ml_labels_table(conn):
    """Stocke séparément les cibles supervisées sans casser le schéma V6."""
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS auditprep.finding_ml_labels (
                finding_id BIGINT PRIMARY KEY,
                priority_label TEXT,
                criticality_score NUMERIC(5, 2),
                label_source TEXT NOT NULL DEFAULT 'expert',
                validated_by TEXT,
                validation_comment TEXT,
                validated_at TIMESTAMP WITHOUT TIME ZONE,
                created_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW(),
                CHECK (priority_label IS NULL OR priority_label IN ('Faible', 'Moyenne', 'Haute')),
                CHECK (criticality_score IS NULL OR (criticality_score >= 0 AND criticality_score <= 100))
            );
            """
        )
        # Compatibilité avec une table créée par V8.0/V8.1.
        cur.execute("ALTER TABLE auditprep.finding_ml_labels ADD COLUMN IF NOT EXISTS validated_by TEXT;")
        cur.execute("ALTER TABLE auditprep.finding_ml_labels ADD COLUMN IF NOT EXISTS validation_comment TEXT;")
        cur.execute("ALTER TABLE auditprep.finding_ml_labels ADD COLUMN IF NOT EXISTS validated_at TIMESTAMP WITHOUT TIME ZONE;")


def upsert_finding_ml_label(
    conn,
    finding_id,
    priority_label=None,
    criticality_score=None,
    label_source="expert",
    validated_by=None,
    validation_comment=None,
):
    finding_id = to_pg_value(finding_id)
    priority_label = normalize_priority_label(priority_label)
    score_value = to_pg_value(criticality_score)
    if finding_id is None or (priority_label is None and score_value is None):
        return
    if score_value is not None:
        score_value = max(0.0, min(100.0, float(score_value)))
    ensure_finding_ml_labels_table(conn)
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO auditprep.finding_ml_labels (
                finding_id, priority_label, criticality_score, label_source,
                validated_by, validation_comment, validated_at, created_at, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s,
                CASE WHEN %s LIKE 'expert%%' THEN NOW() ELSE NULL END,
                NOW(), NOW())
            ON CONFLICT (finding_id) DO UPDATE SET
                priority_label = COALESCE(EXCLUDED.priority_label, auditprep.finding_ml_labels.priority_label),
                criticality_score = COALESCE(EXCLUDED.criticality_score, auditprep.finding_ml_labels.criticality_score),
                label_source = EXCLUDED.label_source,
                validated_by = COALESCE(EXCLUDED.validated_by, auditprep.finding_ml_labels.validated_by),
                validation_comment = COALESCE(EXCLUDED.validation_comment, auditprep.finding_ml_labels.validation_comment),
                validated_at = CASE
                    WHEN EXCLUDED.label_source LIKE 'expert%%' THEN NOW()
                    ELSE auditprep.finding_ml_labels.validated_at
                END,
                updated_at = NOW();
            """,
            (
                int(finding_id), priority_label, score_value, str(label_source or "expert"),
                normalize_blank(validated_by), normalize_blank(validation_comment),
                str(label_source or "expert"),
            ),
        )


def save_expert_label_reviews(host, port, dbname, user, password, reviews_df, reviewer_name=""):
    """Enregistre en une transaction les lignes réellement cochées par l'auditeur."""
    if reviews_df is None or reviews_df.empty:
        return 0
    required = {"finding_id", "priority_experte", "criticite_experte", "valider_expert"}
    if not required.issubset(set(reviews_df.columns)):
        raise ValueError("Colonnes de validation experte incomplètes.")

    selected = reviews_df[reviews_df["valider_expert"].fillna(False).astype(bool)].copy()
    if selected.empty:
        return 0

    saved = 0
    with get_conn(host, port, dbname, user, password) as conn:
        ensure_finding_ml_labels_table(conn)
        for _, row in selected.iterrows():
            upsert_finding_ml_label(
                conn,
                finding_id=row.get("finding_id"),
                priority_label=row.get("priority_experte"),
                criticality_score=row.get("criticite_experte"),
                label_source="expert_validation_v8_3",
                validated_by=reviewer_name,
                validation_comment=row.get("commentaire_expert"),
            )
            saved += 1
        conn.commit()
    return saved


@st.cache_data(show_spinner=False, ttl=60)
def read_history_signals_cached(source_codes_tuple, host, port, dbname, user, password):
    """Construit des signaux métier pour chaque historique disponible.

    V7.9 : la recommandation ne dépend plus seulement du titre de la mission
    ou du contexte saisi manuellement. Elle exploite aussi les signaux déjà
    calculés par le moteur SQL : processus sensibles et clauses ISO sensibles.
    Cela rend le classement plus explicable même quand les historiques n'ont
    pas encore de contexte métier sauvegardé.
    """
    source_codes = [str(c) for c in source_codes_tuple if str(c).strip()]
    if not source_codes:
        return pd.DataFrame(columns=[
            "source_mission_code",
            "history_process_signals",
            "history_clause_signals",
            "history_top_signal",
        ])

    rows = []
    with get_conn(host, port, dbname, user, password) as conn:
        for code in source_codes:
            process_names = []
            clause_names = []
            top_signal = ""

            try:
                process_df = pd.read_sql_query(
                    """
                    SELECT process_name, capped_score, vigilance_level
                    FROM auditprep.vw_dynamic_process_vigilance_dashboard
                    WHERE source_mission_code = %s
                    ORDER BY capped_score DESC NULLS LAST, process_name
                    LIMIT 8;
                    """,
                    conn,
                    params=(code,),
                )
                if not process_df.empty:
                    process_names = [str(x) for x in process_df["process_name"].dropna().tolist()]
                    first = process_df.iloc[0]
                    top_signal = f"{first.get('process_name', '')} ({first.get('capped_score', '')}/100)"
            except Exception:
                process_names = []

            try:
                clause_df = pd.read_sql_query(
                    """
                    SELECT clause_code, clause_title, capped_score, vigilance_level
                    FROM auditprep.vw_dynamic_clause_vigilance_dashboard
                    WHERE source_mission_code = %s
                    ORDER BY capped_score DESC NULLS LAST, clause_code
                    LIMIT 10;
                    """,
                    conn,
                    params=(code,),
                )
                if not clause_df.empty:
                    for _, r in clause_df.iterrows():
                        clause_names.append(f"{r.get('clause_code', '')} {r.get('clause_title', '')}".strip())
            except Exception:
                clause_names = []

            rows.append({
                "source_mission_code": code,
                "history_process_signals": ", ".join(process_names),
                "history_clause_signals": ", ".join(clause_names),
                "history_top_signal": top_signal,
            })

    return pd.DataFrame(rows)


def compact_text(value, max_len=160):
    value = str(value or "").strip()
    if len(value) <= max_len:
        return value
    return value[:max_len - 3].rstrip() + "..."

def create_or_update_mission_from_app(host, port, dbname, user, password, mission_data):
    """Crée une mission depuis l'interface Streamlit.

    Correction V7.3 :
    - mission_title est obligatoire côté PostgreSQL ; on le valide avant l'INSERT.
    - les valeurs saisies dans le formulaire sont réellement envoyées à audit_missions.
    - la fonction reste compatible avec plusieurs schémas possibles : colonnes texte directes
      ou colonnes FK client_id/site_id/primary_standard_id.
    """
    with get_conn(host, port, dbname, user, password) as conn:
        if not table_exists(conn, "audit_missions"):
            raise RuntimeError("La table auditprep.audit_missions est introuvable.")

        cols = table_columns(conn, "audit_missions")
        mission_id_col = pick_col(cols, ["mission_id", "id"])
        mission_code_col = pick_col(cols, ["mission_code", "code"])
        mission_title_col = pick_col(cols, ["mission_title", "title", "name"])

        if not mission_code_col:
            raise RuntimeError("Impossible de trouver la colonne mission_code dans audit_missions.")
        if not mission_title_col:
            raise RuntimeError("Impossible de trouver la colonne mission_title/title dans audit_missions.")

        mission_code = normalize_blank(mission_data.get("mission_code"))
        mission_title = normalize_blank(mission_data.get("mission_title"))
        client_name = normalize_blank(mission_data.get("client_name"))
        site_name = normalize_blank(mission_data.get("site_name"))
        audit_type = normalize_blank(mission_data.get("audit_type")) or "Audit interne"
        standard_name = normalize_blank(mission_data.get("standard_name")) or "ISO 9001:2015"
        planned_audit_date = mission_data.get("planned_audit_date")
        status = normalize_blank(mission_data.get("status")) or "Brouillon"

        if not mission_code:
            raise ValueError("Le code mission est obligatoire.")
        if not mission_title:
            raise ValueError("Le titre mission est obligatoire. Remplis le champ Titre mission avant d’enregistrer.")

        # Si le code existe déjà, on met à jour les champs utiles au lieu de recréer un doublon.
        with conn.cursor() as cur:
            cur.execute(
                sql.SQL("SELECT {} FROM auditprep.audit_missions WHERE {} = %s LIMIT 1").format(
                    sql.Identifier(mission_id_col), sql.Identifier(mission_code_col)
                ),
                (mission_code,),
            )
            existing = cur.fetchone()

        client_id = get_or_create_simple(conn, "clients", ["client_id", "id"], ["client_name", "name"], client_name)
        site_id = None
        if table_exists(conn, "client_sites"):
            extra = {"client_id": client_id} if client_id else {}
            site_id = get_or_create_simple(conn, "client_sites", ["site_id", "id"], ["site_name", "name"], site_name, extra)
        standard_id = get_standard_id(conn, standard_name)
        audit_type_id = get_audit_type_id(conn, audit_type)
        mission_status_id = get_mission_status_id(conn, status)

        if "mission_status_id" in cols and mission_status_id is None:
            raise RuntimeError(
                "La table audit_missions exige mission_status_id, mais aucun statut valide n'a été trouvé. "
                "Vérifie la table de référence des statuts de mission dans PostgreSQL."
            )

        values = {}
        mapping = {
            mission_code_col: mission_code,
            mission_title_col: mission_title,
            "client_name": client_name,
            "site_name": site_name,
            "planned_audit_date": planned_audit_date,
            "audit_date": planned_audit_date,
            "audit_type": audit_type,
            "audit_type_id": audit_type_id,
            "standard_label": standard_name,
            "standard_name": standard_name,
            "status": status,
            "mission_status": status,
            "audit_status": status,
            "mission_status_id": mission_status_id,
            "status_id": mission_status_id,
            "client_id": client_id,
            "site_id": site_id,
            "standard_id": standard_id,
            "primary_standard_id": standard_id,
        }
        for k, v in mapping.items():
            if k in cols and v is not None:
                values[k] = v

        # Filet de sécurité : ces deux colonnes ne doivent jamais être absentes de l'INSERT.
        values[mission_code_col] = mission_code
        values[mission_title_col] = mission_title

        if existing:
            update_values = {k: v for k, v in values.items() if k != mission_code_col and v is not None}
            updated_at_col = pick_col(cols, ["updated_at", "modified_at"])
            if updated_at_col:
                update_values[updated_at_col] = datetime.now()
            if update_values:
                with conn.cursor() as cur:
                    cur.execute(
                        sql.SQL("UPDATE auditprep.audit_missions SET {} WHERE {} = %s").format(
                            sql.SQL(", ").join(
                                sql.SQL("{} = %s").format(sql.Identifier(c)) for c in update_values.keys()
                            ),
                            sql.Identifier(mission_code_col),
                        ),
                        list(update_values.values()) + [mission_code],
                    )
            upsert_mission_context(conn, existing[0], mission_data)
            conn.commit()
            return existing[0], "mise à jour"

        created_at_col = pick_col(cols, ["created_at", "created_on"])
        updated_at_col = pick_col(cols, ["updated_at", "modified_at"])
        now = datetime.now()
        if created_at_col and created_at_col not in values:
            values[created_at_col] = now
        if updated_at_col and updated_at_col not in values:
            values[updated_at_col] = now

        new_id = execute_dynamic_insert(conn, "audit_missions", values, returning_col=mission_id_col)
        upsert_mission_context(conn, new_id, mission_data)
        conn.commit()
        return new_id, "créée"

def ensure_audit_report_for_mission(conn, mission_id, mission_code=None):
    if not table_exists(conn, "audit_reports"):
        raise RuntimeError("La table auditprep.audit_reports est introuvable.")
    cols = table_columns(conn, "audit_reports")
    report_id_col = pick_col(cols, ["audit_report_id", "report_id", "id"])
    mission_id_col = pick_col(cols, ["mission_id"])
    if not report_id_col or not mission_id_col:
        raise RuntimeError("Colonnes audit_report_id/mission_id introuvables dans audit_reports.")

    with conn.cursor() as cur:
        cur.execute(
            sql.SQL("SELECT {} FROM auditprep.audit_reports WHERE {} = %s ORDER BY {} DESC LIMIT 1").format(
                sql.Identifier(report_id_col), sql.Identifier(mission_id_col), sql.Identifier(report_id_col)
            ),
            (mission_id,),
        )
        row = cur.fetchone()
        if row:
            return row[0]

    values = {mission_id_col: mission_id}
    title_col = pick_col(cols, ["report_title", "audit_report_title", "title"])
    if title_col:
        values[title_col] = f"Rapport d’audit - {mission_code or mission_id}"
    date_col = pick_col(cols, ["audit_date", "report_date", "created_at"])
    if date_col:
        values[date_col] = datetime.now()
    return execute_dynamic_insert(conn, "audit_reports", values, returning_col=report_id_col)


def insert_finding_from_app(host, port, dbname, user, password, finding_data):
    mission_code = normalize_blank(finding_data.get("mission_code"))
    if not mission_code:
        raise ValueError("Le code de la mission historique est obligatoire.")

    with get_conn(host, port, dbname, user, password) as conn:
        if not table_exists(conn, "audit_missions") or not table_exists(conn, "audit_findings"):
            raise RuntimeError("Les tables audit_missions ou audit_findings sont introuvables.")

        mission_cols = table_columns(conn, "audit_missions")
        mission_id_col = pick_col(mission_cols, ["mission_id", "id"])
        mission_code_col = pick_col(mission_cols, ["mission_code", "code"])
        with conn.cursor() as cur:
            cur.execute(
                sql.SQL("SELECT {} FROM auditprep.audit_missions WHERE {} = %s LIMIT 1").format(
                    sql.Identifier(mission_id_col), sql.Identifier(mission_code_col)
                ),
                (mission_code,),
            )
            row = cur.fetchone()
            if not row:
                raise ValueError(f"Mission historique introuvable : {mission_code}. Crée d'abord la mission.")
            mission_id = row[0]

        report_id = ensure_audit_report_for_mission(conn, mission_id, mission_code)
        process_id = get_or_create_simple(conn, "processes", ["process_id", "id"], ["process_name", "name", "label"], finding_data.get("process_name"))
        standard_id = get_standard_id(conn, "ISO 9001:2015")
        clause_id = get_clause_id(conn, finding_data.get("clause_code"), finding_data.get("clause_title"), standard_id)
        finding_type_id = get_finding_type_id(conn, finding_data.get("finding_type"))

        cols = table_columns(conn, "audit_findings")
        finding_id_col = pick_col(cols, ["finding_id", "id"])
        values = {}
        mapping = {
            "audit_report_id": report_id,
            "process_id": process_id,
            "clause_id": clause_id,
            "finding_type_id": finding_type_id,
            "finding_title": normalize_blank(finding_data.get("title")),
            "finding_description": normalize_blank(finding_data.get("description")),
            "description": normalize_blank(finding_data.get("description")),
            "finding_text": normalize_blank(finding_data.get("description")),
            "finding_comment": normalize_blank(finding_data.get("description")),
            "severity": normalize_blank(finding_data.get("severity")),
            "severity_level": normalize_blank(finding_data.get("severity")),
            "status": normalize_blank(finding_data.get("status")),
            "finding_status": normalize_blank(finding_data.get("status")),
            "created_at": datetime.now(),
            "finding_date": finding_data.get("finding_date"),
        }
        for k, v in mapping.items():
            if k in cols and v is not None:
                values[k] = v

        new_id = execute_dynamic_insert(conn, "audit_findings", values, returning_col=finding_id_col)

        # V8.3 : l'auditeur fournit les cibles du modèle supervisé.
        # Elles restent dans une table dédiée afin de préserver le moteur SQL V6.
        raw_priority_label = finding_data.get("priority_label")
        raw_criticality_score = finding_data.get("criticality_score")
        explicit_priority = normalize_priority_label(
            raw_priority_label,
            finding_data.get("finding_type"),
        )
        explicit_score = raw_criticality_score if normalize_blank(raw_criticality_score) is not None else None
        if explicit_score is None and explicit_priority:
            explicit_score = priority_to_score(explicit_priority)
        label_source = finding_data.get("label_source", "expert")
        if normalize_blank(raw_priority_label) is None and normalize_blank(raw_criticality_score) is None:
            label_source = "proxy_gravite_type"
        upsert_finding_ml_label(
            conn,
            new_id,
            priority_label=explicit_priority,
            criticality_score=explicit_score,
            label_source=label_source,
        )
        conn.commit()
        return new_id


def normalize_import_columns(df):
    rename_map = {
        "code mission": "mission_code",
        "mission": "mission_code",
        "mission_code": "mission_code",
        "processus": "process_name",
        "process_name": "process_name",
        "clause": "clause_code",
        "clause iso": "clause_code",
        "clause_code": "clause_code",
        "titre clause": "clause_title",
        "clause_title": "clause_title",
        "type": "finding_type",
        "type constat": "finding_type",
        "finding_type": "finding_type",
        "description": "description",
        "constat": "description",
        "finding_description": "description",
        "gravité": "severity",
        "gravite": "severity",
        "severity": "severity",
        "statut": "status",
        "status": "status",
        "date": "finding_date",
        "finding_date": "finding_date",
        "priorité experte": "priority_label",
        "priorite experte": "priority_label",
        "priority_label": "priority_label",
        "score criticité": "criticality_score",
        "score criticite": "criticality_score",
        "criticality_score": "criticality_score",
    }
    out = df.copy()
    normalized = {}
    for c in out.columns:
        key = str(c).strip().lower()
        normalized[c] = rename_map.get(key, c)
    return out.rename(columns=normalized)


def import_findings_dataframe(host, port, dbname, user, password, df):
    df = normalize_import_columns(df)
    required = ["mission_code", "process_name", "finding_type", "description"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError("Colonnes obligatoires manquantes : " + ", ".join(missing))

    inserted = 0
    errors = []
    for idx, row in df.iterrows():
        try:
            insert_finding_from_app(
                host, port, dbname, user, password,
                {
                    "mission_code": row.get("mission_code"),
                    "process_name": row.get("process_name"),
                    "clause_code": row.get("clause_code"),
                    "clause_title": row.get("clause_title"),
                    "finding_type": row.get("finding_type"),
                    "description": row.get("description"),
                    "severity": row.get("severity"),
                    "status": row.get("status"),
                    "finding_date": row.get("finding_date"),
                    "priority_label": row.get("priority_label"),
                    "criticality_score": row.get("criticality_score"),
                    "label_source": "expert_import",
                },
            )
            inserted += 1
        except Exception as exc:
            errors.append(f"Ligne {idx + 2} : {exc}")
    return inserted, errors


def template_findings_excel():
    df = pd.DataFrame(
        [
            {
                "mission_code": "AUD-HIST-2026-001",
                "process_name": "Achats",
                "clause_code": "8.4",
                "clause_title": "Maîtrise des processus, produits et services fournis par des prestataires externes",
                "finding_type": "Non-conformité",
                "description": "Exemple : absence de preuve d’évaluation fournisseur.",
                "severity": "Moyenne",
                "status": "Ouverte",
                "finding_date": datetime.now().date().isoformat(),
                "priority_label": "Haute",
                "criticality_score": 85,
            }
        ]
    )
    return excel_bytes({"modele_constats": df})


@st.cache_data(show_spinner=False, ttl=30)
def load_supervised_dataset(host, port, dbname, user, password):
    """Construit le dataset ML à partir des constats et des étiquettes métier.

    La lecture est volontairement tolérante aux variantes de colonnes du schéma
    PostgreSQL. Les anciennes lignes sans étiquette validée reçoivent une cible
    proxy dérivée de la gravité/type ; cette provenance reste visible.
    """
    with get_conn(host, port, dbname, user, password) as conn:
        ensure_finding_ml_labels_table(conn)
        conn.commit()

        def read_table(table_name, required=False):
            if not table_exists(conn, table_name):
                if required:
                    raise RuntimeError(f"Table auditprep.{table_name} introuvable.")
                return pd.DataFrame()
            return pd.read_sql_query(
                sql.SQL("SELECT * FROM auditprep.{}").format(sql.Identifier(table_name)).as_string(conn),
                conn,
            )

        findings = read_table("audit_findings", required=True)
        reports = read_table("audit_reports")
        missions = read_table("audit_missions")
        processes = read_table("processes")
        clauses = read_table("clauses")
        finding_types = read_table("finding_types")
        contexts = read_table("mission_contexts")
        labels = read_table("finding_ml_labels")

    if findings.empty:
        return pd.DataFrame()

    data = findings.copy()

    def merge_prefixed(base, left_col, lookup, right_candidates, prefix):
        if not left_col or lookup is None or lookup.empty:
            return base
        right_col = pick_col(list(lookup.columns), right_candidates)
        if not right_col:
            return base
        prefixed = lookup.rename(columns={c: f"{prefix}{c}" for c in lookup.columns})
        return base.merge(
            prefixed,
            how="left",
            left_on=left_col,
            right_on=f"{prefix}{right_col}",
        )

    finding_cols = list(findings.columns)
    finding_id_col = pick_col(finding_cols, ["finding_id", "id"])
    report_fk = pick_col(finding_cols, ["audit_report_id", "report_id"])
    process_fk = pick_col(finding_cols, ["process_id"])
    clause_fk = pick_col(finding_cols, ["clause_id"])
    finding_type_fk = pick_col(finding_cols, ["finding_type_id", "type_id"])

    data = merge_prefixed(data, report_fk, reports, ["audit_report_id", "report_id", "id"], "r__")
    report_mission_fk = pick_col(list(reports.columns), ["mission_id"]) if not reports.empty else None
    data = merge_prefixed(
        data,
        f"r__{report_mission_fk}" if report_mission_fk else None,
        missions,
        ["mission_id", "id"],
        "m__",
    )
    data = merge_prefixed(data, process_fk, processes, ["process_id", "id"], "p__")
    data = merge_prefixed(data, clause_fk, clauses, ["clause_id", "id"], "c__")
    data = merge_prefixed(data, finding_type_fk, finding_types, ["finding_type_id", "type_id", "id"], "ft__")
    data = merge_prefixed(data, finding_id_col, labels, ["finding_id"], "ml__")

    mission_id_col = pick_col(list(missions.columns), ["mission_id", "id"]) if not missions.empty else None
    data = merge_prefixed(
        data,
        f"m__{mission_id_col}" if mission_id_col else None,
        contexts,
        ["mission_id"],
        "mc__",
    )

    def coalesce(candidates, default=""):
        available = [c for c in candidates if c in data.columns]
        if not available:
            return pd.Series(default, index=data.index, dtype="object")
        frame = data[available].copy()
        for col in frame.columns:
            if frame[col].dtype == object:
                frame[col] = frame[col].replace(r"^\s*$", pd.NA, regex=True)
        return frame.bfill(axis=1).iloc[:, 0].fillna(default)

    dataset = pd.DataFrame(index=data.index)
    dataset["finding_id"] = coalesce([finding_id_col] if finding_id_col else [], "")
    dataset["mission_code"] = coalesce(["m__mission_code", "m__code"], "")
    dataset["process_name"] = coalesce(["p__process_name", "p__name", "p__label", "process_name"], "Processus non précisé")
    dataset["clause_code"] = coalesce(["c__clause_code", "c__code", "clause_code"], "")
    dataset["clause_title"] = coalesce(["c__clause_title", "c__title", "c__label", "clause_title"], "")
    dataset["finding_type"] = coalesce(["ft__label", "ft__type_name", "ft__name", "finding_type"], "")
    dataset["finding_title"] = coalesce(["finding_title", "title"], "")
    dataset["finding_description"] = coalesce(
        ["finding_description", "description", "finding_text", "finding_comment"],
        "",
    )
    dataset["severity_raw"] = coalesce(["severity", "severity_level", "gravity", "gravite"], "")
    dataset["status"] = coalesce(["status", "finding_status"], "")
    dataset["sector"] = coalesce(["mc__sector"], "")
    dataset["audited_process"] = coalesce(["mc__audited_process"], "")
    dataset["known_risks"] = coalesce(["mc__known_risks"], "")
    dataset["keywords"] = coalesce(["mc__keywords"], "")

    explicit_priority = coalesce(["ml__priority_label"], "")
    dataset["target_priority"] = [
        normalize_priority_label(explicit, ftype)
        or normalize_priority_label(severity, ftype)
        for explicit, severity, ftype in zip(explicit_priority, dataset["severity_raw"], dataset["finding_type"])
    ]

    raw_score = pd.to_numeric(
        coalesce(
            ["ml__criticality_score", "criticality_score", "criticite_score", "risk_score", "score"],
            pd.NA,
        ),
        errors="coerce",
    )
    dataset["target_criticality"] = raw_score
    missing_score = dataset["target_criticality"].isna()
    # V8.5.28 : pandas/Arrow sur Streamlit Cloud conserve parfois le dtype
    # string d'une sélection vide. L'affecter à la colonne float64 provoquait
    # alors ``Invalid value '<ArrowStringArray>' for dtype 'float64'`` même
    # lorsqu'aucun score n'était manquant. On ne calcule les scores proxy que
    # lorsqu'ils sont réellement nécessaires, puis on affecte un tableau
    # explicitement numérique.
    if bool(missing_score.any()):
        proxy_scores = (
            dataset.loc[missing_score, "target_priority"]
            .astype("object")
            .map(priority_to_score)
        )
        dataset.loc[missing_score, "target_criticality"] = pd.to_numeric(
            proxy_scores,
            errors="coerce",
        ).to_numpy(dtype="float64")

    stored_source = coalesce(["ml__label_source"], "")
    dataset["target_source"] = stored_source.astype(str)
    explicit_mask = explicit_priority.astype(str).str.strip().ne("") | raw_score.notna()
    dataset.loc[explicit_mask & dataset["target_source"].str.strip().eq(""), "target_source"] = "historique_explicite"
    dataset.loc[~explicit_mask, "target_source"] = "proxy_gravite_type"
    dataset["validated_by"] = coalesce(["ml__validated_by"], "")
    dataset["validation_comment"] = coalesce(["ml__validation_comment"], "")
    dataset["validated_at"] = coalesce(["ml__validated_at"], "")
    # V8.3 : finding_type et severity_raw servent encore à construire certaines
    # étiquettes proxy. Ils sont donc volontairement exclus des entrées du
    # modèle, tout comme mission_code qui est un identifiant et non un signal
    # généralisable. Cette séparation évite une fuite directe de la cible.
    text_cols = [
        "process_name", "clause_code", "clause_title", "finding_title",
        "finding_description", "sector", "audited_process", "known_risks", "keywords",
    ]
    dataset["model_text_raw"] = dataset[text_cols].fillna("").astype(str).agg(" | ".join, axis=1)
    dataset["model_text"] = dataset["model_text_raw"].map(clean_french_model_text)
    dataset = dataset[
        dataset["target_priority"].notna()
        & dataset["target_criticality"].notna()
        & dataset["model_text"].str.strip().ne("")
    ].copy()
    dataset["target_criticality"] = pd.to_numeric(dataset["target_criticality"], errors="coerce").clip(0, 100)
    return dataset.reset_index(drop=True)


def supervised_dataset_fingerprint(dataset_df):
    if dataset_df is None or dataset_df.empty:
        return "empty"
    cols = [
        c for c in [
            "finding_id", "target_priority", "target_criticality", "target_source",
            "validated_at", "model_text",
        ]
        if c in dataset_df.columns
    ]
    return str(int(pd.util.hash_pandas_object(dataset_df[cols].astype(str), index=True).sum()))


def train_supervised_models(dataset_df, random_state=42):
    """Entraîne et évalue une classification et une régression sur un split 70/30."""
    pack = {
        "status": "insufficient_data",
        "message": "Données insuffisantes pour entraîner les deux modèles.",
        "dataset_rows": 0,
        "train_rows": 0,
        "test_rows": 0,
        "classification_metrics": {},
        "regression_metrics": {},
        "confusion_matrix": pd.DataFrame(),
        "classification_report": pd.DataFrame(),
        "classification_importance": pd.DataFrame(),
        "regression_importance": pd.DataFrame(),
        "training_scope": "Aucune donnée",
        "expert_rows_available": 0,
        "classifier": None,
        "regressor": None,
    }
    if dataset_df is None or dataset_df.empty:
        return pack

    # V8.3 privilégie les étiquettes expertes dès que leur répartition permet
    # une évaluation séparée. Avant ce seuil, le mode proxy reste transitoire.
    selected_df, training_scope, expert_rows_available = select_training_dataset(dataset_df)
    pack["dataset_rows"] = int(len(selected_df))
    pack["training_scope"] = training_scope
    pack["expert_rows_available"] = expert_rows_available

    try:
        import numpy as np
        from sklearn.base import clone
        from sklearn.ensemble import RandomForestRegressor
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.linear_model import LogisticRegression
        from sklearn.metrics import (
            accuracy_score,
            classification_report,
            confusion_matrix,
            mean_absolute_error,
            mean_squared_error,
            r2_score,
        )
        from sklearn.model_selection import train_test_split
        from sklearn.pipeline import Pipeline
    except ImportError as exc:
        pack["status"] = "dependency_missing"
        pack["message"] = f"scikit-learn est requis pour le module V8.3 : {exc}"
        return pack

    data = selected_df.dropna(subset=["model_text", "target_priority", "target_criticality"]).copy()
    data["model_text"] = data["model_text"].astype(str)
    data = data[data["model_text"].str.strip().ne("")]
    n_rows = len(data)
    class_counts = data["target_priority"].value_counts()
    n_classes = int(class_counts.size)
    pack["dataset_rows"] = int(n_rows)

    if n_rows < 10:
        pack["message"] = f"{n_rows} constat(s) étiqueté(s). Il en faut au moins 10 pour créer un test séparé."
        return pack
    if n_classes < 2 or int(class_counts.min()) < 2:
        pack["message"] = "Il faut au moins deux classes de priorité et deux exemples par classe."
        return pack

    test_rows = max(n_classes, int(math.floor(n_rows * 0.30)))
    if n_rows - test_rows < int(math.ceil(n_rows * 0.70)):
        test_rows = n_rows - int(math.ceil(n_rows * 0.70))
    if test_rows < n_classes:
        pack["message"] = "Le volume actuel ne permet pas un split 70/30 stratifié couvrant toutes les classes."
        return pack

    indices = list(range(n_rows))
    train_idx, test_idx = train_test_split(
        indices,
        test_size=test_rows,
        random_state=random_state,
        stratify=data["target_priority"],
    )
    x_train = data.iloc[train_idx]["model_text"]
    x_test = data.iloc[test_idx]["model_text"]
    y_class_train = data.iloc[train_idx]["target_priority"]
    y_class_test = data.iloc[test_idx]["target_priority"]
    y_reg_train = data.iloc[train_idx]["target_criticality"].astype(float)
    y_reg_test = data.iloc[test_idx]["target_criticality"].astype(float)

    classifier_template = Pipeline([
        ("tfidf", TfidfVectorizer(max_features=1200, ngram_range=(1, 2), min_df=1, strip_accents="unicode")),
        ("model", LogisticRegression(max_iter=2500, class_weight="balanced", random_state=random_state)),
    ])
    regression_template = Pipeline([
        ("tfidf", TfidfVectorizer(max_features=1200, ngram_range=(1, 2), min_df=1, strip_accents="unicode")),
        ("model", RandomForestRegressor(
            n_estimators=250,
            max_depth=10,
            min_samples_leaf=1,
            random_state=random_state,
            n_jobs=-1,
        )),
    ])

    classifier_eval = clone(classifier_template).fit(x_train, y_class_train)
    regressor_eval = clone(regression_template).fit(x_train, y_reg_train)
    class_pred = classifier_eval.predict(x_test)
    reg_pred = regressor_eval.predict(x_test)

    class_labels = [p for p in ["Faible", "Moyenne", "Haute"] if p in set(data["target_priority"])]
    cm = confusion_matrix(y_class_test, class_pred, labels=class_labels)
    report = classification_report(
        y_class_test,
        class_pred,
        labels=class_labels,
        output_dict=True,
        zero_division=0,
    )
    mae = float(mean_absolute_error(y_reg_test, reg_pred))
    rmse = float(math.sqrt(mean_squared_error(y_reg_test, reg_pred)))
    r2 = float(r2_score(y_reg_test, reg_pred)) if len(y_reg_test) >= 2 else float("nan")

    # Les modèles finaux sont réentraînés sur 100 % des données après l'évaluation 70/30.
    classifier_final = clone(classifier_template).fit(data["model_text"], data["target_priority"])
    regressor_final = clone(regression_template).fit(data["model_text"], data["target_criticality"].astype(float))

    class_features = classifier_final.named_steps["tfidf"].get_feature_names_out()
    class_coef = np.abs(classifier_final.named_steps["model"].coef_).mean(axis=0)
    class_importance = pd.DataFrame({"variable": class_features, "importance": class_coef})
    class_importance = class_importance.sort_values("importance", ascending=False).head(25).reset_index(drop=True)

    reg_features = regressor_final.named_steps["tfidf"].get_feature_names_out()
    reg_importance = pd.DataFrame({
        "variable": reg_features,
        "importance": regressor_final.named_steps["model"].feature_importances_,
    })
    reg_importance = reg_importance.sort_values("importance", ascending=False).head(25).reset_index(drop=True)

    pack.update({
        "status": "trained",
        "message": "Classification et régression entraînées avec évaluation sur un jeu de test séparé.",
        "dataset_rows": int(n_rows),
        "train_rows": int(len(train_idx)),
        "test_rows": int(len(test_idx)),
        "classification_metrics": {"accuracy": float(accuracy_score(y_class_test, class_pred))},
        "regression_metrics": {"mae": mae, "rmse": rmse, "r2": r2},
        "confusion_matrix": pd.DataFrame(cm, index=[f"Réel {x}" for x in class_labels], columns=[f"Prédit {x}" for x in class_labels]),
        "classification_report": pd.DataFrame(report).transpose().rename_axis("classe").reset_index(),
        "classification_importance": class_importance,
        "regression_importance": reg_importance,
        "classifier": classifier_final,
        "regressor": regressor_final,
        "trained_at": datetime.now(),
    })
    return pack


def checklist_model_text(checklist_df):
    cols = [
        c for c in [
            "question_text", "clause_code", "clause_title", "theme", "recommendation_label",
            "expected_evidence", "context_sector", "context_process", "context_objective", "context_scope",
        ]
        if c in checklist_df.columns
    ]
    if not cols:
        return pd.Series("", index=checklist_df.index, dtype="object")
    return checklist_df[cols].fillna("").astype(str).agg(" | ".join, axis=1)


def priority_from_criticality(score):
    """Transforme le score régressé en classe lisible pour contrôler la cohérence."""
    try:
        value = float(score)
    except (TypeError, ValueError):
        return "Non disponible"
    if value >= 70:
        return "Haute"
    if value >= 45:
        return "Moyenne"
    return "Faible"


def apply_supervised_predictions(checklist_df, ml_pack, confidence_threshold=60.0):
    """Fusion prudente entre règles métier, classification et régression.

    La V8.3 n'abaisse jamais une priorité métier sur la seule base d'un petit
    modèle exploratoire. Une hausse proposée par le ML n'est appliquée que si
    classification et régression sont cohérentes et si la confiance atteint le
    seuil. Tous les autres cas restent à confirmer par l'auditeur.
    """
    if checklist_df is None:
        return pd.DataFrame(), "Aucune check-list à prédire."
    out = checklist_df.copy()
    if "generated_priority" not in out.columns:
        out["generated_priority"] = "Moyenne"
    out["rule_based_priority"] = out["generated_priority"]

    if not ml_pack or ml_pack.get("status") != "trained" or out.empty:
        out["ml_predicted_priority"] = "Non disponible"
        out["ml_prediction_confidence"] = pd.NA
        out["ml_predicted_criticality"] = pd.NA
        out["ml_criticality_band"] = "Non disponible"
        out["ml_decision"] = "Règles conservées : modèle non disponible"
        out["priority_origin"] = "Règles métier - modèle supervisé non disponible"
        out["ml_review_required"] = True
        return out, "Modèle non appliqué : les priorités restent issues des règles métier."

    texts = checklist_model_text(out)
    classifier = ml_pack["classifier"]
    regressor = ml_pack["regressor"]
    predictions = classifier.predict(texts)
    probabilities = classifier.predict_proba(texts)
    confidence = probabilities.max(axis=1) * 100.0
    criticality = regressor.predict(texts)

    out["ml_predicted_priority"] = predictions
    out["ml_prediction_confidence"] = confidence.round(1)
    out["ml_predicted_criticality"] = pd.Series(criticality, index=out.index).clip(0, 100).round(1)
    out["ml_criticality_band"] = out["ml_predicted_criticality"].map(priority_from_criticality)

    priority_rank = {"Faible": 1, "Moyenne": 2, "Haute": 3}
    final_priorities = []
    decisions = []
    origins = []
    reviews = []

    for _, row in out.iterrows():
        rule_priority = str(row.get("rule_based_priority", "Moyenne"))
        ml_priority = str(row.get("ml_predicted_priority", "Moyenne"))
        score_band = str(row.get("ml_criticality_band", "Non disponible"))
        conf = float(row.get("ml_prediction_confidence", 0) or 0)
        coherent = ml_priority == score_band
        confident = conf >= float(confidence_threshold)
        disagreement = ml_priority != rule_priority

        # Fusion prudente : le ML peut renforcer une priorité, jamais la réduire
        # automatiquement. Une prédiction faible ou incohérente reste informative.
        can_raise = (
            confident
            and coherent
            and priority_rank.get(ml_priority, 0) > priority_rank.get(rule_priority, 0)
        )
        if can_raise:
            final_priorities.append(ml_priority)
            decisions.append("Priorité renforcée par un signal ML cohérent")
            origins.append("Fusion prudente règles + ML V8.3")
        else:
            final_priorities.append(rule_priority)
            if not confident:
                decisions.append(f"Règle conservée : confiance ML < {confidence_threshold:.0f} %")
            elif not coherent:
                decisions.append("Règle conservée : classification et criticité divergentes")
            elif priority_rank.get(ml_priority, 0) < priority_rank.get(rule_priority, 0):
                decisions.append("Règle conservée : le ML proposait une baisse")
            else:
                decisions.append("Règle confirmée par le ML")
            origins.append("Règles métier sécurisées par contrôle ML V8.3")

        reviews.append((not confident) or (not coherent) or disagreement)

    out["generated_priority"] = final_priorities
    out["ml_decision"] = decisions
    out["priority_origin"] = origins
    out["ml_review_required"] = reviews
    review_count = int(out["ml_review_required"].sum())
    raised_count = int((out["generated_priority"].astype(str) != out["rule_based_priority"].astype(str)).sum())
    return out, (
        f"Contrôle supervisé appliqué à {len(out)} question(s) : {raised_count} priorité(s) renforcée(s) "
        f"et {review_count} prédiction(s) à confirmer par l’auditeur."
    )


def ml_metrics_export_df(ml_pack):
    if not ml_pack:
        return pd.DataFrame()
    cm = ml_pack.get("classification_metrics", {})
    rm = ml_pack.get("regression_metrics", {})
    return pd.DataFrame([
        {"famille": "Données", "metrique": "Périmètre d'entraînement", "valeur": ml_pack.get("training_scope", "")},
        {"famille": "Données", "metrique": "Étiquettes expertes disponibles", "valeur": ml_pack.get("expert_rows_available", 0)},
        {"famille": "Découpage", "metrique": "Lignes entraînement", "valeur": ml_pack.get("train_rows", 0)},
        {"famille": "Découpage", "metrique": "Lignes test", "valeur": ml_pack.get("test_rows", 0)},
        {"famille": "Classification", "metrique": "Accuracy", "valeur": cm.get("accuracy")},
        {"famille": "Régression", "metrique": "MAE", "valeur": rm.get("mae")},
        {"famille": "Régression", "metrique": "RMSE", "valeur": rm.get("rmse")},
        {"famille": "Régression", "metrique": "R²", "valeur": rm.get("r2")},
    ])


def clean_df(df):
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = out[col].astype(str)
    return out


def safe_first(df, col, default=""):
    if df is None or df.empty or col not in df.columns:
        return default
    val = df.iloc[0][col]
    if pd.isna(val):
        return default
    return val


def priority_badge(priority):
    p = str(priority or "").lower()
    if "haute" in p or "élev" in p:
        return '<span class="audit-badge badge-high">Élevée</span>'
    if "moy" in p or "mod" in p:
        return '<span class="audit-badge badge-medium">Modérée</span>'
    return '<span class="audit-badge badge-low">Faible</span>'



def normalize_context_tokens(*values):
    """Prépare des mots-clés simples pour comparer une mission cible avec les historiques disponibles."""
    raw = " ".join(str(v or "") for v in values).lower()
    replacements = {
        "é": "e", "è": "e", "ê": "e", "ë": "e",
        "à": "a", "â": "a",
        "î": "i", "ï": "i",
        "ô": "o",
        "ù": "u", "û": "u",
        "ç": "c",
    }
    for a, b in replacements.items():
        raw = raw.replace(a, b)
    tokens = []
    stopwords = {
        "audit", "les", "des", "une", "dans", "pour", "avec", "site", "mission",
        "processus", "conformite", "preparation", "preparer", "verifier", "identifier",
        "associes", "associe", "critiques", "critique", "objectif", "perimetre"
    }
    for part in raw.replace("-", " ").replace("_", " ").replace("/", " ").replace(",", " ").replace(";", " ").split():
        part = "".join(ch for ch in part if ch.isalnum())
        if len(part) >= 3 and part not in stopwords:
            tokens.append(part)
    return set(tokens)


def expand_business_tokens(tokens):
    """Ajoute des équivalences métier simples pour éviter une comparaison trop littérale."""
    expanded = set(tokens or set())
    synonyms = {
        "fournisseur": {"fournisseurs", "prestataire", "prestataires", "externe", "externes", "achats", "achat"},
        "fournisseurs": {"fournisseur", "prestataire", "prestataires", "externe", "externes", "achats", "achat"},
        "prestataire": {"prestataires", "fournisseur", "fournisseurs", "externe", "externes", "achats"},
        "prestataires": {"prestataire", "fournisseur", "fournisseurs", "externe", "externes", "achats"},
        "achat": {"achats", "fournisseur", "fournisseurs", "prestataire", "prestataires", "8.4", "84"},
        "achats": {"achat", "fournisseur", "fournisseurs", "prestataire", "prestataires", "8.4", "84"},
        "document": {"documents", "documentaire", "documentees", "information", "informations", "traceabilite", "7.5", "75"},
        "documents": {"document", "documentaire", "documentees", "information", "informations", "traceabilite", "7.5", "75"},
        "documentaire": {"document", "documents", "documentees", "information", "informations", "traceabilite", "7.5", "75"},
        "traceabilite": {"document", "documents", "documentaire", "preuves", "enregistrements"},
        "preuve": {"preuves", "enregistrements", "traceabilite", "documents"},
        "preuves": {"preuve", "enregistrements", "traceabilite", "documents"},
        "formation": {"competence", "competences", "rh", "ressources", "7.2", "72"},
        "competence": {"competences", "formation", "rh", "ressources", "7.2", "72"},
        "competences": {"competence", "formation", "rh", "ressources", "7.2", "72"},
        "qse": {"qualite", "hse", "9001", "iso", "systeme"},
        "hse": {"qse", "securite", "risques", "6.1", "61"},
        "risque": {"risques", "opportunites", "6.1", "61", "hse"},
        "risques": {"risque", "opportunites", "6.1", "61", "hse"},
        "9001": {"iso", "qualite", "qse", "systeme"},
    }
    for t in list(expanded):
        expanded.update(synonyms.get(t, set()))
    return expanded


def score_history_candidate(row, target_row=None, sector="", process="", objective="", scope="", audit_type="", standard=""):
    """Score simple de pertinence d'un historique.

    Le moteur ne choisit plus un historique au hasard : il compare le contexte saisi
    avec les anciennes missions disponibles. Le score reste volontairement explicable.
    """
    target_text = " ".join([
        str(target_row.get("mission_title", "")) if target_row is not None else "",
        str(target_row.get("client_name", "")) if target_row is not None else "",
        str(target_row.get("site_name", "")) if target_row is not None else "",
        sector,
        process,
        objective,
        scope,
        audit_type,
        standard,
    ])
    hist_text = " ".join([
        str(row.get("mission_code", "")),
        str(row.get("mission_title", "")),
        str(row.get("client_name", "")),
        str(row.get("site_name", "")),
        str(row.get("standard_name", "")),
        str(row.get("audit_date", "")),
        str(row.get("sector", "")),
        str(row.get("audited_process", "")),
        str(row.get("audit_objective", "")),
        str(row.get("audit_scope", "")),
        str(row.get("specific_requirements", "")),
        str(row.get("known_risks", "")),
        str(row.get("keywords", "")),
        str(row.get("history_process_signals", "")),
        str(row.get("history_clause_signals", "")),
        str(row.get("history_top_signal", "")),
    ])

    target_tokens = expand_business_tokens(normalize_context_tokens(target_text))
    hist_tokens = expand_business_tokens(normalize_context_tokens(hist_text))
    common = target_tokens.intersection(hist_tokens)

    score = 0
    score += min(len(common) * 15, 60)

    # Les historiques avec plus de constats sont plus utiles.
    try:
        findings_count = int(row.get("findings_count", 0) or 0)
    except Exception:
        findings_count = 0
    score += min(findings_count * 3, 25)

    # Bonus si même client ou même site.
    if target_row is not None:
        if str(row.get("client_name", "")).strip().lower() == str(target_row.get("client_name", "")).strip().lower():
            score += 10
        if str(row.get("site_name", "")).strip().lower() == str(target_row.get("site_name", "")).strip().lower():
            score += 5

    return min(score, 100), ", ".join(sorted(common)) if common else "Aucun mot-clé commun fort"


def build_history_recommendations_df(sources_df, target_row, sector, process, objective, scope, audit_type, standard):
    """Construit le classement des historiques pertinents.

    V7.9 : le classement affiche aussi les signaux de l'historique
    (processus sensibles, clauses sensibles) et un niveau de confiance.
    Cela évite de donner l'impression que le moteur choisit au hasard quand
    une seule mission historique est disponible.
    """
    rows = []
    if sources_df is None or sources_df.empty:
        return pd.DataFrame(columns=[
            "mission_code", "mission_title", "client_name", "audit_date",
            "findings_count", "score_pertinence", "niveau_confiance",
            "signaux_historique", "raison_recommandation"
        ])

    candidate_count = int(len(sources_df))

    for _, row in sources_df.iterrows():
        score, common = score_history_candidate(
            row,
            target_row=target_row,
            sector=sector,
            process=process,
            objective=objective,
            scope=scope,
            audit_type=audit_type,
            standard=standard,
        )

        try:
            findings_count = int(row.get("findings_count", 0) or 0)
        except Exception:
            findings_count = 0

        if score >= 70:
            confidence = "Forte"
            usage = "Historique très pertinent"
        elif score >= 45:
            confidence = "Moyenne"
            usage = "Historique utilisable avec contrôle auditeur"
        else:
            confidence = "Faible à confirmer"
            usage = "Historique utilisé faute de meilleur candidat" if candidate_count == 1 else "Historique peu similaire"

        hist_signals = "; ".join([
            compact_text(row.get("history_top_signal", ""), 80),
            compact_text(row.get("history_process_signals", ""), 120),
            compact_text(row.get("history_clause_signals", ""), 120),
        ]).strip("; ")
        if not hist_signals:
            hist_signals = "Aucun signal de vigilance détaillé disponible"

        if candidate_count == 1:
            reason = (
                f"Score {score}/100. Seul historique exploitable disponible : le moteur l'utilise par défaut, "
                f"mais la similarité doit être confirmée par l'auditeur. Mots-clés/signaux communs : {common}. "
                f"Constats disponibles : {findings_count}."
            )
        else:
            reason = (
                f"Score {score}/100. {usage}. Mots-clés/signaux communs : {common}. "
                f"Constats disponibles : {findings_count}."
            )

        rows.append({
            "mission_code": row.get("mission_code", ""),
            "mission_title": row.get("mission_title", ""),
            "client_name": row.get("client_name", ""),
            "site_name": row.get("site_name", ""),
            "secteur": row.get("sector", ""),
            "processus_audité": row.get("audited_process", ""),
            "audit_date": row.get("audit_date", ""),
            "findings_count": findings_count,
            "score_pertinence": score,
            "niveau_confiance": confidence,
            "signaux_historique": hist_signals,
            "raison_recommandation": reason,
        })

    out = pd.DataFrame(rows)
    return out.sort_values(["score_pertinence", "findings_count"], ascending=[False, False]).reset_index(drop=True)

def build_first_audit_checklist(target_code, target_title, sector="", process="", objective="", scope="", audit_type="", standard="ISO 9001:2015"):
    """Génère une check-list initiale quand aucun historique n'est disponible.

    Ce mode est nécessaire pour un premier audit : la priorisation vient du référentiel,
    du processus audité et du contexte saisi, pas de constats historiques.
    """
    context_process = normalize_blank(process) or "Processus audité"
    context_sector = normalize_blank(sector) or "Secteur non précisé"
    context_scope = normalize_blank(scope) or "Périmètre à préciser pendant la préparation"
    context_objective = normalize_blank(objective) or "Évaluer la conformité et identifier les points de vigilance"

    base_questions = [
        ("4.1", "Contexte de l’organisme", "Contexte", f"Le contexte de {context_sector} et les enjeux internes/externes ont-ils été identifiés et tenus à jour ?", "Moyenne", "Analyse de contexte, SWOT/PESTEL, comptes rendus de revue"),
        ("4.2", "Parties intéressées", "Parties intéressées", "Les besoins et attentes des parties intéressées pertinentes sont-ils identifiés et surveillés ?", "Moyenne", "Liste parties intéressées, exigences clients/réglementaires, suivi"),
        ("5.1", "Leadership", "Pilotage", "La direction démontre-t-elle son engagement dans le système de management et dans le périmètre audité ?", "Moyenne", "Politique qualité, objectifs, communication interne"),
        ("6.1", "Actions face aux risques et opportunités", "Risques", f"Les risques et opportunités liés à {context_process} sont-ils identifiés, évalués et traités ?", "Haute", "Cartographie des risques, plans d’actions, preuves de suivi"),
        ("6.2", "Objectifs qualité", "Objectifs", "Les objectifs qualité sont-ils mesurables, suivis et cohérents avec l’objectif de la mission ?", "Moyenne", "Tableau de bord, indicateurs, plans d’amélioration"),
        ("7.1", "Ressources", "Ressources", "Les ressources nécessaires au fonctionnement du processus sont-elles disponibles et adaptées ?", "Moyenne", "Planning, ressources humaines/matérielles, budget"),
        ("7.2", "Compétences", "Compétences", f"Les compétences nécessaires pour {context_process} sont-elles définies, évaluées et maintenues ?", "Haute", "Fiches de poste, habilitations, plan de formation, évaluations"),
        ("7.5", "Informations documentées", "Documentation", "Les documents et enregistrements nécessaires sont-ils maîtrisés, disponibles et à jour ?", "Haute", "Procédures, versions, enregistrements, preuves de diffusion"),
        ("8.1", "Planification et maîtrise opérationnelles", "Maîtrise opérationnelle", f"Le processus {context_process} est-il planifié, maîtrisé et surveillé dans le périmètre : {context_scope} ?", "Haute", "Procédures opérationnelles, contrôles, enregistrements"),
        ("8.4", "Prestataires externes", "Fournisseurs", "Les prestataires/fournisseurs impactant le processus sont-ils évalués, sélectionnés et surveillés ?", "Haute", "Évaluations fournisseurs, contrats, critères de sélection"),
        ("8.5", "Production et prestation de service", "Réalisation", "Les activités opérationnelles sont-elles réalisées dans des conditions maîtrisées ?", "Moyenne", "Instructions, contrôles, traçabilité, validation"),
        ("8.6", "Libération des produits et services", "Validation", "Les critères d’acceptation/libération sont-ils définis et vérifiés avant livraison ou clôture ?", "Moyenne", "PV de contrôle, validation, signatures, critères"),
        ("8.7", "Éléments non conformes", "Non-conformités", "Les non-conformités sont-elles identifiées, enregistrées, traitées et suivies efficacement ?", "Haute", "Registre NC, actions immédiates, analyses de causes"),
        ("9.1", "Surveillance et mesure", "Indicateurs", "Les performances du processus sont-elles mesurées avec des indicateurs pertinents ?", "Moyenne", "KPI, tableaux de bord, analyses périodiques"),
        ("9.2", "Audit interne", "Audit interne", "Le programme d’audit interne couvre-t-il le processus et les exigences applicables ?", "Faible", "Programme audit, rapports, suivi des actions"),
        ("9.3", "Revue de direction", "Revue", "Les résultats du processus sont-ils intégrés à la revue de direction ou au pilotage ?", "Faible", "Compte rendu revue, décisions, plans d’actions"),
        ("10.2", "Non-conformité et action corrective", "Actions correctives", "Les actions correctives sont-elles définies, suivies et vérifiées en efficacité ?", "Haute", "Plans d’actions, responsables, échéances, preuves d’efficacité"),
        ("10.3", "Amélioration continue", "Amélioration", "Des opportunités d’amélioration sont-elles identifiées et suivies ?", "Moyenne", "Plan d’amélioration, retours d’expérience, suggestions"),
    ]

    batch_code = f"FIRST_AUDIT_SESSION_{target_code}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    temporary_run_id = f"TEMP-{batch_code}"
    rows = []
    for i, (clause_code, clause_title, theme, question, priority, evidence) in enumerate(base_questions, start=1):
        rows.append({
            "generation_run_id": temporary_run_id,
            "generation_batch_code": batch_code,
            "target_mission_code": target_code,
            "target_mission_title": target_title,
            "source_mission_code": "AUCUN_HISTORIQUE",
            "source_mission_title": "Premier audit / génération sans historique",
            "display_order": i,
            "clause_code": clause_code,
            "clause_title": clause_title,
            "theme": theme,
            "question_text": question,
            "generated_priority": priority,
            "recommendation_label": "Question générée depuis le référentiel et le contexte de mission",
            "expected_evidence": evidence,
            "conformity_status": "À vérifier",
            "generation_mode": "Premier audit sans historique",
            "context_sector": context_sector,
            "context_process": context_process,
            "context_objective": context_objective,
            "context_scope": context_scope,
        })
    return pd.DataFrame(rows)


def build_first_audit_datasets(checklist_df):
    """Construit KPI, vigilance et alertes simplifiés pour le mode sans historique."""
    if checklist_df is None or checklist_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    kpi_df = (
        checklist_df.groupby("generated_priority")
        .size()
        .reset_index(name="questions_count")
    )

    process_name = checklist_df["context_process"].iloc[0] if "context_process" in checklist_df.columns else "Processus audité"
    process_df = pd.DataFrame([
        {
            "process_name": process_name,
            "findings_count": 0,
            "nonconformities_count": 0,
            "remarks_count": 0,
            "improvements_count": 0,
            "capped_score": 60,
            "vigilance_level": "Moyenne",
            "explanation_summary": "Premier audit : score initial basé sur le contexte de mission et les clauses ISO critiques.",
        }
    ])

    priority_scores = {"Haute": 75, "Moyenne": 55, "Faible": 35}
    clause_df = (
        checklist_df[["clause_code", "clause_title", "generated_priority"]]
        .drop_duplicates()
        .assign(
            findings_count=0,
            capped_score=lambda d: d["generated_priority"].map(priority_scores).fillna(50),
            vigilance_level=lambda d: d["generated_priority"].replace({"Haute": "Élevée", "Moyenne": "Moyenne", "Faible": "Faible"}),
            explanation_summary="Premier audit : priorité issue du référentiel ISO 9001 et du contexte saisi.",
        )
        .sort_values(["capped_score", "clause_code"], ascending=[False, True])
    )

    alerts_df = clause_df.head(5).copy()
    alerts_df["alert_label"] = alerts_df["clause_code"] + " - " + alerts_df["clause_title"]
    alerts_df["alert_dimension"] = "Clause ISO"
    alerts_df["alert_key"] = alerts_df["clause_code"]

    return kpi_df, process_df, clause_df, alerts_df



def priority_sort_key(priority):
    order = {"Haute": 1, "Moyenne": 2, "Faible": 3}
    return order.get(str(priority), 9)


def recompute_kpi_from_checklist(checklist_df):
    """Recalcule les KPI à partir de la check-list réellement affichée.

    V7.8 : si on ajoute des questions contextuelles côté interface, les cartes KPI
    doivent refléter le lot affiché, pas seulement la sortie brute PostgreSQL.
    """
    if checklist_df is None or checklist_df.empty or "generated_priority" not in checklist_df.columns:
        return pd.DataFrame(columns=["generated_priority", "questions_count"])
    out = (
        checklist_df.groupby("generated_priority")
        .size()
        .reset_index(name="questions_count")
    )
    out["_order"] = out["generated_priority"].map(priority_sort_key)
    return out.sort_values("_order").drop(columns=["_order"]).reset_index(drop=True)


def context_dict_from_target_code(target_code, targets_df, contexts_df=None):
    """Récupère le contexte d'une mission cible à partir de son code."""
    default = {
        "sector": "",
        "audited_process": "",
        "audit_objective": "",
        "audit_scope": "",
        "specific_requirements": "",
        "known_risks": "",
        "keywords": "",
    }
    if targets_df is None or targets_df.empty:
        return default
    try:
        match = targets_df[targets_df["mission_code"].astype(str) == str(target_code)]
        if match.empty:
            return default
        row = match.iloc[0]
        out = default.copy()
        for k in out.keys():
            out[k] = normalize_blank(row.get(k)) or ""
        return out
    except Exception:
        return default


def contextual_questions_for_mission(context):
    """Génère des questions complémentaires selon le secteur/processus/risques.

    Cette couche ne remplace pas le moteur SQL V6. Elle ajoute une adaptation métier
    visible quand deux lots utilisent le même historique mais des missions cibles différentes.
    """
    sector = normalize_blank(context.get("sector")) or ""
    process = normalize_blank(context.get("audited_process")) or ""
    objective = normalize_blank(context.get("audit_objective")) or ""
    scope = normalize_blank(context.get("audit_scope")) or ""
    req = normalize_blank(context.get("specific_requirements")) or ""
    risks = normalize_blank(context.get("known_risks")) or ""
    keywords = normalize_blank(context.get("keywords")) or ""
    full = " ".join([sector, process, objective, scope, req, risks, keywords]).lower()

    rows = []

    def add(clause_code, clause_title, theme, question, priority, evidence):
        rows.append({
            "clause_code": clause_code,
            "clause_title": clause_title,
            "theme": theme,
            "question_text": question,
            "generated_priority": priority,
            "recommendation_label": "Complément contextuel V7.8",
            "expected_evidence": evidence,
            "conformity_status": "À vérifier",
            "source_context": "Contexte métier de la mission cible",
        })

    # Questions communes si un contexte existe.
    if any([sector, process, objective, scope, req, risks, keywords]):
        add(
            "6.1",
            "Actions face aux risques et opportunités",
            "Risques contextualisés",
            f"Les risques spécifiques au contexte '{sector or 'non précisé'}' et au processus '{process or 'non précisé'}' sont-ils identifiés et traités ?",
            "Haute",
            "Cartographie des risques, plan d’actions, responsables, échéances, preuves de suivi",
        )
        add(
            "7.5",
            "Informations documentées",
            "Preuves et traçabilité",
            "Les documents et preuves attendus dans le périmètre de cette mission sont-ils disponibles, à jour et maîtrisés ?",
            "Haute" if any(w in full for w in ["document", "preuve", "tracabil", "enregistrement"]) else "Moyenne",
            "Liste documentaire, procédures, versions, enregistrements, preuves de diffusion",
        )

    if any(w in full for w in ["hse", "securite", "sécurité", "environnement", "risque", "danger"]):
        add(
            "6.1",
            "Risques HSE et opportunités",
            "HSE / sécurité",
            "Les risques HSE liés aux activités auditées sont-ils évalués, maîtrisés et suivis par des actions vérifiables ?",
            "Haute",
            "DUERP/évaluation HSE, consignes sécurité, registres incidents, plans de prévention",
        )

    if any(w in full for w in ["ecole", "école", "formation", "apprenant", "eleve", "élève", "pédagog", "pedagog", "qualiopi"]):
        add(
            "7.2",
            "Compétences",
            "Formation / pédagogie",
            "Les compétences des intervenants et l’organisation pédagogique sont-elles définies, suivies et évaluées ?",
            "Haute",
            "CV/formations, habilitations, évaluations, dossiers apprenants, indicateurs satisfaction",
        )
        add(
            "9.1",
            "Surveillance, mesure et satisfaction",
            "Satisfaction bénéficiaires",
            "La satisfaction des apprenants/clients et les réclamations sont-elles suivies et exploitées ?",
            "Moyenne",
            "Questionnaires, réclamations, plans d’amélioration, indicateurs de suivi",
        )

    if any(w in full for w in ["textile", "production", "manufacturing", "atelier", "cuir", "leather"]):
        add(
            "8.5",
            "Production et prestation de service",
            "Production / atelier",
            "Les étapes de production, contrôles qualité et critères d’acceptation sont-ils définis et tracés ?",
            "Haute",
            "Gammes, fiches contrôle, traçabilité lots, résultats qualité, non-conformités produit",
        )

    if any(w in full for w in ["achat", "achats", "fournisseur", "prestataire", "sous traitant", "sous-traitant"]):
        add(
            "8.4",
            "Maîtrise des prestataires externes",
            "Achats / fournisseurs",
            "Les fournisseurs critiques sont-ils évalués, sélectionnés, surveillés et réévalués selon des critères définis ?",
            "Haute",
            "Évaluations fournisseurs, contrats, critères de sélection, suivis performance, plans d’actions",
        )

    if any(w in full for w in ["logistique", "transport", "stock", "livraison", "import", "export"]):
        add(
            "8.1",
            "Planification et maîtrise opérationnelles",
            "Logistique / transport",
            "Les flux logistiques, délais, contrôles et preuves de livraison sont-ils maîtrisés dans le périmètre audité ?",
            "Haute",
            "Procédures logistiques, bons de livraison, indicateurs délai, contrôles réception/expédition",
        )

    # Déduplication simple par clause + question.
    unique = []
    seen = set()
    for r in rows:
        key = (r["clause_code"], r["question_text"])
        if key not in seen:
            unique.append(r)
            seen.add(key)
    return unique


def augment_checklist_with_target_context(checklist_df, selected_run, targets_df, contexts_df=None):
    """Ajoute une couche d'adaptation métier au lot sélectionné.

    V7.8 : deux lots ayant le même historique peuvent sinon donner le même résultat.
    Cette fonction rend visible l'effet du contexte cible dans la check-list affichée
    et dans l'export, sans modifier le moteur SQL V6.
    """
    if checklist_df is None:
        checklist_df = pd.DataFrame()
    out = checklist_df.copy()
    target_code = str(selected_run.get("target_mission_code", ""))
    target_title = str(selected_run.get("target_mission_title", ""))
    batch_code = str(selected_run.get("generation_batch_code", ""))
    source_code = str(selected_run.get("source_mission_code", ""))
    source_title = str(selected_run.get("source_mission_title", ""))
    mode = str(selected_run.get("generation_mode", "Historique d’audit"))

    # Le mode premier audit a déjà sa check-list contextuelle.
    if "Premier audit" in mode or "AUCUN_HISTORIQUE" in source_code:
        return out, "Mode premier audit : la check-list est déjà générée depuis le contexte et le référentiel."

    context = context_dict_from_target_code(target_code, targets_df, contexts_df)
    questions = contextual_questions_for_mission(context)
    if not questions:
        return out, "Aucun contexte métier sauvegardé : le lot repose sur l’historique."

    existing_questions = set(out.get("question_text", pd.Series(dtype=str)).astype(str).str.lower().tolist()) if not out.empty else set()
    rows = []
    start_order = int(out["display_order"].max()) if not out.empty and "display_order" in out.columns and pd.notna(out["display_order"].max()) else 0
    for i, q in enumerate(questions, start=1):
        if str(q["question_text"]).lower() in existing_questions:
            continue
        rows.append({
            "generation_run_id": selected_run.get("generation_run_id", None),
            "generation_batch_code": batch_code,
            "target_mission_code": target_code,
            "target_mission_title": target_title,
            "source_mission_code": source_code,
            "source_mission_title": source_title,
            "display_order": start_order + i,
            "clause_code": q["clause_code"],
            "clause_title": q["clause_title"],
            "theme": q["theme"],
            "question_text": q["question_text"],
            "generated_priority": q["generated_priority"],
            "recommendation_label": q["recommendation_label"],
            "expected_evidence": q["expected_evidence"],
            "conformity_status": q["conformity_status"],
            "generation_mode": mode + " + contexte cible V7.8",
            "context_sector": context.get("sector", ""),
            "context_process": context.get("audited_process", ""),
            "context_objective": context.get("audit_objective", ""),
            "context_scope": context.get("audit_scope", ""),
        })
    if rows:
        out = pd.concat([out, pd.DataFrame(rows)], ignore_index=True, sort=False)
        return out, f"{len(rows)} question(s) contextuelle(s) ajoutée(s) selon le secteur, le processus et les risques de la mission cible."
    return out, "Contexte cible lu, mais aucune question contextuelle nouvelle n’a été ajoutée."


def add_context_rows_to_vigilance(process_df, clause_df, selected_run, targets_df):
    """Ajoute une ligne de vigilance contextuelle pour mieux distinguer les lots."""
    context = context_dict_from_target_code(str(selected_run.get("target_mission_code", "")), targets_df)
    process_name = normalize_blank(context.get("audited_process"))
    sector = normalize_blank(context.get("sector"))
    risks = normalize_blank(context.get("known_risks"))
    if not any([process_name, sector, risks]):
        return process_df, clause_df

    p = process_df.copy() if process_df is not None else pd.DataFrame()
    c = clause_df.copy() if clause_df is not None else pd.DataFrame()

    if process_name and "process_name" in p.columns and process_name not in p["process_name"].astype(str).tolist():
        p = pd.concat([pd.DataFrame([{
            "process_name": process_name,
            "findings_count": 0,
            "nonconformities_count": 0,
            "remarks_count": 0,
            "improvements_count": 0,
            "capped_score": 65 if risks else 55,
            "vigilance_level": "Moyenne",
            "explanation_summary": "Vigilance ajoutée par la couche contextuelle V7.8 à partir du processus de la mission cible.",
        }]), p], ignore_index=True, sort=False)

    if risks and "clause_code" in c.columns:
        c = pd.concat([pd.DataFrame([{
            "clause_code": "6.1",
            "clause_title": "Actions face aux risques et opportunités",
            "findings_count": 0,
            "capped_score": 70,
            "vigilance_level": "Élevée",
            "explanation_summary": "Vigilance contextuelle V7.8 : risques connus déclarés dans la fiche mission.",
        }]), c], ignore_index=True, sort=False)
    return p, c

def build_traceability_from_checklist(checklist_df, selected_run):
    """Construit une traçabilité lisible, différente de la check-list brute."""
    trace_cols = [
        "generation_run_id",
        "generation_batch_code",
        "target_mission_code",
        "target_mission_title",
        "source_mission_code",
        "source_mission_title",
        "display_order",
        "clause_code",
        "clause_title",
        "theme",
        "question_text",
        "rule_based_priority",
        "ml_predicted_priority",
        "ml_prediction_confidence",
        "ml_predicted_criticality",
        "ml_criticality_band",
        "ml_decision",
        "generated_priority",
        "priority_origin",
        "ml_review_required",
        "recommendation_label",
        "expected_evidence",
        "generation_mode",
        "context_sector",
        "context_process",
        "context_objective",
        "context_scope",
    ]
    trace_cols = [c for c in trace_cols if c in checklist_df.columns]
    trace_df = checklist_df[trace_cols].copy()

    source_code = str(selected_run.get("source_mission_code", ""))
    target_code = str(selected_run.get("target_mission_code", ""))
    mode = str(selected_run.get("generation_mode", "Historique d’audit"))

    if "AUCUN_HISTORIQUE" in source_code or "Premier audit" in mode:
        reason = (
            "Premier audit sans historique : priorité calculée à partir du référentiel, "
            "du processus audité et du contexte saisi par l’auditeur."
        )
    else:
        reason = (
            "Priorité calculée à partir du score de vigilance de l'historique "
            + source_code
            + " puis appliquée à la mission cible "
            + target_code
        )
    if (
        "ml_predicted_priority" in trace_df.columns
        and trace_df["ml_predicted_priority"].astype(str).ne("Non disponible").any()
    ):
        trace_df["raison_priorisation"] = (
            reason
            + " La V8.3 confronte ensuite classification, régression et règles métier ; "
            + "une prédiction peu fiable ou susceptible d'abaisser la vigilance reste soumise à validation humaine."
        )
    else:
        trace_df["raison_priorisation"] = reason
    return trace_df


def make_session_run_from_first_audit(checklist_df, target_code, target_title):
    """Crée une ligne de lot visible dans la liste sans dépendre de PostgreSQL."""
    if checklist_df is None or checklist_df.empty:
        return {}
    batch_code = checklist_df["generation_batch_code"].iloc[0]
    return {
        "generation_run_id": f"TEMP-{batch_code}",
        "generation_batch_code": batch_code,
        "generated_checklist_title": f"Check-list initiale sans historique - {target_code}",
        "target_mission_code": target_code,
        "target_mission_title": target_title,
        "source_mission_code": "AUCUN_HISTORIQUE",
        "source_mission_title": "Premier audit / génération sans historique",
        "checklist_items_count": int(checklist_df.shape[0]),
        "recommendations_count": int(checklist_df.shape[0]),
        "generated_at": datetime.now(),
        "generation_mode": "Premier audit sans historique",
    }


def make_run_label(row):
    dt = row.get("generated_at", "")
    try:
        dt = pd.to_datetime(dt).strftime("%d/%m/%Y %H:%M")
    except Exception:
        dt = str(dt)
    mode = resolve_generation_mode(row)
    mode_part = f" | {mode}" if mode else ""
    return (
        f"{row.get('generation_batch_code', '')} | "
        f"{row.get('target_mission_code', '')} ← {row.get('source_mission_code', '')} | {dt}{mode_part}"
    )


def is_first_audit_run(row):
    """Détecte un lot généré sans historique, même s'il est temporaire côté session."""
    source_code = str(row.get("source_mission_code", ""))
    mode = str(row.get("generation_mode", ""))
    batch = str(row.get("generation_batch_code", ""))
    return (
        "AUCUN_HISTORIQUE" in source_code
        or "Premier audit" in mode
        or batch.startswith("FIRST_AUDIT_SESSION_")
    )


def safe_display_value(value, default="Non précisé"):
    """Évite d'afficher nan/NaT/None dans les cartes de démonstration."""
    try:
        if value is None or pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if not text or text.lower() in {"nan", "nat", "none", "null"}:
        return default
    return text


def resolve_generation_mode(row):
    """Retourne un mode lisible, y compris pour les anciens lots incomplets."""
    raw_mode = safe_display_value(row.get("generation_mode", ""), "")
    if raw_mode:
        return raw_mode
    source = safe_display_value(row.get("source_mission_code", ""), "")
    batch = safe_display_value(row.get("generation_batch_code", ""), "")
    if "AUCUN_HISTORIQUE" in source or batch.startswith("FIRST_AUDIT_SESSION_"):
        return "Premier audit sans historique"
    if source:
        return "Audit préparé à partir d'un historique"
    return "Mode historique non renseigné"


def generation_mode_labels(row):
    """Libellés d'interface adaptés au mode de génération du lot sélectionné."""
    if is_first_audit_run(row):
        return {
            "source_caption": "Source utilisée",
            "source_value": "Aucun historique",
            "result_sentence": "Résultats affichés selon le lot sélectionné : génération initiale depuis le référentiel ISO et le contexte métier saisi.",
            "alerts_title": "Alertes prioritaires issues du premier audit",
            "vigilance_title": "Vigilance métier issue du référentiel et du contexte",
            "vigilance_caption": "Analyse des points de vigilance initiaux par processus et par clause ISO, sans constats historiques.",
            "exec_note": "ce lot ne réutilise aucun audit passé : les priorités viennent du référentiel ISO 9001, du processus audité et du contexte métier saisi.",
            "trace_note": "le moteur priorise les questions à partir du référentiel, du contexte de mission et des risques déclarés par l’auditeur.",
            "export_note": "L’export contient uniquement le lot sélectionné, la check-list initiale, les KPI, la vigilance référentiel/contexte et la traçabilité du raisonnement.",
        }
    return {
        "source_caption": "Source historique",
        "source_value": str(row.get("source_mission_code", "")),
        "result_sentence": "Résultats affichés selon le lot sélectionné : historique SQL, recommandation historique ou couche contextuelle de la mission cible.",
        "alerts_title": "Alertes prioritaires liées à l’historique",
        "vigilance_title": "Vigilance métier issue de la mission historique",
        "vigilance_caption": "Analyse des risques de préparation par processus métier et par clause ISO à partir des constats historiques.",
        "exec_note": "ce lot traduit automatiquement les signaux de l’audit historique en une préparation ciblée de la mission sélectionnée. Les scores de vigilance structurent les priorités de la check-list.",
        "trace_note": "le moteur retient le signal le plus fort entre la vigilance d’une clause ISO et celle d’un processus métier, puis le traduit en priorité de préparation.",
        "export_note": "L’export contient uniquement le lot sélectionné, la mission cible, les KPI, les alertes, les scores de vigilance, la check-list produite et la traçabilité du raisonnement.",
    }


def build_selected_run_export_df(selected_run, checklist_df, kpi_df, ml_pack=None):
    """Construit une feuille d’export limitée au lot sélectionné, avec compteurs recalculés après post-traitement."""
    row = dict(selected_run)
    questions_count = int(checklist_df.shape[0]) if checklist_df is not None else 0
    row["checklist_items_count"] = questions_count
    row["recommendations_count"] = questions_count
    row["displayed_questions_count"] = questions_count
    if kpi_df is not None and not kpi_df.empty and "generated_priority" in kpi_df.columns:
        for priority in ["Haute", "Moyenne", "Faible"]:
            match = kpi_df[kpi_df["generated_priority"].astype(str).str.lower() == priority.lower()]
            row[f"questions_priorite_{priority.lower()}"] = int(match.iloc[0].get("questions_count", 0)) if not match.empty else 0
    if is_first_audit_run(row):
        row["persistence_status"] = "Lot temporaire de session Streamlit - génération sans historique"
    else:
        row["persistence_status"] = "Lot issu de PostgreSQL avec post-traitement d’affichage éventuel"
    if ml_pack:
        row["ml_status"] = ml_pack.get("status", "")
        row["ml_training_scope"] = ml_pack.get("training_scope", "")
        row["ml_expert_rows_available"] = ml_pack.get("expert_rows_available", 0)
        row["ml_dataset_rows"] = ml_pack.get("dataset_rows", 0)
        row["ml_train_rows"] = ml_pack.get("train_rows", 0)
        row["ml_test_rows"] = ml_pack.get("test_rows", 0)
        row["ml_classification_accuracy"] = ml_pack.get("classification_metrics", {}).get("accuracy")
        row["ml_regression_mae"] = ml_pack.get("regression_metrics", {}).get("mae")
        row["ml_regression_rmse"] = ml_pack.get("regression_metrics", {}).get("rmse")
        row["ml_regression_r2"] = ml_pack.get("regression_metrics", {}).get("r2")
    return pd.DataFrame([row])



def make_option_label(row):
    audit_date = row.get("audit_date", "")
    if pd.notna(audit_date) and audit_date != "":
        try:
            audit_date = pd.to_datetime(audit_date).strftime("%d/%m/%Y")
        except Exception:
            audit_date = str(audit_date)
    else:
        audit_date = "date non renseignée"

    return (
        f"{row.get('mission_code', '')} — {row.get('mission_title', '')} | "
        f"{row.get('client_name', '')} | {audit_date}"
    )


def excel_bytes(sheets):
    buffer = io.BytesIO()

    column_labels = {
        "generation_run_id": "ID génération",
        "generation_batch_code": "Code du lot généré",
        "generated_checklist_title": "Titre de la check-list",
        "checklist_title": "Titre de la check-list",
        "target_mission_code": "Code mission cible",
        "target_mission_title": "Titre mission cible",
        "source_mission_code": "Code mission historique",
        "source_mission_title": "Titre mission historique",
        "recommendations_count": "Nombre de recommandations",
        "checklist_items_count": "Nombre de questions",
        "generated_at": "Date de génération",
        "ml_training_scope": "Périmètre d'entraînement ML",
        "ml_expert_rows_available": "Étiquettes expertes disponibles",

        "display_order": "Ordre",
        "clause_code": "Clause ISO",
        "clause_title": "Titre clause",
        "theme": "Thème",
        "question_text": "Question d’audit",
        "generated_priority": "Priorité générée",
        "rule_based_priority": "Priorité règles métier V7.10",
        "ml_predicted_priority": "Priorité prédite ML",
        "ml_prediction_confidence": "Confiance ML (%)",
        "ml_predicted_criticality": "Criticité prédite ML (/100)",
        "ml_criticality_band": "Classe issue de la criticité",
        "ml_decision": "Décision de fusion règles/ML",
        "priority_origin": "Origine de la priorité finale",
        "ml_review_required": "Validation auditeur requise",
        "recommendation_label": "Recommandation",
        "expected_evidence": "Preuves attendues",
        "conformity_status": "Statut conformité",

        "process_name": "Processus",
        "findings_count": "Nombre de constats",
        "nonconformities_count": "Non-conformités",
        "remarks_count": "Remarques",
        "improvements_count": "Améliorations",
        "open_corrective_actions_count": "Actions correctives ouvertes",
        "raw_score": "Score brut",
        "capped_score": "Score final",
        "vigilance_level": "Niveau de vigilance",
        "explanation_summary": "Justification du score",
        "computed_at": "Date de calcul",
        "generation_mode": "Mode de génération",
        "context_sector": "Secteur cible",
        "context_process": "Processus cible",
        "context_objective": "Objectif cible",
        "context_scope": "Périmètre cible",
        "source_context": "Origine contextuelle",
        "raison_priorisation": "Raison de priorisation",
        "target_priority": "Cible classification",
        "target_criticality": "Cible régression",
        "target_source": "Origine de la cible",
    }

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            clean = clean_df(df)

            # Renommer les colonnes techniques en français
            clean = clean.rename(columns=column_labels)

            safe_sheet_name = str(sheet_name)[:31]

            clean.to_excel(
                writer,
                index=False,
                sheet_name=safe_sheet_name
            )

            ws = writer.book[safe_sheet_name]

            # Figer la première ligne
            ws.freeze_panes = "A2"

            # Activer le filtre automatique
            if ws.max_row > 1 and ws.max_column > 1:
                ws.auto_filter.ref = ws.dimensions

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            header_fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78"
            )

            header_font = Font(
                bold=True,
                color="FFFFFF"
            )

            thin_border = Border(
                left=Side(style="thin", color="D9E2F3"),
                right=Side(style="thin", color="D9E2F3"),
                top=Side(style="thin", color="D9E2F3"),
                bottom=Side(style="thin", color="D9E2F3")
            )

            # Style général
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )
                    cell.border = thin_border

            # Style de l’en-tête
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

            ws.row_dimensions[1].height = 32

            # Important : on garde le filtre automatique simple.
            # On n'ajoute PAS de tableau Excel natif, car certaines versions d'Excel
            # réparent/suppriment les tables générées automatiquement par openpyxl
            # quand il existe des en-têtes longs, similaires ou issus de vues SQL.

            # Largeur automatique des colonnes
            for column_cells in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    value = cell.value
                    if value is not None:
                        value_length = len(str(value))
                        if value_length > max_length:
                            max_length = value_length

                adjusted_width = min(max(max_length + 2, 14), 55)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Hauteur raisonnable pour les lignes longues
            for row_number in range(2, ws.max_row + 1):
                ws.row_dimensions[row_number].height = 38

    return buffer.getvalue()
# ============================================================
# 4. REQUÊTES SQL
# ============================================================

SQL_AVAILABLE_TARGET_MISSIONS = """
SELECT
    mission_id,
    mission_code,
    mission_title,
    COALESCE(client_name, 'Client non renseigné') AS client_name,
    COALESCE(site_name, 'Site non renseigné') AS site_name,
    planned_audit_date AS audit_date,
    COALESCE(standard_label, 'Référentiel non renseigné') AS standard_name
FROM auditprep.vw_available_target_missions
ORDER BY planned_audit_date DESC NULLS LAST, mission_code;
"""

SQL_AVAILABLE_HISTORICAL_MISSIONS = """
SELECT
    mission_id,
    mission_code,
    mission_title,
    COALESCE(client_name, 'Client non renseigné') AS client_name,
    COALESCE(site_name, 'Site non renseigné') AS site_name,
    planned_audit_date AS audit_date,
    findings_count
FROM auditprep.vw_available_historical_missions
ORDER BY planned_audit_date DESC NULLS LAST, mission_code;
"""

SQL_GENERATION_RUNS = """
SELECT *
FROM auditprep.vw_dynamic_generation_runs
ORDER BY generated_at DESC;
"""

SQL_KPI_BY_PRIORITY = """
SELECT *
FROM auditprep.vw_dynamic_smart_checklist_kpi_by_priority
WHERE generation_batch_code = %s
ORDER BY
    CASE generated_priority
        WHEN 'Haute' THEN 1
        WHEN 'Moyenne' THEN 2
        WHEN 'Faible' THEN 3
        ELSE 4
    END;
"""

SQL_CHECKLIST_ITEMS = """
SELECT *
FROM auditprep.vw_dynamic_smart_checklist_items
WHERE generation_batch_code = %s
ORDER BY display_order;
"""

SQL_TRACEABILITY = """
-- Fallback V6.2 : la vue auditprep.vw_dynamic_smart_checklist_traceability
-- n'existe pas dans certaines bases. On reconstruit donc une traçabilité simple
-- depuis la vue des items de check-list, qui existe déjà dans le moteur V6.
SELECT *
FROM auditprep.vw_dynamic_smart_checklist_items
WHERE generation_batch_code = %s
ORDER BY display_order;
"""

SQL_PROCESS_VIGILANCE = """
SELECT *
FROM auditprep.vw_dynamic_process_vigilance_dashboard
WHERE source_mission_code = %s
ORDER BY capped_score DESC, process_name;
"""

SQL_CLAUSE_VIGILANCE = """
SELECT *
FROM auditprep.vw_dynamic_clause_vigilance_dashboard
WHERE source_mission_code = %s
ORDER BY capped_score DESC, clause_code;
"""

SQL_TOP_ALERTS = """
SELECT *
FROM auditprep.vw_dynamic_top_vigilance_alerts
WHERE source_mission_code = %s
ORDER BY capped_score DESC, alert_dimension, alert_key
LIMIT 8;
"""


# ============================================================
# 5. AUTHENTIFICATION ET CONFIGURATION SÉCURISÉE
# ============================================================

PASSWORD_SCHEME = "pbkdf2_sha256"
PASSWORD_ITERATIONS = 600_000


def _hash_password(password):
    """Produit une empreinte PBKDF2 salée pour le premier compte entreprise."""
    salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt, PASSWORD_ITERATIONS
    )
    return "{}${}${}${}".format(
        PASSWORD_SCHEME,
        PASSWORD_ITERATIONS,
        base64.urlsafe_b64encode(salt).decode("ascii"),
        base64.urlsafe_b64encode(digest).decode("ascii"),
    )


def _toml_string(value):
    return '"' + str(value).replace("\\", "\\\\").replace('"', '\\"') + '"'


def _project_root():
    script_dir = Path(__file__).resolve().parent
    return script_dir.parent if script_dir.name.lower() == "app" else script_dir


def render_brand_nav(active="Accueil", user_info=None):
    """Barre de navigation interne AuditPrep avec liens réellement cliquables."""
    if user_info:
        right_label = (
            f"{html.escape(str(user_info.get('display_name', 'Utilisateur')))} · "
            f"{html.escape(str(user_info.get('role', 'Auditeur')))}"
        )
    else:
        right_label = "Accès sécurisé"

    def _nav_link(label, anchor):
        active_class = " audit-nav-active" if active == label else ""
        return (
            f'<a class="audit-nav-link{active_class}" '
            f'href="#{anchor}" target="_self">{html.escape(label)}</a>'
        )

    nav_html = "".join(
        [
            _nav_link("Accueil", "audit-accueil"),
            _nav_link("Préparation", "audit-preparation"),
            _nav_link("Résultats", "audit-resultats"),
            _nav_link("Espace audit", "audit-espace"),
        ]
    )

    st.markdown(
        f"""
<div class="audit-topbar">
    <div class="audit-brand">
        <span class="audit-brand-mark">AP</span>
        <span class="audit-brand-name">AuditPrep IA</span>
    </div>
    <nav class="audit-nav-items" aria-label="Navigation AuditPrep">
        {nav_html}
        <span class="audit-nav-user">{right_label}</span>
    </nav>
</div>
        """,
        unsafe_allow_html=True,
    )


def _save_first_configuration(auth_data, database_data):
    """Crée automatiquement le fichier local de secrets au premier lancement."""
    secrets_dir = _project_root() / ".streamlit"
    secrets_path = secrets_dir / "secrets.toml"
    username = auth_data["initial_username"]
    account = auth_data["users"][username]
    lines = [
        "# Configuration locale AuditPrep IA - ne jamais partager ce fichier",
        "[auth]",
        f"company_name = {_toml_string(auth_data['company_name'])}",
        f"session_timeout_minutes = {int(auth_data['session_timeout_minutes'])}",
        "",
        f"[auth.users.{username}]",
        f"display_name = {_toml_string(account['display_name'])}",
        f"role = {_toml_string(account['role'])}",
        f"password_hash = {_toml_string(account['password_hash'])}",
        "active = true",
        "",
        "[database]",
        f"host = {_toml_string(database_data['host'])}",
        f"port = {int(database_data['port'])}",
        f"dbname = {_toml_string(database_data['dbname'])}",
        f"user = {_toml_string(database_data['user'])}",
        f"password = {_toml_string(database_data['password'])}",
        "",
    ]
    secrets_dir.mkdir(parents=True, exist_ok=True)
    secrets_path.write_text("\n".join(lines), encoding="utf-8")


def _render_first_run_setup():
    """Assistant visuel intégré : aucun second fichier Python n'est nécessaire."""
    render_brand_nav("Configuration")
    st.markdown(
        """
<div class="audit-login-shell">
    <div class="audit-login-mark">⚙</div>
    <div class="audit-login-company">Première utilisation</div>
    <div class="audit-login-title">Configurer AuditPrep IA</div>
    <div class="audit-subtitle">Crée le compte administrateur et relie la base PostgreSQL.</div>
</div>
        """,
        unsafe_allow_html=True,
    )

    left, center, right = st.columns([0.7, 1.6, 0.7])
    with center:
        with st.form("auditprep_first_setup"):
            st.markdown("#### Identité de l'application")
            company_name = st.text_input("Entreprise", value="Convergence")

            st.markdown("#### Compte administrateur")
            c1, c2 = st.columns(2)
            with c1:
                admin_username = st.text_input("Identifiant", value="admin")
            with c2:
                admin_name = st.text_input("Nom affiché", value="Administrateur AuditPrep")
            admin_password = st.text_input("Mot de passe", type="password")
            admin_confirmation = st.text_input("Confirmer le mot de passe", type="password")

            st.markdown("#### Connexion PostgreSQL")
            d1, d2 = st.columns([2, 1])
            with d1:
                db_host = st.text_input("Hôte PostgreSQL", value="localhost")
            with d2:
                db_port = st.number_input("Port", min_value=1, max_value=65535, value=5432)
            d3, d4 = st.columns(2)
            with d3:
                db_name = st.text_input("Base de données", value="auditprep_ia")
            with d4:
                db_user = st.text_input("Utilisateur PostgreSQL", value="postgres")
            db_password = st.text_input("Mot de passe PostgreSQL", type="password")

            create_config = st.form_submit_button(
                "Créer la configuration sécurisée",
                type="primary",
                use_container_width=True,
            )

        if create_config:
            normalized_username = admin_username.strip().lower()
            errors = []
            if not re.fullmatch(r"[a-z0-9_-]{3,30}", normalized_username):
                errors.append("L'identifiant doit contenir 3 à 30 lettres, chiffres, _ ou -.")
            if not company_name.strip() or not admin_name.strip():
                errors.append("Le nom de l'entreprise et le nom affiché sont obligatoires.")
            if len(admin_password) < 10:
                errors.append("Le mot de passe administrateur doit contenir au moins 10 caractères.")
            if admin_password != admin_confirmation:
                errors.append("Les deux mots de passe administrateur ne correspondent pas.")
            if not db_host.strip() or not db_name.strip() or not db_user.strip() or not db_password:
                errors.append("Tous les paramètres PostgreSQL sont obligatoires.")

            if errors:
                for error in errors:
                    st.error(error)
                st.stop()

            account = {
                "display_name": admin_name.strip(),
                "role": "Administrateur",
                "password_hash": _hash_password(admin_password),
                "active": True,
            }
            auth_data = {
                "company_name": company_name.strip(),
                "session_timeout_minutes": 60,
                "initial_username": normalized_username,
                "users": {normalized_username: account},
            }
            database_data = {
                "host": db_host.strip(),
                "port": int(db_port),
                "dbname": db_name.strip(),
                "user": db_user.strip(),
                "password": db_password,
            }
            try:
                _save_first_configuration(auth_data, database_data)
            except Exception as exc:
                st.error("Impossible d'enregistrer la configuration locale.")
                st.code(str(exc))
                st.stop()

            st.session_state.auditprep_bootstrap = {
                "auth": auth_data,
                "database": database_data,
            }
            st.success("Configuration créée. Ouverture de l'écran de connexion…")
            time.sleep(0.5)
            st.rerun()

        st.caption(
            "Cette étape n'apparaît qu'au premier lancement. Les mots de passe ne sont pas affichés."
        )
    st.stop()


def _verify_password(password, encoded_hash):
    """Vérifie un secret PBKDF2 sans stocker de mot de passe en clair."""
    try:
        scheme, iterations, salt_b64, digest_b64 = str(encoded_hash).split("$", 3)
        if scheme != PASSWORD_SCHEME:
            return False
        salt = base64.urlsafe_b64decode(salt_b64.encode("ascii"))
        expected = base64.urlsafe_b64decode(digest_b64.encode("ascii"))
        candidate = hashlib.pbkdf2_hmac(
            "sha256", password.encode("utf-8"), salt, int(iterations)
        )
        return hmac.compare_digest(candidate, expected)
    except Exception:
        return False


def _auth_settings():
    bootstrap = st.session_state.get("auditprep_bootstrap")
    if bootstrap:
        auth = bootstrap["auth"]
        return auth, auth["users"]
    try:
        auth = st.secrets["auth"]
        users = auth["users"]
        return auth, users
    except Exception:
        return None, None


def _clear_authenticated_session():
    for key in (
        "auditprep_user",
        "auditprep_last_activity",
        "loaded",
        "selected_batch_code",
        "last_generated_batch_code",
    ):
        st.session_state.pop(key, None)


def require_authentication():
    """Bloque toute l'application tant qu'un utilisateur autorisé n'est pas connecté."""
    auth, configured_users = _auth_settings()

    if auth is None or configured_users is None:
        _render_first_run_setup()

    timeout_minutes = int(auth.get("session_timeout_minutes", 60))
    current_user = st.session_state.get("auditprep_user")
    last_activity = float(st.session_state.get("auditprep_last_activity", 0.0))
    now = time.time()

    if current_user and last_activity and now - last_activity > timeout_minutes * 60:
        _clear_authenticated_session()
        current_user = None
        st.warning("La session a expiré. Connecte-toi de nouveau.")

    if current_user:
        st.session_state.auditprep_last_activity = now
        return current_user

    company_name = str(auth.get("company_name", "Convergence"))
    render_brand_nav("Connexion")
    st.markdown(
        f"""
<div class="audit-login-shell">
    <div class="audit-login-mark">✓</div>
    <div class="audit-login-company">{company_name}</div>
    <div class="audit-login-title">AuditPrep IA</div>
    <div class="audit-subtitle">Espace sécurisé de préparation des audits</div>
</div>
        """,
        unsafe_allow_html=True,
    )

    locked_until = float(st.session_state.get("auditprep_locked_until", 0.0))
    if locked_until > now:
        remaining = max(1, int(locked_until - now))
        st.error(f"Trop de tentatives. Réessaie dans {remaining} seconde(s).")
        st.stop()

    left, center, right = st.columns([1.25, 1, 1.25])
    with center:
        with st.form("auditprep_login", clear_on_submit=False):
            username = st.text_input("Identifiant", autocomplete="username")
            password = st.text_input(
                "Mot de passe", type="password", autocomplete="current-password"
            )
            submitted = st.form_submit_button(
                "Se connecter", type="primary", use_container_width=True
            )

        if submitted:
            normalized_username = username.strip().lower()
            account = configured_users.get(normalized_username)
            account_hash = account.get("password_hash", "") if account else ""

            if account and bool(account.get("active", True)) and _verify_password(password, account_hash):
                st.session_state.auditprep_user = {
                    "username": normalized_username,
                    "display_name": str(account.get("display_name", normalized_username)),
                    "role": str(account.get("role", "Auditeur")),
                }
                st.session_state.auditprep_last_activity = now
                st.session_state.auditprep_failed_logins = 0
                st.session_state.pop("auditprep_locked_until", None)
                st.rerun()

            failures = int(st.session_state.get("auditprep_failed_logins", 0)) + 1
            st.session_state.auditprep_failed_logins = failures
            if failures >= 5:
                st.session_state.auditprep_failed_logins = 0
                st.session_state.auditprep_locked_until = now + 30
                st.error("Cinq tentatives incorrectes. Accès temporairement verrouillé.")
            else:
                st.error("Identifiant ou mot de passe incorrect.")

        st.caption(
            "Accès réservé aux collaborateurs autorisés. Les mots de passe ne sont jamais enregistrés en clair."
        )
    st.stop()


def read_database_settings():
    """Charge le compte technique PostgreSQL depuis les secrets ou l'environnement."""
    bootstrap = st.session_state.get("auditprep_bootstrap")
    if bootstrap:
        database = bootstrap["database"]
    else:
        try:
            database = st.secrets["database"]
        except Exception:
            database = {}

    settings = {
        "host": str(database.get("host", os.getenv("AUDITPREP_DB_HOST", "localhost"))),
        "port": int(database.get("port", os.getenv("AUDITPREP_DB_PORT", "5432"))),
        "dbname": str(database.get("dbname", os.getenv("AUDITPREP_DB_NAME", "auditprep_ia"))),
        "user": str(database.get("user", os.getenv("AUDITPREP_DB_USER", ""))),
        "password": str(database.get("password", os.getenv("AUDITPREP_DB_PASSWORD", ""))),
    }
    if not settings["user"] or not settings["password"]:
        st.error("Le compte technique PostgreSQL n'est pas configuré.")
        st.info("Renseigne la section `[database]` du fichier `.streamlit/secrets.toml`.")
        st.stop()
    return settings


current_user = require_authentication()
database_settings = read_database_settings()
host = database_settings["host"]
port = database_settings["port"]
dbname = database_settings["dbname"]
user = database_settings["user"]
password = database_settings["password"]
is_admin = current_user["role"].strip().lower() in {
    "administrateur",
    "admin",
    "responsable qualité",
}


def show_technical_error(error):
    """Évite d'exposer les détails internes aux comptes non administrateurs."""
    if is_admin:
        st.code(str(error))
    else:
        st.caption("Un détail technique a été masqué. Contacte l'administrateur AuditPrep.")


# ============================================================
# 6. SIDEBAR ENTREPRISE
# ============================================================

with st.sidebar:
    st.markdown("## AuditPrep IA")
    st.markdown(
        f"""
<div class="audit-user-chip audit-requested-black">
    <b>{current_user['display_name']}</b><br>
    <span>{current_user['role']}</span>
</div>
        """,
        unsafe_allow_html=True,
    )

    load_btn = st.button("Charger les missions", use_container_width=True, type="primary")
    if st.button("Se déconnecter", use_container_width=True):
        _clear_authenticated_session()
        st.rerun()

    st.markdown("---")
    st.markdown("### Version")
    st.caption("AuditPrep V8.5.28 — compatibilité Cloud, contraste Light/Dark, moteur SQL V6 et IA supervisée par l’auditeur.")

render_brand_nav("Espace audit", current_user)


# ============================================================
# 7. ÉTAT INITIAL
# ============================================================

if "loaded" not in st.session_state:
    st.session_state.loaded = False

if load_btn:
    st.session_state.loaded = True
    st.cache_data.clear()


# ============================================================
# 8. PAGE D'ACCUEIL SANS CHARGEMENT
# ============================================================

if not st.session_state.loaded:
    st.markdown(
        """
<div class="audit-hero">
    <div class="audit-hero-title">Préparez vos audits avec une IA explicable</div>
    <div class="audit-hero-subtitle">
        AuditPrep transforme les constats historiques en une check-list priorisée,
        traçable et contrôlée par l’auditeur.
    </div>
</div>
        """,
        unsafe_allow_html=True,
    )

    cta_left, cta_center, cta_right = st.columns([1.2, 1, 1.2])
    with cta_center:
        if st.button(
            "Démarrer une préparation",
            type="primary",
            use_container_width=True,
            key="v85_home_start",
        ):
            st.session_state.loaded = True
            st.cache_data.clear()
            st.rerun()

    st.markdown(
        """
<div class="audit-grid-3">
    <div class="audit-mini-card">
        <div class="audit-kpi-value">3</div>
        <div class="audit-kpi-label">Modes de génération</div>
        <p>Historique manuel, recommandé ou premier audit.</p>
    </div>
    <div class="audit-mini-card">
        <div class="audit-kpi-value">2</div>
        <div class="audit-kpi-label">Modèles supervisés</div>
        <p>Classification de priorité et régression de criticité.</p>
    </div>
    <div class="audit-mini-card">
        <div class="audit-kpi-value">1</div>
        <div class="audit-kpi-label">Livrable consolidé</div>
        <p>Une check-list justifiée, contrôlable et exportable.</p>
    </div>
</div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("## Comment ça fonctionne ?")
    st.markdown(
        """
<div class="audit-grid-3">
    <div class="audit-mini-card">
        <div class="audit-step-bubble">1</div>
        <h3>Sélectionner</h3>
        <p>Choisissez la mission cible et le mode de préparation.</p>
    </div>
    <div class="audit-mini-card">
        <div class="audit-step-bubble">2</div>
        <h3>Analyser</h3>
        <p>Le moteur confronte historique, règles métier et modèles supervisés.</p>
    </div>
    <div class="audit-mini-card">
        <div class="audit-step-bubble">3</div>
        <h3>Contrôler</h3>
        <p>L’auditeur valide les priorités puis exporte la check-list.</p>
    </div>
</div>
<div class="audit-warning">
    <b>Principe de sécurité :</b> l’IA assiste la préparation ; l’auditeur reste décisionnaire.
</div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()


# ============================================================
# 9. CHARGEMENT DES DONNÉES
# ============================================================

try:
    targets_df = read_sql_cached(SQL_AVAILABLE_TARGET_MISSIONS, host, port, dbname, user, password)
    sources_df = read_sql_cached(SQL_AVAILABLE_HISTORICAL_MISSIONS, host, port, dbname, user, password)
    runs_df = read_sql_cached(SQL_GENERATION_RUNS, host, port, dbname, user, password)
except Exception as e:
    st.error("Connexion ou chargement impossible.")
    if is_admin:
        with st.expander("Détails techniques réservés à l'administrateur"):
            show_technical_error(e)
    st.stop()

if targets_df.empty:
    st.warning("Aucune mission cible disponible. Vérifie les vues SQL V6 ou crée une mission depuis PostgreSQL.")
    st.stop()

if sources_df.empty:
    st.info(
        "Aucune mission historique exploitable n’est encore disponible. "
        "Tu peux quand même utiliser le mode **Premier audit / sans historique**."
    )

# Listes de référence réelles PostgreSQL pour éviter les erreurs FK/NOT NULL.
try:
    audit_type_labels = get_reference_labels_cached(host, port, dbname, user, password, "audit_types")
except Exception:
    audit_type_labels = ["Audit interne", "Audit de conformité", "Audit système", "Audit processus", "Audit fournisseur"]

try:
    mission_status_labels = get_reference_labels_cached(host, port, dbname, user, password, "mission_statuses")
except Exception:
    mission_status_labels = ["Brouillon", "En cours", "À compléter", "Presque prêt", "Prêt", "Archivé"]

if not audit_type_labels:
    audit_type_labels = ["Audit interne"]
if not mission_status_labels:
    mission_status_labels = ["Brouillon"]


# V7.8 : contexte métier persistant. La table est créée automatiquement si besoin.
try:
    mission_contexts_df = read_mission_contexts_cached(host, port, dbname, user, password)
except Exception as e:
    mission_contexts_df = pd.DataFrame()
    st.warning("Le contexte métier persistant n'a pas pu être chargé. L'application reste utilisable, mais les recommandations seront moins précises.")
    show_technical_error(e)

targets_df = enrich_missions_with_context(targets_df, mission_contexts_df)
sources_df = enrich_missions_with_context(sources_df, mission_contexts_df)

# V7.9 : on enrichit les historiques avec les signaux calculés par le moteur SQL
# (processus et clauses sensibles), même si aucun contexte manuel n'a encore été saisi
# pour ces anciennes missions.
try:
    history_signals_df = read_history_signals_cached(
        tuple(sources_df["mission_code"].astype(str).tolist()) if not sources_df.empty else tuple(),
        host, port, dbname, user, password,
    )
    if not history_signals_df.empty:
        sources_df = sources_df.merge(
            history_signals_df,
            left_on="mission_code",
            right_on="source_mission_code",
            how="left",
        )
    for col in ["history_process_signals", "history_clause_signals", "history_top_signal"]:
        if col not in sources_df.columns:
            sources_df[col] = ""
        else:
            sources_df[col] = sources_df[col].fillna("")
except Exception:
    for col in ["history_process_signals", "history_clause_signals", "history_top_signal"]:
        if col not in sources_df.columns:
            sources_df[col] = ""


# ============================================================
# 9. HEADER
# ============================================================

st.markdown('<div id="audit-accueil" class="audit-anchor"></div>', unsafe_allow_html=True)

st.markdown(
    """
<div class="audit-hero">
    <div class="audit-hero-title">AuditPrep IA – Prédire la criticité pour mieux préparer l’audit</div>
    <div class="audit-hero-subtitle">
        Capitalise les constats historiques, entraîne les modèles supervisés, sélectionne une mission cible,
        puis consulte les prédictions, KPI, alertes, check-list produite et traçabilité.
    </div>
</div>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    '<div class="audit-section-head audit-section-head--why"><h2>Pourquoi cet outil ?</h2></div>',
    unsafe_allow_html=True,
)
st.markdown(
    """
<div class="audit-grid-3">
    <div class="audit-mini-card"><b>PROBLÈME MÉTIER</b><h3>Préparation trop manuelle</h3><p>Les constats passés sont peu exploités.</p></div>
    <div class="audit-mini-card"><b>RÉPONSE AUDITPREP</b><h3>Capitaliser les signaux</h3><p>L’IA priorise les risques de façon explicable.</p></div>
    <div class="audit-mini-card"><b>RÉSULTAT</b><h3>Une check-list ciblée</h3><p>La check-list est priorisée et exportable.</p></div>
</div>
    """,
    unsafe_allow_html=True,
)
st.markdown(
    '<div class="audit-section-head audit-section-head--chain"><h2>Chaîne de valeur du prototype</h2></div>',
    unsafe_allow_html=True,
)
st.markdown(
    """
<div class="audit-grid-4">
    <div class="audit-mini-card"><div class="audit-step-bubble">1</div><h3>Données</h3><p>Missions et constats.</p></div>
    <div class="audit-mini-card"><div class="audit-step-bubble">2</div><h3>Apprentissage</h3><p>Classification et régression.</p></div>
    <div class="audit-mini-card"><div class="audit-step-bubble">3</div><h3>Fusion prudente</h3><p>Règles et prédictions confrontées.</p></div>
    <div class="audit-mini-card"><div class="audit-step-bubble">4</div><h3>Livrable</h3><p>Contrôle, traçabilité et export.</p></div>
</div>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# 10. ALIMENTER LES DONNÉES D'AUDIT DEPUIS L'APPLICATION
# ============================================================

st.markdown('<hr class="audit-separator">', unsafe_allow_html=True)
st.markdown(
    '<div class="audit-section-head audit-section-head--tools"><h2>Outils avancés : gérer les données</h2></div>',
    unsafe_allow_html=True,
)
st.caption(
    "Cette zone sert à créer les missions et alimenter la base de constats."
)
st.markdown(
    """
<div class="audit-note">
<b>Différence importante :</b> cette partie sert à remplir la base de connaissances. Une mission créée seule peut être préparée comme cible ; elle devient historique exploitable seulement lorsqu’elle possède des constats ou un rapport d’audit.
</div>
    """,
    unsafe_allow_html=True,
)

with st.expander("Créer une mission, ajouter ou importer des constats", expanded=False):
    if not is_admin:
        st.info(
            "Consultation autorisée. La création et l'import de données sont réservés "
            "au rôle Administrateur."
        )
    data_tabs = st.tabs([
        "Créer une mission",
        "Ajouter un constat",
        "Importer constats Excel/CSV",
        "Modèle fichier"
    ])

    with data_tabs[0]:
        st.markdown("### Créer une mission cible ou historique")
        st.caption(
            "V7.8 : la mission n’est plus seulement administrative. On enregistre aussi le contexte métier "
            "pour que le moteur puisse comparer textile, école, logistique, formation, industrie, etc."
        )

        c1, c2 = st.columns(2)
        with c1:
            new_mission_code = st.text_input("Code mission", placeholder="Exemple : AUD-2026-002", key="v7_mission_code")
            new_mission_title = st.text_input("Titre mission", placeholder="Exemple : Audit préparation ISO 9001 – Processus achats", key="v7_mission_title")
            new_client_name = st.text_input("Client", placeholder="Exemple : Entreprise Beta", key="v7_client_name")
            new_site_name = st.text_input("Site", placeholder="Exemple : Site principal", key="v7_site_name")
        with c2:
            new_audit_type = st.selectbox(
                "Type d’audit",
                audit_type_labels,
                key="v7_audit_type",
            )
            new_standard_name = st.text_input("Référentiel", placeholder="Exemple : ISO 9001:2015", key="v7_standard")
            new_audit_date = st.date_input("Date prévue de l’audit", key="v7_audit_date")
            new_status = st.selectbox("Statut", mission_status_labels, key="v7_status")

        st.markdown("#### Contexte métier de la mission")
        mctx1, mctx2 = st.columns(2)
        with mctx1:
            new_sector = st.text_input(
                "Secteur / domaine d’activité",
                placeholder="Exemple : textile, école, logistique, formation, industrie...",
                key="v77_new_sector",
            )
            new_audited_process = st.text_input(
                "Processus audité",
                placeholder="Exemple : achats, RH, production, formation, maintenance...",
                key="v77_new_audited_process",
            )
            new_requirements = st.text_area(
                "Exigences spécifiques",
                placeholder="Exemple : ISO 9001, HSE, conformité réglementaire, exigences client, Qualiopi...",
                key="v77_new_requirements",
                height=95,
            )
        with mctx2:
            new_objective = st.text_area(
                "Objectif de l’audit",
                placeholder="Exemple : vérifier la conformité du processus achats et identifier les risques prioritaires.",
                key="v77_new_objective",
                height=95,
            )
            new_scope = st.text_area(
                "Périmètre de l’audit",
                placeholder="Exemple : site principal, service achats, fournisseurs critiques, documents associés.",
                key="v77_new_scope",
                height=95,
            )
            new_known_risks = st.text_area(
                "Risques connus / points d’attention",
                placeholder="Exemple : fournisseurs non évalués, documents non à jour, absence de preuves, sécurité...",
                key="v77_new_known_risks",
                height=95,
            )

        new_keywords = st.text_input(
            "Mots-clés métier",
            placeholder="Exemple : achats, fournisseur, textile, HSE, école, formation, traçabilité...",
            key="v77_new_keywords",
        )

        if not normalize_blank(new_mission_title):
            st.warning("Le titre mission est obligatoire pour PostgreSQL.")

        if st.button(
            "Enregistrer la mission et son contexte dans PostgreSQL",
            key="v7_save_mission",
            disabled=not is_admin,
        ):
            try:
                mission_id, state = create_or_update_mission_from_app(
                    host, port, dbname, user, password,
                    {
                        "mission_code": new_mission_code,
                        "mission_title": new_mission_title,
                        "client_name": new_client_name,
                        "site_name": new_site_name,
                        "audit_type": new_audit_type,
                        "standard_name": new_standard_name,
                        "planned_audit_date": new_audit_date,
                        "status": new_status,
                        "sector": new_sector,
                        "audited_process": new_audited_process,
                        "audit_objective": new_objective,
                        "audit_scope": new_scope,
                        "specific_requirements": new_requirements,
                        "known_risks": new_known_risks,
                        "keywords": new_keywords,
                    },
                )
                st.success(f"Mission {state} avec succès. ID interne : {mission_id}")
                st.cache_data.clear()
                st.info("Clique ensuite sur Charger les missions dans la barre latérale pour rafraîchir les listes.")
            except Exception as e:
                st.error("Impossible d’enregistrer la mission.")
                show_technical_error(e)

    with data_tabs[1]:
        st.markdown("### Ajouter un constat à une mission historique")
        st.caption("Une mission devient utilisable comme historique dès qu’elle possède au moins un rapport et un constat exploitable.")

        hist_codes = list(sources_df["mission_code"].astype(str).unique())
        target_codes_for_manual = list(targets_df["mission_code"].astype(str).unique())
        all_codes = sorted(set(hist_codes + target_codes_for_manual))

        c1, c2 = st.columns(2)
        with c1:
            finding_mission_code = st.selectbox("Mission concernée", all_codes, key="v7_finding_mission_code")
            finding_process = st.text_input("Processus", placeholder="Achats", key="v7_finding_process")
            finding_clause = st.text_input("Clause ISO", placeholder="8.4", key="v7_finding_clause")
            finding_clause_title = st.text_input("Titre de la clause", placeholder="Optionnel", key="v7_finding_clause_title")
        with c2:
            finding_type = st.selectbox("Type de constat", ["Non-conformité", "Remarque", "Amélioration"], key="v7_finding_type")
            finding_severity = st.selectbox("Gravité", ["Faible", "Moyenne", "Élevée"], key="v7_finding_severity")
            finding_status = st.selectbox("Statut", ["Ouverte", "En cours", "Clôturée", "Non applicable"], key="v7_finding_status")
            finding_date = st.date_input("Date du constat", key="v7_finding_date")

        mlc1, mlc2 = st.columns(2)
        with mlc1:
            finding_priority_label = st.selectbox(
                "Priorité experte pour l’apprentissage",
                ["Faible", "Moyenne", "Haute"],
                index=1,
                key="v8_finding_priority_label",
                help="Cette valeur est la cible de la classification supervisée.",
            )
        with mlc2:
            finding_criticality_score = st.slider(
                "Score de criticité expert (0 à 100)",
                min_value=0,
                max_value=100,
                value=60,
                key="v8_finding_criticality_score",
                help="Cette valeur est la cible de la régression supervisée.",
            )

        finding_description = st.text_area(
            "Description du constat",
            placeholder="Décrire le constat relevé pendant l’audit historique...",
            key="v7_finding_description",
        )

        if st.button(
            "Enregistrer le constat dans PostgreSQL",
            key="v7_save_finding",
            disabled=not is_admin,
        ):
            try:
                finding_id = insert_finding_from_app(
                    host, port, dbname, user, password,
                    {
                        "mission_code": finding_mission_code,
                        "process_name": finding_process,
                        "clause_code": finding_clause,
                        "clause_title": finding_clause_title,
                        "finding_type": finding_type,
                        "description": finding_description,
                        "severity": finding_severity,
                        "status": finding_status,
                        "finding_date": finding_date,
                        "priority_label": finding_priority_label,
                        "criticality_score": finding_criticality_score,
                        "label_source": "expert_interface",
                    },
                )
                st.success(f"Constat enregistré avec succès. ID interne : {finding_id}")
                st.cache_data.clear()
                st.info("Recharge les missions puis régénère une check-list pour intégrer ce nouveau constat.")
            except Exception as e:
                st.error("Impossible d’enregistrer le constat.")
                show_technical_error(e)

    with data_tabs[2]:
        st.markdown("### Importer plusieurs constats")
        st.caption("Colonnes acceptées : mission_code, process_name, clause_code, clause_title, finding_type, description, severity, status, finding_date, priority_label et criticality_score. Les deux dernières colonnes alimentent directement les modèles supervisés.")

        uploaded_findings = st.file_uploader("Fichier Excel ou CSV des constats", type=["xlsx", "csv"], key="v7_upload_findings")
        if uploaded_findings is not None:
            try:
                if uploaded_findings.name.lower().endswith(".csv"):
                    import_df = pd.read_csv(uploaded_findings)
                else:
                    import_df = pd.read_excel(uploaded_findings)
                import_df = normalize_import_columns(import_df)
                st.success(f"Fichier lu : {len(import_df)} ligne(s).")
                st.dataframe(clean_df(import_df), use_container_width=True, hide_index=True)

                required_cols = ["mission_code", "process_name", "finding_type", "description"]
                missing_cols = [c for c in required_cols if c not in import_df.columns]
                if missing_cols:
                    st.warning("Colonnes obligatoires manquantes : " + ", ".join(missing_cols))
                else:
                    if st.button(
                        "Importer ces constats dans PostgreSQL",
                        key="v7_import_findings",
                        disabled=not is_admin,
                    ):
                        inserted, errors = import_findings_dataframe(host, port, dbname, user, password, import_df)
                        st.success(f"Import terminé : {inserted} constat(s) inséré(s).")
                        if errors:
                            st.warning(f"{len(errors)} ligne(s) non importée(s).")
                            st.code("\n".join(errors[:20]))
                        st.cache_data.clear()
            except Exception as e:
                st.error("Impossible de lire ou importer le fichier.")
                show_technical_error(e)

    with data_tabs[3]:
        st.markdown("### Télécharger un modèle d’import")
        st.caption("Ce modèle évite les erreurs de colonnes lors de l’import des constats.")
        st.download_button(
            "Télécharger le modèle Excel des constats",
            data=template_findings_excel(),
            file_name="modele_import_constats_auditprep.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="v7_template_findings",
        )

st.markdown(
    """
<div class="audit-note">
<b>Lecture projet :</b> La V8.3 clarifie les lots et les décisions IA pour faciliter la validation humaine.
</div>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# 10B. V8.3 - MODULE D'APPRENTISSAGE SUPERVISÉ
# ============================================================

st.markdown('<div id="audit-espace" class="audit-anchor"></div>', unsafe_allow_html=True)
st.markdown('<hr class="audit-separator">', unsafe_allow_html=True)
st.markdown(
    """
<div class="audit-workflow-head">
    <span class="audit-workflow-num">01</span>
    <h2>Étape 1 — Vérifier que le modèle est prêt</h2>
</div>
    """,
    unsafe_allow_html=True,
)
st.markdown(
    '<p class="audit-caption-white">Les constats historiques deviennent des exemples étiquetés. '
    'La classification prédit Faible / Moyenne / Haute ; la régression estime un score de criticité entre 0 et 100.</p>',
    unsafe_allow_html=True,
)
st.markdown(
    '<p class="audit-caption-white">Sécurité V8.3 : la gravité et le type de constat sont exclus des entrées du modèle, '
    'car ils servent encore à construire certaines étiquettes proxy.</p>',
    unsafe_allow_html=True,
)

try:
    ml_dataset_df = load_supervised_dataset(host, port, dbname, user, password)
    ml_dataset_error = ""
except Exception as exc:
    ml_dataset_df = pd.DataFrame()
    ml_dataset_error = str(exc)

# V8.3 : l'auditeur peut transformer les anciennes cibles proxy en véritables
# étiquettes supervisées sans passer par PostgreSQL ou par un fichier Excel.
if not ml_dataset_df.empty:
    pending_mask = ~ml_dataset_df["target_source"].map(is_expert_label_source)
    pending_reviews_df = ml_dataset_df[pending_mask].copy()
else:
    pending_reviews_df = pd.DataFrame()

validation_title = (
    f"Valider les constats historiques ({len(pending_reviews_df)} restant(s))"
    if not pending_reviews_df.empty
    else "Validation experte des constats"
)
with st.expander(validation_title, expanded=False):
    if pending_reviews_df.empty:
        st.success("Tous les constats disponibles possèdent déjà une étiquette experte.")
    else:
        st.markdown(
            "Pour chaque constat vérifié, corrige si nécessaire la priorité et le score, "
            "puis coche **Valider par l'expert**. Seules les lignes cochées seront enregistrées."
        )
        reviewer_name = st.text_input(
            "Nom ou rôle du validateur (facultatif)",
            placeholder="Exemple : auditeur qualité",
            key="v83_reviewer_name",
        )
        review_table = pd.DataFrame({
            "finding_id": pending_reviews_df["finding_id"],
            "mission": pending_reviews_df["mission_code"],
            "clause": pending_reviews_df["clause_code"],
            "processus": pending_reviews_df["process_name"],
            "constat": pending_reviews_df["finding_description"].astype(str).str.slice(0, 220),
            "priority_experte": pending_reviews_df["target_priority"],
            "criticite_experte": pending_reviews_df["target_criticality"].round(0).astype(int),
            "commentaire_expert": "",
            "valider_expert": False,
        }).reset_index(drop=True)
        edited_reviews = st.data_editor(
            review_table,
            use_container_width=True,
            hide_index=True,
            disabled=["finding_id", "mission", "clause", "processus", "constat"],
            column_config={
                "finding_id": st.column_config.NumberColumn("ID"),
                "mission": st.column_config.TextColumn("Mission"),
                "clause": st.column_config.TextColumn("Clause"),
                "processus": st.column_config.TextColumn("Processus"),
                "constat": st.column_config.TextColumn("Constat"),
                "priority_experte": st.column_config.SelectboxColumn(
                    "Priorité experte", options=["Faible", "Moyenne", "Haute"], required=True,
                ),
                "criticite_experte": st.column_config.NumberColumn(
                    "Criticité /100", min_value=0, max_value=100, step=1, required=True,
                ),
                "commentaire_expert": st.column_config.TextColumn("Commentaire"),
                "valider_expert": st.column_config.CheckboxColumn("Valider par l'expert"),
            },
            key="v83_expert_review_editor",
        )
        selected_review_count = int(edited_reviews["valider_expert"].fillna(False).sum())
        st.caption(f"{selected_review_count} ligne(s) sélectionnée(s) pour validation.")
        if st.button(
            "Enregistrer les validations expertes",
            type="primary",
            disabled=selected_review_count == 0,
            key="v83_save_expert_reviews",
        ):
            try:
                saved_reviews = save_expert_label_reviews(
                    host, port, dbname, user, password, edited_reviews, reviewer_name,
                )
                st.cache_data.clear()
                st.session_state.pop("ml_pack_v8", None)
                st.session_state.pop("ml_dataset_fingerprint_v8", None)
                st.success(f"{saved_reviews} constat(s) validé(s) par l'expert.")
                st.rerun()
            except Exception as exc:
                st.error("Impossible d'enregistrer les validations expertes.")
                st.code(str(exc))

ml_fingerprint = supervised_dataset_fingerprint(ml_dataset_df)
ml_retrain = st.button(
    "Entraîner / réentraîner les modèles supervisés",
    key="v82_train_models_advanced",
    help="Les métriques sont calculées sur 30 % de données de test.",
)

if (
    ml_retrain
    or "ml_pack_v8" not in st.session_state
    or st.session_state.get("ml_dataset_fingerprint_v8") != ml_fingerprint
):
    with st.spinner("Préparation du dataset et entraînement des modèles..."):
        st.session_state["ml_pack_v8"] = train_supervised_models(ml_dataset_df)
        st.session_state["ml_dataset_fingerprint_v8"] = ml_fingerprint

ml_pack = st.session_state.get("ml_pack_v8", train_supervised_models(pd.DataFrame()))

if ml_dataset_error:
    st.warning("Le dataset supervisé n'a pas pu être construit. Le reste de l'application reste opérationnel.")
    st.code(ml_dataset_error)

dataset_rows = int(len(ml_dataset_df))
model_dataset_rows = int(ml_pack.get("dataset_rows", dataset_rows))
expert_rows = 0
proxy_rows = 0
if not ml_dataset_df.empty and "target_source" in ml_dataset_df.columns:
    expert_mask = ml_dataset_df["target_source"].map(is_expert_label_source)
    expert_rows = int(expert_mask.sum())
    proxy_rows = int((~expert_mask).sum())

mc1, mc2, mc3, mc4 = st.columns(4)
mc1.metric("Constats étiquetés", dataset_rows)
mc2.metric("Validations expertes", expert_rows)
mc3.metric("Étiquettes proxy", proxy_rows)
mc4.metric("État du modèle", "Prêt" if ml_pack.get("status") == "trained" else "En attente")

validation_ratio = expert_rows / max(dataset_rows, 1)
st.progress(min(max(validation_ratio, 0.0), 1.0))
if expert_rows < 10:
    st.caption(
        f"Progression de la validation : {expert_rows}/{dataset_rows}. "
        f"Encore {max(10 - expert_rows, 0)} validation(s) experte(s) minimum avant d'étudier une bascule sans proxy."
    )
else:
    st.markdown(
        f'<p class="audit-caption-white">Progression de la validation : {expert_rows}/{dataset_rows}. '
        "La bascule vers les seules données expertes dépend aussi de la présence d'au moins deux classes équilibrées.</p>",
        unsafe_allow_html=True,
    )

if ml_pack.get("status") == "dependency_missing":
    st.error(ml_pack.get("message", "scikit-learn est manquant."))
    st.code("conda run -n auditprep pip install scikit-learn")
elif ml_pack.get("status") != "trained":
    st.warning(ml_pack.get("message", "Données insuffisantes."))
    st.info(
        "Ajoute ou importe des constats avec une priorité experte et un score de criticité. "
        "La V8.3 attend au moins 10 lignes, deux classes et deux exemples par classe. "
        "Tant que ce seuil n'est pas atteint, la V7.10 continue avec ses règles métier."
    )
else:
    train_rows = int(ml_pack.get("train_rows", 0))
    test_rows = int(ml_pack.get("test_rows", 0))
    total_split = max(train_rows + test_rows, 1)
    train_pct = 100.0 * train_rows / total_split
    test_pct = 100.0 * test_rows / total_split
    class_metrics = ml_pack.get("classification_metrics", {})
    reg_metrics = ml_pack.get("regression_metrics", {})

    st.success(
        f"Modèles entraînés : {train_rows} lignes d'entraînement ({train_pct:.1f} %) et "
        f"{test_rows} lignes de test ({test_pct:.1f} %)."
    )
    st.caption(
        f"Périmètre utilisé : {ml_pack.get('training_scope', 'non précisé')} "
        f"— {model_dataset_rows} ligne(s) retenue(s) sur {dataset_rows}."
    )
    if dataset_rows < 30:
        st.warning(
            "Le pipeline est fonctionnel, mais les métriques restent exploratoires avec moins de 30 constats. "
            "Elles ne doivent pas être présentées comme une performance généralisable."
        )
    if proxy_rows:
        st.info(
            f"{proxy_rows} ligne(s) utilisent encore une cible proxy dérivée de la gravité/type. "
            "Ouvre la zone de validation experte ci-dessus pour les remplacer progressivement."
        )

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Accuracy classification", f"{100 * class_metrics.get('accuracy', 0):.1f} %")
    k2.metric("MAE régression", f"{reg_metrics.get('mae', 0):.2f}")
    k3.metric("RMSE régression", f"{reg_metrics.get('rmse', 0):.2f}")
    r2_value = reg_metrics.get("r2", float("nan"))
    k4.metric("R² régression", "n/a" if pd.isna(r2_value) else f"{r2_value:.3f}")

    ml_tab_class, ml_tab_reg, ml_tab_data = st.tabs(
        ["Classification", "Régression", "Données d'apprentissage"]
    )
    with ml_tab_class:
        left, right = st.columns(2)
        with left:
            st.markdown("### Matrice de confusion")
            st.dataframe(ml_pack["confusion_matrix"], use_container_width=True)
            st.markdown("### Rapport par classe")
            st.dataframe(clean_df(ml_pack["classification_report"]), use_container_width=True, hide_index=True)
        with right:
            st.markdown('<h3 class="audit-requested-black-heading">Variables textuelles les plus influentes</h3>', unsafe_allow_html=True)
            class_imp = ml_pack["classification_importance"].head(15)
            st.dataframe(clean_df(class_imp), use_container_width=True, hide_index=True)
            if not class_imp.empty:
                st.bar_chart(class_imp.set_index("variable")["importance"])

    with ml_tab_reg:
        st.markdown(
            "Le Random Forest Regressor estime la criticité continue. MAE et RMSE mesurent "
            "l'erreur en points sur 100 ; R² compare le modèle à une prédiction moyenne."
        )
        reg_imp = ml_pack["regression_importance"].head(15)
        st.dataframe(clean_df(reg_imp), use_container_width=True, hide_index=True)
        if not reg_imp.empty:
            st.bar_chart(reg_imp.set_index("variable")["importance"])

    with ml_tab_data:
        visible_cols = [
            c for c in [
                "finding_id", "mission_code", "process_name", "clause_code", "finding_type",
                "severity_raw", "target_priority", "target_criticality", "target_source",
                "validated_by", "validated_at", "validation_comment",
            ]
            if c in ml_dataset_df.columns
        ]
        st.dataframe(clean_df(ml_dataset_df[visible_cols]), use_container_width=True, hide_index=True, height=360)
        st.download_button(
            "Télécharger le dataset supervisé en CSV",
            data=clean_df(ml_dataset_df).to_csv(index=False).encode("utf-8-sig"),
            file_name="auditprep_dataset_supervise_v8_3.csv",
            mime="text/csv",
            key="v82_download_dataset",
        )


# ============================================================
# 11. PARAMÉTRAGE DE GÉNÉRATION INTELLIGENTE
# ============================================================

last_run_text = "Aucun lot généré pour le moment."
if not runs_df.empty:
    r = runs_df.iloc[0]
    gen_date = r.get("generated_at", "")
    try:
        gen_date = pd.to_datetime(gen_date).strftime("%d/%m/%Y %H:%M")
    except Exception:
        gen_date = str(gen_date)
    last_run_text = (
        f"{r.get('generation_batch_code', '')} · "
        f"Cible : {r.get('target_mission_code', '')} · "
        f"Historique : {r.get('source_mission_code', '')} · "
        f"Questions : {r.get('checklist_items_count', '')} · "
        f"Généré le : {gen_date}"
    )

st.markdown(
    f"""
<div class="audit-card audit-requested-black">
<b>Mode démonstration opérationnel</b><br>
L’auditeur choisit d’abord la mission à préparer, précise le contexte métier, puis sélectionne un mode :
historique manuel, historique recommandé ou premier audit sans historique.
</div>
<div class="audit-card audit-requested-black">
<b>Dernier lot généré :</b> {last_run_text}
</div>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div id="audit-preparation" class="audit-anchor"></div>', unsafe_allow_html=True)
st.markdown(
    """
<div class="audit-workflow-head">
    <span class="audit-workflow-num">02</span>
    <h2>Étape 2 — Choisir la mission et le mode</h2>
</div>
    """,
    unsafe_allow_html=True,
)
st.markdown(
    '<p class="audit-caption-white">Le moteur peut travailler avec un historique précis, proposer un historique pertinent, '
    'ou générer une première check-list sans historique.</p>',
    unsafe_allow_html=True,
)

target_labels = [make_option_label(row) for _, row in targets_df.iterrows()]
source_labels = [make_option_label(row) for _, row in sources_df.iterrows()] if not sources_df.empty else []

col_a, col_b = st.columns(2)
with col_a:
    target_label = st.selectbox("Mission cible à préparer", target_labels, index=0)
    target_row = targets_df.iloc[target_labels.index(target_label)]

with col_b:
    generation_mode = st.radio(
        "Mode de génération",
        [
            "Avec historique manuel",
            "Historique recommandé",
            "Premier audit / sans historique",
        ],
        index=2 if sources_df.empty else 1,
        horizontal=False,
        help=(
            "Avec historique manuel : l’auditeur choisit la mission source. "
            "Historique recommandé : le moteur propose les historiques les plus proches. "
            "Premier audit : génération depuis le référentiel et le contexte."
        ),
    )

target_code = str(target_row["mission_code"])
target_title = str(target_row.get("mission_title", ""))
saved_context = context_for_mission(mission_contexts_df, target_row.get("mission_id"))

context_container = st.container()
with context_container:
    st.markdown("### Contexte métier de la mission")
    st.markdown(
        '<p class="audit-caption-white">Les valeurs déjà enregistrées sont utilisées automatiquement. '
        'Modifie-les seulement si le contexte a changé.</p>',
        unsafe_allow_html=True,
    )
    ctx1, ctx2 = st.columns(2)
    with ctx1:
        context_sector = st.text_input(
            "Secteur / domaine d’activité",
            value=saved_context.get("sector", ""),
            placeholder="Exemple : textile, école, logistique, formation, industrie...",
            key=f"v82_context_sector_{target_code}",
        )
        context_process = st.text_input(
            "Processus audité",
            value=saved_context.get("audited_process", ""),
            placeholder="Exemple : achats, RH, production, formation, maintenance...",
            key=f"v82_context_process_{target_code}",
        )
        context_requirements = st.text_area(
            "Exigences spécifiques",
            value=saved_context.get("specific_requirements", ""),
            placeholder="Exemple : ISO 9001, HSE, conformité réglementaire, exigences client, Qualiopi...",
            key=f"v82_context_requirements_{target_code}",
            height=90,
        )
    with ctx2:
        context_objective = st.text_area(
            "Objectif de l’audit",
            value=saved_context.get("audit_objective", ""),
            placeholder="Exemple : vérifier la conformité ISO 9001 du processus achats et identifier les risques prioritaires.",
            key=f"v82_context_objective_{target_code}",
            height=90,
        )
        context_scope = st.text_area(
            "Périmètre de l’audit",
            value=saved_context.get("audit_scope", ""),
            placeholder="Exemple : site principal, service achats, fournisseurs critiques, documents et enregistrements associés.",
            key=f"v82_context_scope_{target_code}",
            height=90,
        )
        context_known_risks = st.text_area(
            "Risques connus / points d’attention",
            value=saved_context.get("known_risks", ""),
            placeholder="Exemple : fournisseurs non évalués, documents non à jour, sécurité, conformité HSE...",
            key=f"v82_context_known_risks_{target_code}",
            height=90,
        )
    context_keywords = st.text_input(
        "Mots-clés métier",
        value=saved_context.get("keywords", ""),
        placeholder="Exemple : achats, fournisseur, textile, HSE, école, formation, traçabilité...",
        key=f"v82_context_keywords_{target_code}",
    )

    if st.button("Sauvegarder le contexte", key=f"v82_save_context_{target_code}"):
        try:
            save_context_for_existing_mission(
                host, port, dbname, user, password,
                target_row.get("mission_id"),
                {
                    "sector": context_sector,
                    "audited_process": context_process,
                    "audit_objective": context_objective,
                    "audit_scope": context_scope,
                    "specific_requirements": context_requirements,
                    "known_risks": context_known_risks,
                    "keywords": context_keywords,
                },
            )
            st.success("Contexte métier sauvegardé pour cette mission.")
            st.cache_data.clear()
        except Exception as e:
            st.error("Impossible de sauvegarder le contexte métier.")
            show_technical_error(e)

selected_recommendations_df = pd.DataFrame()
source_row = None
source_code = None
source_label = None

if generation_mode == "Avec historique manuel":
    if sources_df.empty:
        st.warning("Aucun historique exploitable disponible. Utilise le mode Premier audit / sans historique.")
    else:
        source_label = st.selectbox("Mission historique utilisée comme référence", source_labels, index=0)
        source_row = sources_df.iloc[source_labels.index(source_label)]
        source_code = str(source_row["mission_code"])

elif generation_mode == "Historique recommandé":
    if sources_df.empty:
        st.warning("Aucun historique exploitable disponible. Utilise le mode Premier audit / sans historique.")
    else:
        selected_recommendations_df = build_history_recommendations_df(
            sources_df,
            target_row,
            sector=context_sector,
            process=context_process,
            objective=" ".join([context_objective, context_requirements, context_known_risks, context_keywords]),
            scope=context_scope,
            audit_type="",
            standard=target_row.get("standard_name", ""),
        )
        st.markdown("### Historiques recommandés par le moteur")
        st.markdown(
            '<p class="audit-caption-white">Le classement est explicable : il compare le contexte saisi, le client/site, '
            'le référentiel, les constats disponibles et les signaux historiques déjà calculés par le moteur SQL.</p>',
            unsafe_allow_html=True,
        )
        st.dataframe(
            clean_df(selected_recommendations_df.head(5)),
            use_container_width=True,
            hide_index=True,
        )

        rec_labels = [
            f"{r.get('mission_code', '')} — score {r.get('score_pertinence', 0)}/100 | {r.get('mission_title', '')}"
            for _, r in selected_recommendations_df.iterrows()
        ]
        if rec_labels:
            rec_label = st.selectbox("Historique recommandé à utiliser", rec_labels, index=0)
            rec_code = selected_recommendations_df.iloc[rec_labels.index(rec_label)]["mission_code"]
            source_row = sources_df[sources_df["mission_code"].astype(str) == str(rec_code)].iloc[0]
            source_code = str(source_row["mission_code"])
        else:
            st.warning("Aucun historique recommandé disponible.")

else:
    st.markdown(
        """
<div class="audit-info-box audit-first-audit-black"
     style="color:#111111 !important;">
<b style="color:#111111 !important;">Premier audit / sans historique :</b><br>
<span style="color:#111111 !important;">
Aucun constat passé n’est nécessaire. Le moteur génère une check-list initiale à partir du référentiel,
du secteur, du processus audité, de l’objectif et du périmètre saisis.
</span>
</div>
        """,
        unsafe_allow_html=True,
    )

col_a, col_b = st.columns(2)
with col_a:
    st.markdown(
        f"""
<div class="audit-card audit-requested-black">
    <b>Mission cible : {target_code}</b><br>
    {target_row.get('mission_title', '')}<br>
    Client : {target_row.get('client_name', '')} · Site : {target_row.get('site_name', '')} · Audit : {target_row.get('audit_date', '')}
</div>
        """,
        unsafe_allow_html=True,
    )

with col_b:
    if generation_mode == "Premier audit / sans historique" or source_row is None:
        st.markdown(
            """
<div class="audit-card audit-requested-black">
    <b>Source utilisée :</b> Aucun historique<br>
    Génération initiale depuis le référentiel ISO, le contexte métier et les informations saisies.
</div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f"""
<div class="audit-card audit-requested-black">
    <b>Mission historique : {source_code}</b><br>
    {source_row.get('mission_title', '')}<br>
    Constats exploitables : {source_row.get('findings_count', '')} · Audit : {source_row.get('audit_date', '')}
</div>
            """,
            unsafe_allow_html=True,
        )

same_mission = bool(source_code) and target_code == source_code
if same_mission:
    st.markdown(
        """
<div class="audit-warning">
La mission cible et la mission historique sont identiques. Sélectionne une mission historique distincte ou utilise le mode Premier audit.
</div>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    """
<div class="audit-workflow-head">
    <span class="audit-workflow-num">03</span>
    <h2>Étape 3 — Lancer la génération</h2>
</div>
    """,
    unsafe_allow_html=True,
)
st.markdown(
    '<p class="audit-caption-white">Vérifie les deux cartes ci-dessus, puis clique une seule fois sur le bouton.</p>',
    unsafe_allow_html=True,
)
generate_btn = st.button(
    "Générer la check-list et afficher le nouveau résultat",
    disabled=same_mission or (generation_mode != "Premier audit / sans historique" and not source_code),
    type="primary",
    use_container_width=False,
)

if generate_btn:
    try:
        if generation_mode == "Premier audit / sans historique":
            first_checklist_df = build_first_audit_checklist(
                target_code=target_code,
                target_title=target_title,
                sector=context_sector,
                process=context_process,
                objective=" ".join([context_objective, context_requirements, context_known_risks, context_keywords]),
                scope=context_scope,
                audit_type="",
                standard=target_row.get("standard_name", "ISO 9001:2015"),
            )
            first_kpi_df, first_process_df, first_clause_df, first_alerts_df = build_first_audit_datasets(first_checklist_df)
            first_run = make_session_run_from_first_audit(first_checklist_df, target_code, target_title)

            st.session_state["first_audit_pack"] = {
                "run": first_run,
                "checklist_df": first_checklist_df,
                "kpi_df": first_kpi_df,
                "process_df": first_process_df,
                "clause_df": first_clause_df,
                "alerts_df": first_alerts_df,
            }
            st.session_state["preferred_run_batch_v81"] = first_run.get("generation_batch_code")
            st.session_state.pop("v82_run_advanced", None)

            st.markdown(
                f"""
<div class="audit-success">
Check-list initiale générée sans historique : {first_checklist_df.shape[0]} questions.
</div>
                """,
                unsafe_allow_html=True,
            )

        else:
            result_df = execute_generation(host, port, dbname, user, password, target_code, source_code)
            generated_batch = safe_first(result_df, "generation_batch_code", None)
            if generated_batch:
                st.session_state["preferred_run_batch_v81"] = str(generated_batch)
                st.session_state.pop("v82_run_advanced", None)
            st.cache_data.clear()
            st.markdown(
                f"""
<div class="audit-success">
Check-list générée : {safe_first(result_df, 'recommendations_count', 'n/a')} recommandations et {safe_first(result_df, 'checklist_items_count', 'n/a')} questions.
</div>
                """,
                unsafe_allow_html=True,
            )
            runs_df = read_sql_cached(SQL_GENERATION_RUNS, host, port, dbname, user, password)
            if not generated_batch and not runs_df.empty:
                generated_batch = str(runs_df.iloc[0].get("generation_batch_code", ""))
                if generated_batch:
                    st.session_state["preferred_run_batch_v81"] = generated_batch
                    st.session_state.pop("v82_run_advanced", None)

            if generation_mode == "Historique recommandé" and not selected_recommendations_df.empty:
                st.info("Historique recommandé utilisé. La justification du choix reste visible dans le tableau de recommandations ci-dessus.")
            st.info("Le nouveau lot est sélectionné automatiquement dans les résultats.")

    except Exception as e:
        st.error("La génération a échoué.")
        show_technical_error(e)



# ============================================================
# 11B. SÉLECTION DU LOT
# ============================================================

st.markdown('<div id="audit-resultats" class="audit-anchor"></div>', unsafe_allow_html=True)
st.markdown('<hr class="audit-separator">', unsafe_allow_html=True)
st.markdown(
    """
<div class="audit-workflow-head">
    <span class="audit-workflow-num">04</span>
    <h2>Étape 4 — Contrôler le résultat généré</h2>
</div>
    """,
    unsafe_allow_html=True,
)
st.markdown(
    '<p class="audit-caption-white">Le dernier lot généré est sélectionné automatiquement. '
    'Ouvre la liste seulement si tu veux consulter un ancien résultat.</p>',
    unsafe_allow_html=True,
)

session_run = None
if "first_audit_pack" in st.session_state:
    session_run = st.session_state["first_audit_pack"].get("run")

session_runs_df = pd.DataFrame([session_run]) if session_run else pd.DataFrame()
runs_display_df = pd.concat([session_runs_df, runs_df], ignore_index=True, sort=False)

if runs_display_df.empty:
    st.info("Aucun lot de génération n’est encore disponible. Lance une génération.")
    st.stop()

run_labels = [make_run_label(row) for _, row in runs_display_df.iterrows()]
preferred_batch = str(st.session_state.get("preferred_run_batch_v81", ""))
preferred_index = 0
if preferred_batch:
    matches = runs_display_df.index[
        runs_display_df["generation_batch_code"].astype(str).eq(preferred_batch)
    ].tolist()
    if matches:
        preferred_index = int(matches[0])

selected_run_label = st.selectbox(
    "Lot à afficher",
    run_labels,
    index=preferred_index,
    key="v82_run_advanced",
)
selected_run = runs_display_df.iloc[run_labels.index(selected_run_label)]
batch_code = selected_run["generation_batch_code"]
selected_mode_display = resolve_generation_mode(selected_run)
selected_target_display = safe_display_value(selected_run.get("target_mission_code", ""))
selected_source_display = safe_display_value(selected_run.get("source_mission_code", ""), "Aucun historique")
selected_count_display = safe_display_value(selected_run.get("checklist_items_count", ""), "Non calculé")
selected_date_display = safe_display_value(selected_run.get("generated_at", ""))

st.markdown(
    f"""
<div class="audit-card audit-requested-black">
    <span class="audit-badge badge-low">Lot sélectionné</span><br><br>
    <b>{batch_code}</b><br>
    Cible : {selected_target_display} ·
    Historique : {selected_source_display} ·
    Questions : {selected_count_display} ·
    Généré le : {selected_date_display}<br>
    Mode : {selected_mode_display}
</div>
    """,
    unsafe_allow_html=True,
)

contextual_message = ""
ml_application_message = ""
try:
    is_first_audit_session = (
        "first_audit_pack" in st.session_state
        and str(batch_code) == str(st.session_state["first_audit_pack"]["run"].get("generation_batch_code"))
    )

    if is_first_audit_session:
        pack = st.session_state["first_audit_pack"]
        checklist_df = pack["checklist_df"]
        kpi_df = pack["kpi_df"]
        process_df = pack["process_df"]
        clause_df = pack["clause_df"]
        alerts_df = pack["alerts_df"]
        trace_df = build_traceability_from_checklist(checklist_df, selected_run)
        contextual_message = "Mode premier audit : la check-list est générée directement depuis le contexte et le référentiel."

    else:
        source_for_scores = str(selected_run.get("source_mission_code", source_code))

        kpi_df = read_sql_cached(SQL_KPI_BY_PRIORITY, host, port, dbname, user, password, params=(batch_code,))
        checklist_df = read_sql_cached(SQL_CHECKLIST_ITEMS, host, port, dbname, user, password, params=(batch_code,))

        # V7.8 : le lot PostgreSQL reste la base, puis on applique une couche contextuelle
        # liée à la mission cible pour éviter que deux lots avec le même historique soient identiques.
        checklist_df, contextual_message = augment_checklist_with_target_context(
            checklist_df,
            selected_run,
            targets_df,
            mission_contexts_df,
        )
        kpi_df = recompute_kpi_from_checklist(checklist_df)

        # Les vues de vigilance et d'alertes sont filtrées par la mission historique du lot sélectionné.
        process_df = read_sql_cached(SQL_PROCESS_VIGILANCE, host, port, dbname, user, password, params=(source_for_scores,))
        clause_df = read_sql_cached(SQL_CLAUSE_VIGILANCE, host, port, dbname, user, password, params=(source_for_scores,))
        alerts_df = read_sql_cached(SQL_TOP_ALERTS, host, port, dbname, user, password, params=(source_for_scores,))
        process_df, clause_df = add_context_rows_to_vigilance(process_df, clause_df, selected_run, targets_df)

        trace_df = build_traceability_from_checklist(checklist_df, selected_run)

    # V8.3 : classification, régression et règles sont confrontées. La règle
    # V7.10 reste la barrière de sécurité lorsque le signal ML est faible.
    checklist_df, ml_application_message = apply_supervised_predictions(checklist_df, ml_pack)
    kpi_df = recompute_kpi_from_checklist(checklist_df)
    trace_df = build_traceability_from_checklist(checklist_df, selected_run)

except Exception as e:
    st.error("Impossible de charger les données du lot sélectionné.")
    show_technical_error(e)
    st.stop()

mode_labels = generation_mode_labels(selected_run)
if ml_pack.get("status") == "trained":
    mode_labels["result_sentence"] += " La V8.3 confronte ensuite la classification, la criticité prédite et les règles métier."
    mode_labels["exec_note"] += " Une prédiction faible ou incohérente ne remplace pas automatiquement la priorité métier."
    mode_labels["trace_note"] += " La V8.3 conserve la règle d'origine, les deux prédictions, la décision de fusion et l'indicateur de validation humaine."
    mode_labels["export_note"] += " Les métriques, variables importantes, prédictions et données d'apprentissage sont également exportées."
selected_run_export_df = build_selected_run_export_df(selected_run, checklist_df, kpi_df, ml_pack)

if contextual_message:
    st.markdown(
        f"""
<div class="audit-info-box audit-requested-black">
<b>Lecture du lot sélectionné :</b> {contextual_message}
</div>
        """,
        unsafe_allow_html=True,
    )

if ml_application_message:
    st.markdown(
        f"""
<div class="audit-info-box audit-requested-black">
<b>Contrôle supervisé V8.3 :</b> {ml_application_message}
</div>
        """,
        unsafe_allow_html=True,
    )

# ============================================================
# 12. FICHE MISSION
# ============================================================

mission_display_title = safe_first(
    pd.DataFrame([selected_run]),
    "target_mission_title",
    "Mission cible",
)

st.markdown(
    f'<div class="audit-section-head audit-section-head--summary">'
    f'<h2>{html.escape(str(mission_display_title))}</h2>'
    f'</div>',
    unsafe_allow_html=True,
)

st.markdown(
    f"""
<div class="audit-card audit-requested-black">
    <b>Code mission :</b> {selected_run.get('target_mission_code', '')}
    &nbsp; | &nbsp; <b>{mode_labels['source_caption']} :</b> {mode_labels['source_value']}<br>
    {mode_labels['result_sentence']}
</div>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# 13. ONGLETS
# ============================================================

tab_exec, tab_checklist, tab_vigilance, tab_trace = st.tabs(
    ["1 · Synthèse", "2 · Check-list et validation", "3 · Vigilance", "4 · Traçabilité"]
)


# ---------------------- TAB 1 : VUE EXÉCUTIVE ----------------------
with tab_exec:
    st.markdown(
        '<div class="audit-section-head audit-section-head--summary"><h2>Synthèse de la génération</h2></div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p class="audit-caption-white">Vue globale du lot sélectionné et des priorités de préparation produites automatiquement.</p>',
        unsafe_allow_html=True,
    )

    total_questions = int(checklist_df.shape[0]) if checklist_df is not None else 0

    def count_priority(name):
        if kpi_df.empty or "generated_priority" not in kpi_df.columns:
            return 0
        row = kpi_df[kpi_df["generated_priority"].astype(str).str.lower() == name.lower()]
        if row.empty:
            return 0
        return int(row.iloc[0].get("questions_count", 0))

    high_count = count_priority("Haute")
    medium_count = count_priority("Moyenne")
    low_count = count_priority("Faible")
    ml_review_count = int(checklist_df.get("ml_review_required", pd.Series(dtype=bool)).fillna(False).astype(bool).sum())

    st.markdown(
        f"""
<div class="audit-grid-4">
    <div class="audit-mini-card">
        <div class="audit-kpi-label">Questions générées</div>
        <div class="audit-kpi-value">{total_questions}</div>
    </div>
    <div class="audit-mini-card">
        <div class="audit-kpi-label">Priorité haute</div>
        <div class="audit-kpi-value">{high_count}</div>
    </div>
    <div class="audit-mini-card">
        <div class="audit-kpi-label">Priorité moyenne</div>
        <div class="audit-kpi-value">{medium_count}</div>
    </div>
    <div class="audit-mini-card">
        <div class="audit-kpi-label">Priorité faible</div>
        <div class="audit-kpi-value">{low_count}</div>
    </div>
</div>
        """,
        unsafe_allow_html=True,
    )
    if ml_pack.get("status") == "trained":
        st.caption(
            f"Priorités finales issues de la fusion prudente entre règles et ML V8.3. "
            f"{ml_review_count} question(s) nécessitent une confirmation humaine (faible confiance ou divergence avec la règle métier)."
        )

    st.markdown(
        f'<div class="audit-section-head audit-section-head--alerts"><h2>{html.escape(str(mode_labels["alerts_title"]))}</h2></div>',
        unsafe_allow_html=True,
    )

    if alerts_df.empty:
        st.info("Aucune alerte prioritaire disponible.")
    else:
        left, right = st.columns([2, 1])
        with left:
            for _, row in alerts_df.head(5).iterrows():
                label = row.get("alert_label", row.get("process_name", row.get("clause_title", "")))
                score = row.get("capped_score", row.get("score", ""))
                level = row.get("vigilance_level", row.get("level", ""))
                st.markdown(
                    f"""
<div class="audit-alert-card">
    <b>{label}</b> {priority_badge(level)}<br>
    Score de vigilance : <b>{score}/100</b>
</div>
                    """,
                    unsafe_allow_html=True,
                )

        with right:
            top_process = safe_first(process_df, "process_name", "Non disponible")
            top_process_score = safe_first(process_df, "capped_score", "")
            top_clause = safe_first(clause_df, "clause_code", "Non disponible")
            top_clause_score = safe_first(clause_df, "capped_score", "")

            st.markdown(
                '<div class="audit-section-head audit-section-head--decisions"><h2>Décisions immédiates</h2></div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                f"""
<div class="audit-mini-card">
    <div class="audit-kpi-label">Processus le plus sensible</div>
    <div class="audit-kpi-value">{top_process}</div>
    <span class="audit-badge badge-low">↑ {top_process_score}/100</span>
</div>
<div class="audit-mini-card">
    <div class="audit-kpi-label">Clause ISO la plus sensible</div>
    <div class="audit-kpi-value">{top_clause}</div>
    <span class="audit-badge badge-low">↑ {top_clause_score}/100</span>
</div>
                """,
                unsafe_allow_html=True,
            )

    st.markdown(
        f"""
<div class="audit-note">
<b>Lecture métier :</b> {mode_labels['exec_note']}
</div>
        """,
        unsafe_allow_html=True,
    )


# ---------------------- TAB 2 : VIGILANCE ----------------------
with tab_vigilance:
    st.markdown(f"## {mode_labels['vigilance_title']}")
    st.caption(mode_labels["vigilance_caption"])

    c1, c2 = st.columns(2)

    with c1:
        st.markdown("### Par processus")
        show_cols = [c for c in ["process_name", "findings_count", "nonconformities_count", "remarks_count", "improvements_count", "capped_score", "vigilance_level"] if c in process_df.columns]
        st.dataframe(clean_df(process_df[show_cols]), use_container_width=True, hide_index=True)
        if "process_name" in process_df.columns and "capped_score" in process_df.columns:
            chart_df = process_df[["process_name", "capped_score"]].rename(columns={"process_name": "Processus", "capped_score": "Score"}).set_index("Processus")
            st.bar_chart(chart_df)

    with c2:
        st.markdown("### Par clause ISO")
        show_cols = [c for c in ["clause_code", "clause_title", "findings_count", "capped_score", "vigilance_level"] if c in clause_df.columns]
        st.dataframe(clean_df(clause_df[show_cols]), use_container_width=True, hide_index=True)
        if "clause_code" in clause_df.columns and "capped_score" in clause_df.columns:
            chart_df = clause_df[["clause_code", "capped_score"]].rename(columns={"clause_code": "Clause", "capped_score": "Score"}).set_index("Clause")
            st.bar_chart(chart_df)

    with st.expander("Voir les explications détaillées des scores", expanded=False):
        st.markdown("### Processus")
        process_cols = [c for c in ["process_name", "capped_score", "vigilance_level", "explanation_summary"] if c in process_df.columns]
        st.dataframe(clean_df(process_df[process_cols]), use_container_width=True, hide_index=True)

        st.markdown("### Clauses ISO")
        clause_cols = [c for c in ["clause_code", "clause_title", "capped_score", "vigilance_level", "explanation_summary"] if c in clause_df.columns]
        st.dataframe(clean_df(clause_df[clause_cols]), use_container_width=True, hide_index=True)


# ---------------------- TAB 3 : CHECK-LIST ----------------------
with tab_checklist:
    st.markdown("## Check-list priorisée produite par le moteur")
    st.caption("Filtre, recherche et export du livrable de préparation d’audit.")

    if checklist_df.empty:
        st.info("Aucun item de check-list pour ce lot.")
    else:
        priorities = sorted(checklist_df["generated_priority"].dropna().unique().tolist()) if "generated_priority" in checklist_df.columns else []
        c1, c2 = st.columns([1.2, 1])

        with c1:
            selected_priorities = st.multiselect("Filtrer par priorité", priorities, default=priorities)
        with c2:
            search = st.text_input("Rechercher dans la check-list", placeholder="Ex. compétences, non-conformité, procédures...")

        filtered = checklist_df.copy()

        if selected_priorities and "generated_priority" in filtered.columns:
            filtered = filtered[filtered["generated_priority"].isin(selected_priorities)]

        if search:
            mask = pd.Series(False, index=filtered.index)
            for col in filtered.columns:
                mask = mask | filtered[col].astype(str).str.contains(search, case=False, na=False)
            filtered = filtered[mask]

        preferred_cols = [
            "display_order",
            "clause_code",
            "clause_title",
            "theme",
            "question_text",
            "rule_based_priority",
            "ml_predicted_priority",
            "ml_prediction_confidence",
            "ml_predicted_criticality",
            "ml_criticality_band",
            "generated_priority",
            "priority_origin",
            "ml_review_required",
            "ml_decision",
            "recommendation_label",
            "conformity_status",
        ]
        show_cols = [c for c in preferred_cols if c in filtered.columns]
        display_labels = {
            "display_order": "N°",
            "clause_code": "Clause",
            "clause_title": "Exigence ISO",
            "theme": "Thème",
            "question_text": "Question d’audit",
            "rule_based_priority": "Priorité des règles",
            "ml_predicted_priority": "Priorité proposée par l’IA",
            "ml_prediction_confidence": "Confiance IA (%)",
            "ml_predicted_criticality": "Criticité IA (/100)",
            "ml_criticality_band": "Classe selon criticité",
            "generated_priority": "Priorité finale",
            "priority_origin": "Origine de la décision",
            "ml_review_required": "À confirmer",
            "ml_decision": "Explication simple",
            "recommendation_label": "Recommandation",
            "conformity_status": "Conformité",
        }
        display_df = clean_df(filtered[show_cols]).rename(columns=display_labels)
        st.dataframe(display_df, use_container_width=True, hide_index=True, height=520)

        st.download_button(
            "Télécharger la check-list filtrée en CSV",
            data=clean_df(filtered).to_csv(index=False).encode("utf-8-sig"),
            file_name=f"checklist_{batch_code}.csv",
            mime="text/csv",
        )

        with st.expander("Voir les justifications détaillées", expanded=False):
            just_cols = [c for c in [
                "display_order", "clause_code", "question_text", "rule_based_priority",
                "ml_predicted_priority", "ml_prediction_confidence", "ml_predicted_criticality",
                "ml_criticality_band", "generated_priority", "priority_origin", "ml_review_required",
                "ml_decision", "expected_evidence",
            ] if c in checklist_df.columns]
            st.dataframe(
                clean_df(checklist_df[just_cols]).rename(columns=display_labels),
                use_container_width=True,
                hide_index=True,
            )


# ---------------------- TAB 4 : TRAÇABILITÉ ----------------------
with tab_trace:
    st.markdown("## Traçabilité du raisonnement de priorisation")
    st.caption("Lien entre le lot généré, la question, la priorité et les éléments de justification disponibles.")

    if trace_df.empty:
        st.info("Aucune trace disponible pour ce lot.")
    else:
        st.dataframe(clean_df(trace_df), use_container_width=True, hide_index=True, height=620)

    st.markdown(
        f"""
<div class="audit-note">
<b>Principe :</b> {mode_labels['trace_note']}
</div>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# 14. EXPORT GLOBAL
# ============================================================

st.markdown('<hr class="audit-separator">', unsafe_allow_html=True)
st.markdown(
    """
<div class="audit-workflow-head">
    <span class="audit-workflow-num">05</span>
    <h2>5. Exports du lot sélectionné</h2>
</div>
    """,
    unsafe_allow_html=True,
)
st.caption("Téléchargement des données consolidées pour l’entreprise, le rapport PFE ou une démonstration.")

export = excel_bytes(
    {
        "generation_run_selectionne": selected_run_export_df,
        "kpi_priorites": kpi_df,
        "checklist": checklist_df,
        "traceabilite": trace_df,
        "vigilance_processus": process_df,
        "vigilance_clauses": clause_df,
        "alertes": alerts_df,
        "metriques_ia": ml_metrics_export_df(ml_pack),
        "matrice_confusion": ml_pack.get("confusion_matrix", pd.DataFrame()).reset_index(),
        "importance_classification": ml_pack.get("classification_importance", pd.DataFrame()),
        "importance_regression": ml_pack.get("regression_importance", pd.DataFrame()),
        "dataset_apprentissage": ml_dataset_df,
    }
)

c1, c2 = st.columns([1, 2])
with c1:
    st.download_button(
        "Télécharger l’export Excel complet",
        data=export,
        file_name=f"auditprep_export_{batch_code}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
with c2:
    st.markdown(
        f"""
<div class="audit-note">
{mode_labels['export_note']}
</div>
        """,
        unsafe_allow_html=True,
    )

st.caption("AuditPrep IA – Version finale entreprise | Streamlit + HTML/CSS | Authentification sécurisée | Classification + régression supervisées | PostgreSQL + supervision humaine")
